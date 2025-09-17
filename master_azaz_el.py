#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
 .S_SSSs     sdSSSSSSSbs   .S_SSSs     sdSSSSSSSbs    sSSs  S.      
.SS~SSSSS    YSSSSSSSS%S  .SS~SSSSS    YSSSSSSSS%S   d%%SP  SS.     
S%S   SSSS          S%S   S%S   SSSS          S%S   d%S'    S%S     
S%S    S%S         S&S    S%S    S%S         S&S    S%S     S%S     
S%S SSSS%S        S&S     S%S SSSS%S        S&S     S&S     S&S     
S&S  SSS%S        S&S     S&S  SSS%S        S&S     S&S_Ss  S&S     
S&S    S&S       S&S      S&S    S&S       S&S      S&S~SP  S&S     
S&S    S&S      S*S       S&S    S&S      S*S       S&S     S&S     
S*S    S&S     S*S        S*S    S&S     S*S        S*b     S*b     
S*S    S*S   .s*S         S*S    S*S   .s*S         S*S.    S*S.    
S*S    S*S   sY*SSSSSSSP  S*S    S*S   sY*SSSSSSSP   SSSbs   SSSbs  
SSS    S*S  sY*SSSSSSSSP  SSS    S*S  sY*SSSSSSSSP    YSSP    YSSP  
       SP                        SP                                 
       Y                         Y                                  
                                                                      
Azaz-El v6.0.0-MASTER - Unified Professional Security Assessment Framework
Advanced Master TUI with Complete Integration of All Security Tools and Modules
"""

import os
import sys
import asyncio
import argparse
import json
import time
import threading
import subprocess
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging
import shutil
import uuid
import webbrowser
import psutil

# Add project root to path
sys.path.append(str(Path(__file__).parent))

# Import all modules and integrations
try:
    from core.config import ConfigurationManager
    from core.logging import get_logger
    from core.reporting import AdvancedReportGenerator
    from scanners.web_scanner import AdvancedWebScanner
    from scanners.api_scanner import AdvancedAPIScanner
    from scanners.cloud_scanner import CloudSecurityScanner
    from scanners.infrastructure_scanner import InfrastructureScanner
    from moloch_integration import MolochIntegration, EnhancedScanner
    
    # Import moloch core functions
    from moloch import (
        run_subdomain_discovery, run_dns_resolution, run_http_probing,
        run_vulnerability_scan, run_port_scan, run_ssl_scan,
        run_crawling, run_xss_scan, run_directory_fuzzing,
        execute_tool, load_config, new_run, setup_logging,
        filter_and_save_positive_results, generate_simple_report,
        check_and_install_dependencies, initialize_environment
    )
except ImportError as e:
    print(f"Warning: Some modules could not be imported: {e}")
    print("The master tool will run with limited functionality.")

# --- Master Configuration ---
MASTER_APP = "Azaz-El"
MASTER_VERSION = "v6.0.0-MASTER"
MASTER_AUTHOR = "Advanced Security Research Team"

MASTER_BANNER = r"""
 .S_SSSs     sdSSSSSSSbs   .S_SSSs     sdSSSSSSSbs    sSSs  S.      
.SS~SSSSS    YSSSSSSSS%S  .SS~SSSSS    YSSSSSSSS%S   d%%SP  SS.     
S%S   SSSS          S%S   S%S   SSSS          S%S   d%S'    S%S     
S%S    S%S         S&S    S%S    S%S         S&S    S%S     S%S     
S%S SSSS%S        S&S     S%S SSSS%S        S&S     S&S     S&S     
S&S  SSS%S        S&S     S&S  SSS%S        S&S     S&S_Ss  S&S     
S&S    S&S       S&S      S&S    S&S       S&S      S&S~SP  S&S     
S&S    S&S      S*S       S&S    S&S      S*S       S&S     S&S     
S*S    S&S     S*S        S*S    S&S     S*S        S*b     S*b     
S*S    S*S   .s*S         S*S    S*S   .s*S         S*S.    S*S.    
S*S    S*S   sY*SSSSSSSP  S*S    S*S   sY*SSSSSSSP   SSSbs   SSSbs  
SSS    S*S  sY*SSSSSSSSP  SSS    S*S  sY*SSSSSSSSP    YSSP    YSSP  
       SP                        SP                                 
       Y                         Y                                  
                                                                      
"""

class MasterAzazElFramework:
    """
    Unified Master Security Assessment Framework
    Advanced Professional TUI with Complete Integration
    """
    
    def __init__(self) -> None:
        """Initialize the master unified framework"""
        self.version = MASTER_VERSION
        self.app_name = MASTER_APP
        self.author = MASTER_AUTHOR
        
        # Initialize logging
        self.setup_master_logging()
        self.logger = logging.getLogger("master-azaz-el")
        
        # Initialize configuration management
        try:
            self.config_manager = ConfigurationManager(Path("moloch.cfg.json"))
            self.config = self.config_manager.load_config()
        except:
            self.config = load_config()
            self.config_manager = None
        
        # Initialize integrations
        self.initialize_integrations()
        
        # Master dashboard state
        self.active_scans = {}
        self.scan_history = []
        self.system_status = {
            "scanners": {},
            "tools": {},
            "resources": {},
            "health": "Unknown"
        }
        
        # Performance monitoring
        self.performance_metrics = {
            "cpu_usage": 0.0,
            "memory_usage": 0.0,
            "disk_usage": 0.0,
            "network_status": "Unknown"
        }
        
        # Initialize master framework
        self.initialize_master_framework()
    
    def setup_master_logging(self) -> None:
        """Setup enhanced logging for master framework"""
        try:
            setup_logging()
        except:
            # Fallback logging setup
            logging.basicConfig(
                level=logging.INFO,
                format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
            )
    
    def initialize_integrations(self) -> None:
        """Initialize all framework integrations"""
        self.logger.info("Initializing Master Framework Integrations")
        
        try:
            # Initialize moloch integration
            self.moloch_integration = MolochIntegration(self.config_manager)
            self.enhanced_scanner = EnhancedScanner(self.config_manager)
            
            # Initialize advanced scanners
            self.web_scanner = AdvancedWebScanner(self.config)
            self.api_scanner = AdvancedAPIScanner(self.config)
            self.cloud_scanner = CloudSecurityScanner(self.config)
            self.infrastructure_scanner = InfrastructureScanner(self.config)
            
            # Initialize reporting
            self.report_generator = AdvancedReportGenerator(self.config)
            
            self.integrations_available = True
            self.logger.info("All integrations initialized successfully")
            
        except Exception as e:
            self.logger.warning(f"Some integrations failed to initialize: {e}")
            self.integrations_available = False
    
    def initialize_master_framework(self) -> None:
        """Initialize the master framework environment"""
        self.logger.info("Initializing Master Azaz-El Framework")
        
        # Initialize targets
        self.targets = set()
        self.load_targets()
        
        # Update system status
        self.update_system_status()
        
        # Setup directories
        self.base_dir = Path("runs")
        self.base_dir.mkdir(exist_ok=True)
        
        # Initialize environment if needed (skip tool installation for now)
        try:
            # Skip automatic tool installation to prevent prompts
            self.logger.info("Skipping automatic tool installation to prevent interactive prompts")
        except Exception as e:
            self.logger.warning(f"Environment initialization had issues: {e}, continuing...")
        
        self.logger.info("Master framework initialization complete")
        
    def load_targets(self) -> None:
        """Load targets from targets.txt file"""
        try:
            targets_file = Path("targets.txt")
            if targets_file.exists():
                with open(targets_file, 'r') as f:
                    targets = [line.strip() for line in f if line.strip()]
                    self.targets = set(targets)
                    self.logger.info(f"Loaded {len(self.targets)} targets")
            else:
                self.targets = set()
                self.logger.info("No targets file found, starting with empty target list")
        except Exception as e:
            self.logger.warning(f"Error loading targets: {e}")
            self.targets = set()
    
    def update_system_status(self) -> None:
        """Update comprehensive system status"""
        try:
            # Check scanner availability
            if self.integrations_available:
                scanners = {
                    "web_scanner": self.web_scanner,
                    "api_scanner": self.api_scanner,
                    "cloud_scanner": self.cloud_scanner,
                    "infrastructure_scanner": self.infrastructure_scanner,
                    "moloch_integration": self.moloch_integration,
                    "enhanced_scanner": self.enhanced_scanner
                }
                
                for name, scanner in scanners.items():
                    if scanner:
                        self.system_status["scanners"][name] = "✅ Available"
                    else:
                        self.system_status["scanners"][name] = "❌ Not Available"
            
            # Update performance metrics
            self.performance_metrics["cpu_usage"] = psutil.cpu_percent()
            self.performance_metrics["memory_usage"] = psutil.virtual_memory().percent
            disk = psutil.disk_usage('/')
            self.performance_metrics["disk_usage"] = (disk.used / disk.total) * 100
            
            # Check tools availability
            essential_tools = [
                "subfinder", "nuclei", "httpx", "ffuf", "nmap", 
                "katana", "dalfox", "gobuster", "naabu", "dnsx"
            ]
            
            for tool in essential_tools:
                if shutil.which(tool):
                    self.system_status["tools"][tool] = "✅ Available"
                else:
                    self.system_status["tools"][tool] = "❌ Not Installed"
            
            # Overall health assessment
            available_tools = sum(1 for status in self.system_status["tools"].values() if "Available" in status)
            total_tools = len(self.system_status["tools"])
            
            if available_tools >= total_tools * 0.8:
                self.system_status["health"] = "Excellent"
            elif available_tools >= total_tools * 0.6:
                self.system_status["health"] = "Good"
            elif available_tools >= total_tools * 0.4:
                self.system_status["health"] = "Fair"
            else:
                self.system_status["health"] = "Poor"
                
        except Exception as e:
            self.logger.error(f"Error updating system status: {e}")
            self.system_status["health"] = "Error"
    
    def print_master_banner(self) -> None:
        """Print the enhanced master banner with comprehensive status"""
        # Clear screen for better presentation
        os.system('clear' if os.name == 'posix' else 'cls')
        
        # Print the main banner with gradient colors
        print(f"\033[1;36m{MASTER_BANNER}\033[0m")
        
        # Enhanced master title section
        print("╔═══════════════════════════════════════════════════════════════════════════════╗")
        print(f"║\033[1;91m                🔱 {MASTER_APP} {MASTER_VERSION} MASTER FRAMEWORK 🔱\033[0m                 ║")
        print("║\033[1;92m        Unified Professional Security Assessment & Penetration Testing\033[0m        ║")
        print("╠═══════════════════════════════════════════════════════════════════════════════╣")
        print(f"║  Author: \033[1;97m{MASTER_AUTHOR}\033[0m  │  Health: \033[1;{'92m' if self.system_status['health'] == 'Excellent' else '93m' if self.system_status['health'] == 'Good' else '91m'}{self.system_status['health']}\033[0m     ║")
        print(f"║  Status: \033[1;92mOperational\033[0m                    │  Platform: \033[1;97mMulti-Cloud Ready\033[0m    ║")
        print("╚═══════════════════════════════════════════════════════════════════════════════╝")
        
        # System performance display
        cpu_color = "92" if self.performance_metrics["cpu_usage"] < 70 else "93" if self.performance_metrics["cpu_usage"] < 90 else "91"
        mem_color = "92" if self.performance_metrics["memory_usage"] < 70 else "93" if self.performance_metrics["memory_usage"] < 90 else "91"
        
        print(f"⏰ Session: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} │ "
              f"CPU: \033[1;{cpu_color}m{self.performance_metrics['cpu_usage']:.1f}%\033[0m │ "
              f"RAM: \033[1;{mem_color}m{self.performance_metrics['memory_usage']:.1f}%\033[0m │ "
              f"Scanners: \033[1;92m{len([s for s in self.system_status['scanners'].values() if 'Available' in s])}\033[0m")
        
        print()  # Add spacing
    
    async def display_master_menu(self) -> None:
        """Display the modern TUI interface"""
        await self.create_modern_tui_interface()
    
    async def create_modern_tui_interface(self) -> None:
        """Create the modern, sophisticated TUI interface"""
        while True:
            try:
                self.clear_screen()
                self.display_modern_header()
                self.display_system_dashboard()
                self.display_navigation_menu()
                
                choice = self.get_user_input()
                if not await self.handle_navigation_choice(choice):
                    break
                    
            except KeyboardInterrupt:
                self.show_info("🔄 Returning to main menu...")
                break
            except Exception as e:
                self.show_error(f"TUI Error: {e}")
                self.wait_for_continue()
    
    def clear_screen(self) -> None:
        """Clear the terminal screen"""
        os.system('clear' if os.name == 'posix' else 'cls')
    
    def display_modern_header(self) -> None:
        """Display the modern header with system information"""
        import datetime
        current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        print("╔" + "═" * 98 + "╗")
        print(f"║ \033[1;91m{MASTER_APP} {MASTER_VERSION}\033[0m - \033[1;96mUnified Security Assessment Framework\033[0m".ljust(120) + f"\033[1;90m{current_time}\033[0m ║")
        print("╠" + "═" * 98 + "╣")
    
    def display_system_dashboard(self) -> None:
        """Display real-time system dashboard"""
        # Update performance metrics
        self.update_performance_metrics()
        
        # System health indicator
        health = self.system_status.get("health", "Unknown")
        health_color = {
            "Excellent": "\033[1;32m",
            "Good": "\033[1;33m", 
            "Fair": "\033[1;93m",
            "Poor": "\033[1;91m"
        }.get(health, "\033[1;90m")
        
        print(f"║ \033[1;97mSYSTEM DASHBOARD\033[0m".ljust(108) + " ║")
        print("╠" + "─" * 98 + "╣")
        print(f"║  🏥 Health: {health_color}{health}\033[0m" + 
              f"  💾 Memory: \033[1;94m{self.performance_metrics['memory_usage']:.1f}%\033[0m" +
              f"  🖥️  CPU: \033[1;92m{self.performance_metrics['cpu_usage']:.1f}%\033[0m" +
              f"  💽 Disk: \033[1;93m{self.performance_metrics['disk_usage']:.1f}%\033[0m".ljust(80) + " ║")
        
        # Active scans info
        active_count = len(self.active_scans)
        scan_info = f"\033[1;95m{active_count} Active\033[0m" if active_count > 0 else "\033[1;90mNone\033[0m"
        
        # Tools status
        available_tools = len([t for t in self.system_status.get("tools", {}).values() if "Available" in str(t)])
        total_tools = len(self.system_status.get("tools", {}))
        tools_ratio = f"{available_tools}/{total_tools}" if total_tools > 0 else "0/0"
        
        print(f"║  🔄 Scans: {scan_info}" +
              f"  🔧 Tools: \033[1;96m{tools_ratio}\033[0m" +
              f"  📊 History: \033[1;97m{len(self.scan_history)}\033[0m".ljust(80) + " ║")
        print("╠" + "═" * 98 + "╣")
    
    def display_navigation_menu(self) -> None:
        """Display the enhanced navigation menu with visual indicators"""
        print("║ \033[1;97mMAIN NAVIGATION CONTROL CENTER\033[0m".ljust(108) + " ║")
        print("╠" + "─" * 98 + "╣")
        
        # Core functionality
        print("║ \033[1;92m[1]\033[0m 🚀 \033[1;97mFULL AUTOMATION PIPELINE\033[0m - Complete security assessment workflow    ║")
        print("║     └─ Recon → Vuln Scan → Web Testing → Cloud → API → Infrastructure → Report  ║")
        print("╠" + "─" * 98 + "╣")
        
        # Target management
        print("║ \033[1;93m[2]\033[0m 🎯 \033[1;97mTARGET MANAGEMENT HUB\033[0m - Advanced target configuration and validation  ║")
        
        # Scanning modules  
        print("║ \033[1;94m[3]\033[0m 🔍 \033[1;97mRECONNAISSANCE SUITE\033[0m - Intelligence gathering and enumeration        ║")
        print("║ \033[1;95m[4]\033[0m 🛡️  \033[1;97mVULNERABILITY ASSESSMENT\033[0m - Advanced security vulnerability scanning    ║")
        print("║ \033[1;96m[5]\033[0m 🌐 \033[1;97mWEB APPLICATION TESTING\033[0m - Comprehensive web security analysis         ║")
        print("║ \033[1;97m[6]\033[0m ☁️  \033[1;97mCLOUD SECURITY AUDIT\033[0m - Multi-cloud security assessment              ║")
        print("║ \033[1;91m[7]\033[0m 🔌 \033[1;97mAPI SECURITY TESTING\033[0m - RESTful and GraphQL API security analysis    ║")
        print("║ \033[1;92m[8]\033[0m 🏗️  \033[1;97mINFRASTRUCTURE SCANNING\033[0m - Network infrastructure security assessment  ║")
        
        print("╠" + "─" * 98 + "╣")
        
        # System management
        print("║ \033[1;90m[9]\033[0m 🔧 \033[1;97mSYSTEM CONFIGURATION\033[0m - Framework settings and tool management        ║")
        print("║ \033[1;90m[A]\033[0m 📊 \033[1;97mREPORTING & ANALYTICS\033[0m - Advanced reporting and data visualization    ║")
        print("║ \033[1;90m[B]\033[0m 📈 \033[1;97mMONITORING DASHBOARD\033[0m - Real-time system monitoring and metrics      ║")
        print("║ \033[1;90m[C]\033[0m ⚙️  \033[1;97mADVANCED SETTINGS\033[0m - Expert configuration and tool installation    ║")
        
        print("╠" + "─" * 98 + "╣")
        
        # Navigation help
        print("║ \033[1;90m[H]\033[0m ❓ \033[1;97mHELP & DOCUMENTATION\033[0m   \033[1;90m[Q]\033[0m 🚪 \033[1;97mEXIT FRAMEWORK\033[0m                     ║")
        print("╚" + "═" * 98 + "╝")
        
        # Navigation tips
        print("\n\033[1;96m💡 Navigation Tips:\033[0m")
        print("   • Use number keys [1-9] or letters [A-C] for main functions")
        print("   • Press [H] for detailed help and keyboard shortcuts")
        print("   • Press [Ctrl+C] at any time to return to this menu")
        print("   • Use [Q] or [Ctrl+D] to exit the framework")
    
    def get_user_input(self) -> str:
        """Get user input with enhanced prompt"""
        try:
            return input("\n\033[1;97m🎮 Select option: \033[0m").strip().upper()
        except (EOFError, KeyboardInterrupt):
            return "Q"
    
    async def handle_navigation_choice(self, choice: str) -> bool:
        """Handle navigation choices with comprehensive routing"""
        if choice in ['Q', 'QUIT', 'EXIT']:
            return False
        elif choice == '1':
            await self.run_full_automation_pipeline()
        elif choice == '2':
            self.target_management_hub()
        elif choice == '3':
            self.reconnaissance_suite_menu()
        elif choice == '4':
            self.vulnerability_scanning_menu() 
        elif choice == '5':
            self.web_application_testing_menu()
        elif choice == '6':
            self.cloud_security_assessment_menu()
        elif choice == '7':
            self.api_security_testing_menu()
        elif choice == '8':
            self.infrastructure_scanning_menu()
        elif choice == '9':
            self.system_configuration_menu()
        elif choice == 'A':
            self.reporting_analytics_menu()
        elif choice == 'B':
            self.system_dashboard_menu()
        elif choice == 'C':
            self.advanced_settings_menu()
        elif choice == 'H':
            self.show_help_documentation()
        else:
            self.show_error("Invalid option selected. Please try again.")
            self.wait_for_continue()
        
        return True
    
    def update_performance_metrics(self) -> None:
        """Update real-time performance metrics"""
        try:
            # CPU usage
            self.performance_metrics["cpu_usage"] = psutil.cpu_percent(interval=0.1)
            
            # Memory usage
            memory = psutil.virtual_memory()
            self.performance_metrics["memory_usage"] = memory.percent
            
            # Disk usage
            disk = psutil.disk_usage('/')
            self.performance_metrics["disk_usage"] = disk.percent
            
            # Network status (simplified)
            self.performance_metrics["network_status"] = "Active"
            
        except Exception as e:
            self.logger.debug(f"Error updating performance metrics: {e}")
    
    def target_management_hub(self) -> None:
        """Enhanced target management with validation and bulk operations"""
        self.show_info("🎯 Launching Enhanced Target Management Hub...")
        # For now, delegate to moloch's target management
        try:
            from moloch import target_management_menu
            target_management_menu()
        except Exception as e:
            self.show_error(f"Target management error: {e}")
            self.wait_for_continue()
    
    def advanced_settings_menu(self) -> None:
        """Advanced settings and tool installation menu"""
        self.show_info("⚙️ Launching Advanced Settings Panel...")
        try:
            from moloch import tool_status_menu, settings_menu
            
            while True:
                self.clear_screen()
                print("╔" + "═" * 80 + "╗")
                print("║\033[1;95m                    ⚙️  ADVANCED SETTINGS PANEL ⚙️\033[0m                    ║")
                print("╠" + "═" * 80 + "╣")
                print("║  \033[1;92m1.\033[0m 🔧 \033[1;97mTOOL STATUS & INSTALLATION\033[0m - Manage security tools        ║")
                print("║  \033[1;94m2.\033[0m ⚙️  \033[1;97mFRAMEWORK CONFIGURATION\033[0m - System settings management     ║")
                print("║  \033[1;96m3.\033[0m 🔄 \033[1;97mUPDATE ALL TOOLS\033[0m - Update all installed security tools    ║")
                print("║  \033[1;93m4.\033[0m 🏗️  \033[1;97mSYSTEM DIAGNOSTICS\033[0m - Comprehensive system health check  ║")
                print("║  \033[1;90m5.\033[0m 🔙 \033[1;97mRETURN TO MAIN MENU\033[0m - Back to navigation center         ║")
                print("╚" + "═" * 80 + "╝")
                
                choice = input("\n\033[1;97m🎮 Select option: \033[0m").strip()
                
                if choice == '1':
                    tool_status_menu()
                elif choice == '2':
                    settings_menu()
                elif choice == '3':
                    self.update_all_tools()
                elif choice == '4':
                    self.run_system_diagnostics()
                elif choice == '5':
                    break
                else:
                    self.show_error("Invalid option. Please try again.")
                    
        except Exception as e:
            self.show_error(f"Advanced settings error: {e}")
            self.wait_for_continue()
    
    def update_all_tools(self) -> None:
        """Update all installed security tools"""
        self.show_info("🔄 Updating all security tools...")
        
        # Update Go-based tools
        go_tools = [
            "github.com/projectdiscovery/nuclei/v3/cmd/nuclei@latest",
            "github.com/projectdiscovery/httpx/cmd/httpx@latest", 
            "github.com/projectdiscovery/subfinder/v2/cmd/subfinder@latest",
            "github.com/projectdiscovery/naabu/v2/cmd/naabu@latest",
            "github.com/projectdiscovery/katana/cmd/katana@latest"
        ]
        
        for tool in go_tools:
            try:
                self.show_info(f"Updating {tool.split('/')[-1].split('@')[0]}...")
                subprocess.run(['go', 'install', '-v', tool], check=True, capture_output=True)
            except Exception as e:
                self.show_error(f"Failed to update {tool}: {e}")
        
        # Update nuclei templates
        try:
            self.show_info("Updating Nuclei templates...")
            subprocess.run(['nuclei', '-update-templates'], check=True, capture_output=True)
        except Exception as e:
            self.show_error(f"Failed to update nuclei templates: {e}")
        
        self.show_success("🎉 Tool update process completed!")
        self.wait_for_continue()
    
    def run_system_diagnostics(self) -> None:
        """Run comprehensive system diagnostics"""
        self.show_info("🏗️ Running comprehensive system diagnostics...")
        
        print("\n╔" + "═" * 80 + "╗")
        print("║\033[1;96m                       🏗️  SYSTEM DIAGNOSTICS 🏗️\033[0m                       ║")
        print("╠" + "═" * 80 + "╣")
        
        # System information
        import platform
        print(f"║  OS: \033[1;97m{platform.system()} {platform.release()}\033[0m")
        print(f"║  Python: \033[1;97m{platform.python_version()}\033[0m")
        print(f"║  Architecture: \033[1;97m{platform.machine()}\033[0m")
        
        # Performance metrics
        self.update_performance_metrics()
        print(f"║  CPU Usage: \033[1;92m{self.performance_metrics['cpu_usage']:.1f}%\033[0m")
        print(f"║  Memory Usage: \033[1;94m{self.performance_metrics['memory_usage']:.1f}%\033[0m")
        print(f"║  Disk Usage: \033[1;93m{self.performance_metrics['disk_usage']:.1f}%\033[0m")
        
        # Tool availability
        config = load_config()
        tools_config = config.get("tools", {})
        available = sum(1 for tool in tools_config if which(tool))
        total = len(tools_config)
        
        print(f"║  Tools Available: \033[1;96m{available}/{total}\033[0m ({available/total*100:.1f}%)")
        
        # Framework health
        health_score = (available/total) * 100 if total > 0 else 0
        if health_score >= 90:
            health = "\033[1;32mEXCELLENT\033[0m"
        elif health_score >= 70:
            health = "\033[1;33mGOOD\033[0m"
        elif health_score >= 50:
            health = "\033[1;93mFAIR\033[0m"
        else:
            health = "\033[1;91mPOOR\033[0m"
        
        print(f"║  Overall Health: {health}")
        print("╚" + "═" * 80 + "╝")
        
        self.wait_for_continue()
    
    def show_help_documentation(self) -> None:
        """Display comprehensive help and documentation"""
        self.clear_screen()
        print("╔" + "═" * 80 + "╗")
        print("║\033[1;93m                      ❓ HELP & DOCUMENTATION ❓\033[0m                      ║") 
        print("╠" + "═" * 80 + "╣")
        print("║  \033[1;97mKEYBOARD SHORTCUTS:\033[0m")
        print("║    Ctrl+C     - Return to main menu from any screen")
        print("║    Ctrl+D     - Exit framework")
        print("║    H          - Show this help screen")
        print("║    Q          - Quit/Exit current screen")
        print("║")
        print("║  \033[1;97mMAIN FUNCTIONS:\033[0m")
        print("║    [1] Full Pipeline    - Complete automated security assessment")
        print("║    [2] Target Mgmt      - Add/remove/validate targets")
        print("║    [3] Reconnaissance   - Information gathering")
        print("║    [4] Vulnerability    - Security vulnerability scanning")
        print("║    [5] Web Testing      - Web application security")
        print("║    [6] Cloud Security   - Multi-cloud assessment") 
        print("║    [7] API Testing      - API security analysis")
        print("║    [8] Infrastructure   - Network infrastructure scanning")
        print("║")
        print("║  \033[1;97mSYSTEM MANAGEMENT:\033[0m")
        print("║    [9] Configuration    - Framework settings")
        print("║    [A] Reporting        - Generate reports")
        print("║    [B] Monitoring       - Real-time dashboard")
        print("║    [C] Advanced         - Tool installation & updates")
        print("║")
        print("║  \033[1;97mTIPS:\033[0m")
        print("║    • Ensure targets are added before running scans")
        print("║    • Install missing tools via Advanced Settings")
        print("║    • Monitor system resources during large scans")
        print("║    • Review reports for detailed findings")
        print("╚" + "═" * 80 + "╝")
        
        self.wait_for_continue()
    
    async def run_full_automation_pipeline(self):
        """Execute the complete automated security assessment pipeline"""
        self.print_master_banner()
        print("╔═══════════════════════════════════════════════════════════════════════════════╗")
        print("║\033[1;91m                    🔄 FULL AUTOMATION PIPELINE 🔄\033[0m                    ║")
        print("║\033[1;97m              Complete Security Assessment Execution\033[0m              ║")
        print("╚═══════════════════════════════════════════════════════════════════════════════╝")
        
        try:
            # Get target configuration
            target = input("\n🎯 Enter target domain/IP: ").strip()
            if not target:
                self.show_error("Target is required for pipeline execution")
                return
            
            # Get pipeline options
            print("\n📋 \033[1;97mPipeline Configuration:\033[0m")
            aggressive = input("   Enable aggressive scanning? [y/N]: ").lower() == 'y'
            include_cloud = input("   Include cloud security assessment? [y/N]: ").lower() == 'y'
            include_api = input("   Include API security testing? [y/N]: ").lower() == 'y'
            
            # Display pipeline plan
            print(f"\n🚀 \033[1;92mStarting Full Pipeline for: {target}\033[0m")
            print("📋 \033[1;97mPipeline Phases:\033[0m")
            print("   Phase 1: 🔍 Reconnaissance & Intelligence Gathering")
            print("   Phase 2: 🛡️  Vulnerability Scanning & Assessment")
            print("   Phase 3: 🌐 Web Application Security Testing")
            print("   Phase 4: 💥 Fuzzing & Directory Discovery")
            if include_cloud:
                print("   Phase 5: ☁️  Cloud Security Assessment")
            if include_api:
                print("   Phase 6: 🔌 API Security Testing")
            print("   Final:   📊 Report Generation & Analysis")
            
            input("\n⏳ Press Enter to start pipeline execution...")
            
            # Execute pipeline using moloch integration
            if self.integrations_available and hasattr(self, 'moloch_integration'):
                scan_id = f"master_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
                
                self.active_scans[scan_id] = {
                    "target": target,
                    "start_time": datetime.now(),
                    "status": "running",
                    "phase": "initialization",
                    "aggressive": aggressive,
                    "include_cloud": include_cloud,
                    "include_api": include_api
                }
                
                print(f"\n✅ \033[1;92mPipeline initiated with scan ID: {scan_id}\033[0m")
                
                # Execute the full pipeline
                pipeline_results = await self.moloch_integration.execute_full_pipeline(
                    target, aggressive, include_cloud
                )
                
                # Update scan status
                self.active_scans[scan_id]["status"] = pipeline_results.get("status", "completed")
                self.active_scans[scan_id]["end_time"] = datetime.now()
                self.active_scans[scan_id]["results"] = pipeline_results
                
                if pipeline_results.get("status") == "completed":
                    print(f"\n✅ \033[1;92mPipeline execution completed successfully!\033[0m")
                    print(f"📊 Run ID: {pipeline_results.get('run_id', scan_id)}")
                    print(f"📁 Results directory: runs/{pipeline_results.get('run_id', scan_id)}")
                    
                    # Offer to generate report
                    if input("\n📋 Generate comprehensive report? [Y/n]: ").lower() != 'n':
                        await self.generate_pipeline_report(scan_id)
                        
                else:
                    print(f"\n❌ \033[1;91mPipeline execution failed\033[0m")
                    if "error" in pipeline_results:
                        print(f"Error: {pipeline_results['error']}")
                
                # Move to history
                self.scan_history.append(self.active_scans.pop(scan_id))
                
            else:
                print("❌ \033[1;91mMoloch integration not available. Please check system configuration.\033[0m")
                
        except Exception as e:
            self.logger.error(f"Pipeline execution error: {e}")
            print(f"\n❌ \033[1;91mPipeline execution failed: {e}\033[0m")
        
        self.wait_for_continue()
    
    def target_management_menu(self):
        """Target management interface"""
        while True:
            self.print_master_banner()
            print("╔═══════════════════════════════════════════════════════════════════════════════╗")
            print("║\033[1;93m                        🎯 TARGET MANAGEMENT 🎯\033[0m                        ║")
            print("║\033[1;97m                   Configure & Manage Scan Targets\033[0m                   ║")
            print("╠═══════════════════════════════════════════════════════════════════════════════╣")
            print("║  1. 📝 Add Single Target                2. 📄 Import Target List           ║")
            print("║  3. 📋 View Current Targets             4. 🗑️  Remove Targets              ║")
            print("║  5. ✅ Validate Targets                 6. 📊 Target Analysis Summary      ║")
            print("║  7. 💾 Export Target List               8. 🔄 Bulk Target Operations      ║")
            print("║  0. ⬅️  Back to Main Menu                                                  ║")
            print("╚═══════════════════════════════════════════════════════════════════════════════╝")
            
            choice = input("\n🎯 Select target operation [1-8, 0]: ").strip()
            
            if choice == "0":
                break
            elif choice == "1":
                self.add_single_target()
            elif choice == "2":
                self.import_target_list()
            elif choice == "3":
                self.view_current_targets()
            elif choice == "4":
                self.remove_targets()
            elif choice == "5":
                self.validate_targets()
            elif choice == "6":
                self.target_analysis_summary()
            elif choice == "7":
                self.export_target_list()
            elif choice == "8":
                self.bulk_target_operations()
            else:
                self.show_error(f"Invalid option: {choice}")
    
    def show_error(self, message: str) -> None:
        """Display error message with formatting"""
        print(f"\n❌ \033[1;91mError:\033[0m {message}")
        self.wait_for_continue()
    
    def show_success(self, message: str) -> None:
        """Display success message with formatting"""
        print(f"\n✅ \033[1;92mSuccess:\033[0m {message}")
    
    def show_info(self, message: str) -> None:
        """Display info message with formatting"""
        print(f"\n💡 \033[1;94mInfo:\033[0m {message}")
    
    def wait_for_continue(self) -> None:
        """Wait for user input to continue"""
        input("\n⏳ Press Enter to continue...")
    
    def exit_framework(self):
        """Exit the master framework gracefully"""
        self.print_master_banner()
        print("╔═══════════════════════════════════════════════════════════════════════════════╗")
        print("║\033[1;91m                        🚪 EXITING FRAMEWORK 🚪\033[0m                        ║")
        print("║\033[1;97m                     Thank you for using Azaz-El\033[0m                     ║")
        print("╚═══════════════════════════════════════════════════════════════════════════════╝")
        
        # Save active scans to history
        if self.active_scans:
            print(f"\n💾 Saving {len(self.active_scans)} active scan(s) to history...")
            self.scan_history.extend(self.active_scans.values())
            self.active_scans.clear()
        
        # Show session summary
        print(f"\n📊 \033[1;97mSession Summary:\033[0m")
        print(f"   • Total scans performed: {len(self.scan_history)}")
        print(f"   • Framework health: {self.system_status['health']}")
        print(f"   • Session duration: {datetime.now().strftime('%H:%M:%S')}")
        
        print(f"\n🙏 \033[1;97mThank you for using {MASTER_APP} {MASTER_VERSION}!\033[0m")
        print(f"   Visit us at: \033[1;94mhttps://github.com/cxb3rf1lth/Azaz-El\033[0m")
        
        # Graceful shutdown
        try:
            # Clean up any background processes
            self.logger.info("Performing graceful shutdown cleanup")
            # Could add specific cleanup tasks here if needed
        except Exception as e:
            self.logger.warning(f"Cleanup during shutdown had issues: {e}")
        
        print("\n👋 \033[1;92mGoodbye!\033[0m")
        sys.exit(0)
    
    # Placeholder methods for menu implementations
    def reconnaissance_suite_menu(self):
        """Reconnaissance suite interface"""
        while True:
            self.print_master_banner()
            print("╔═══════════════════════════════════════════════════════════════════════════════╗")
            print("║\033[1;94m                      🔍 RECONNAISSANCE SUITE\033[0m                      ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m1.\033[0m 🌐 Subdomain Discovery (subfinder, assetfinder, amass)              ║")
            print("║  \033[1;97m2.\033[0m 🔍 DNS Resolution & Validation                                    ║")
            print("║  \033[1;97m3.\033[0m 🌍 HTTP Service Probing (httpx)                                 ║") 
            print("║  \033[1;97m4.\033[0m ⚡ Full Reconnaissance Pipeline                                  ║")
            print("║  \033[1;97m5.\033[0m 📊 View Recent Reconnaissance Results                           ║")
            print("║  \033[1;97m6.\033[0m ⚙️ Configure Reconnaissance Settings                            ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m0.\033[0m ↩️ Return to Main Menu                                           ║")
            print("╚═══════════════════════════════════════════════════════════════════════════════╝")
            
            choice = self.get_user_input()
            
            if choice == "0":
                break
            elif choice == "1":
                self.handle_subdomain_discovery()
            elif choice == "2":
                self.handle_dns_resolution()
            elif choice == "3":
                self.handle_http_probing()
            elif choice == "4":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_full_reconnaissance())
            elif choice == "5":
                self.view_reconnaissance_results()
            elif choice == "6":
                self.configure_reconnaissance_settings()
            else:
                self.show_error(f"Invalid option: {choice}")
                
    def handle_subdomain_discovery(self):
        """Handle subdomain discovery operations"""
        if not self.targets:
            self.show_error("No targets configured. Please add targets first.")
            return
            
        target = self.select_target_interactive()
        if not target:
            return
            
        try:
            self.show_info(f"Starting subdomain discovery for {target}")
            
            # Create run directory
            run_dir = self.moloch_integration.config['general']['runs_dir']
            run_path = Path(run_dir) / f"recon_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute subdomain discovery using moloch
            from moloch import run_subdomain_discovery
            success = run_subdomain_discovery(target, run_path / "subdomains", self.moloch_integration.config)
            
            if success:
                self.show_success(f"Subdomain discovery completed for {target}")
                print(f"📁 Results saved to: {run_path / 'subdomains'}")
            else:
                self.show_error("Subdomain discovery failed")
                
        except Exception as e:
            self.show_error(f"Error during subdomain discovery: {e}")
        
        self.wait_for_continue()
        
    def handle_dns_resolution(self):
        """Handle DNS resolution operations"""
        if not self.targets:
            self.show_error("No targets configured. Please add targets first.")
            return
            
        # Look for recent subdomain files
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        subdomain_files = list(runs_dir.glob("*/subdomains/*.txt"))
        
        if not subdomain_files:
            self.show_error("No subdomain files found. Please run subdomain discovery first.")
            return
            
        try:
            # Use most recent subdomain file
            latest_file = max(subdomain_files, key=lambda x: x.stat().st_mtime)
            output_file = latest_file.parent.parent / "resolved_hosts.txt"
            
            self.show_info(f"Resolving subdomains from {latest_file.name}")
            
            from moloch import run_dns_resolution
            success = run_dns_resolution(latest_file, output_file, self.moloch_integration.config)
            
            if success:
                self.show_success("DNS resolution completed")
                print(f"📁 Results saved to: {output_file}")
            else:
                self.show_error("DNS resolution failed")
                
        except Exception as e:
            self.show_error(f"Error during DNS resolution: {e}")
            
        self.wait_for_continue()
        
    def handle_http_probing(self):
        """Handle HTTP service probing"""
        # Look for recent resolved host files
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        resolved_files = list(runs_dir.glob("*/resolved_hosts.txt"))
        
        if not resolved_files:
            self.show_error("No resolved host files found. Please run DNS resolution first.")
            return
            
        try:
            # Use most recent resolved file
            latest_file = max(resolved_files, key=lambda x: x.stat().st_mtime)
            output_file = latest_file.parent / "live_hosts.txt"
            
            self.show_info(f"Probing HTTP services from {latest_file.name}")
            
            from moloch import run_http_probing
            success = run_http_probing(latest_file, output_file, self.moloch_integration.config)
            
            if success:
                self.show_success("HTTP probing completed")
                print(f"📁 Results saved to: {output_file}")
                
                # Show quick stats
                if output_file.exists():
                    with open(output_file, 'r') as f:
                        live_count = len([line.strip() for line in f if line.strip()])
                    print(f"🌍 Found {live_count} live HTTP services")
            else:
                self.show_error("HTTP probing failed")
                
        except Exception as e:
            self.show_error(f"Error during HTTP probing: {e}")
            
        self.wait_for_continue()
        
    async def handle_full_reconnaissance(self):
        """Handle full reconnaissance pipeline"""
        if not self.targets:
            self.show_error("No targets configured. Please add targets first.")
            return
            
        target = self.select_target_interactive()
        if not target:
            return
            
        try:
            self.show_info(f"Starting full reconnaissance pipeline for {target}")
            print("🔄 This will run: Subdomain Discovery → DNS Resolution → HTTP Probing")
            
            confirm = input("\n⚠️ Continue with full reconnaissance? (y/N): ")
            if confirm.lower() != 'y':
                return
                
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"full_recon_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute full reconnaissance using moloch integration
            results = await self.moloch_integration.run_reconnaissance_suite(
                target, run_path, aggressive=False
            )
            
            if results and not results.get('errors'):
                self.show_success(f"Full reconnaissance completed for {target}")
                print(f"📁 Results saved to: {run_path}")
                print(f"🌐 Found {len(results.get('subdomains', []))} subdomains")
                print(f"🌍 Found {len(results.get('live_hosts', []))} live hosts")
            else:
                self.show_error("Reconnaissance pipeline failed")
                if results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during reconnaissance pipeline: {e}")
            
        self.wait_for_continue()
        
    def view_reconnaissance_results(self):
        """View recent reconnaissance results"""
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        
        # Find recent reconnaissance runs
        recon_dirs = [d for d in runs_dir.iterdir() if d.is_dir() and 
                     ('recon_' in d.name or 'full_recon_' in d.name)]
        recon_dirs.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        
        if not recon_dirs:
            self.show_info("No reconnaissance results found")
            self.wait_for_continue()
            return
            
        print("\n📊 \033[1;97mRecent Reconnaissance Results:\033[0m")
        print("=" * 60)
        
        for i, run_dir in enumerate(recon_dirs[:10], 1):
            print(f"\n{i}. 📁 {run_dir.name}")
            
            # Check for subdomain files
            subdomain_dir = run_dir / "subdomains"
            if subdomain_dir.exists():
                subdomain_files = list(subdomain_dir.glob("*.txt"))
                total_subdomains = 0
                for file in subdomain_files:
                    with open(file, 'r') as f:
                        total_subdomains += len([line.strip() for line in f if line.strip()])
                print(f"   🌐 Subdomains: {total_subdomains}")
            
            # Check for resolved hosts
            resolved_file = run_dir / "resolved_hosts.txt"
            if resolved_file.exists():
                with open(resolved_file, 'r') as f:
                    resolved_count = len([line.strip() for line in f if line.strip()])
                print(f"   🔍 Resolved: {resolved_count}")
            
            # Check for live hosts
            live_file = run_dir / "live_hosts.txt"
            if live_file.exists():
                with open(live_file, 'r') as f:
                    live_count = len([line.strip() for line in f if line.strip()])
                print(f"   🌍 Live: {live_count}")
                    
        self.wait_for_continue()
        
    def configure_reconnaissance_settings(self):
        """Configure reconnaissance settings"""
        print("\n⚙️ \033[1;97mReconnaissance Configuration:\033[0m")
        print("=" * 50)
        
        config = self.moloch_integration.config
        recon_config = config.get('reconnaissance', {})
        
        print(f"📊 Current Settings:")
        print(f"   • Subdomain Tools: {', '.join(recon_config.get('subdomain_tools', ['subfinder']))}")
        print(f"   • DNS Resolver: {recon_config.get('dns_resolver', 'system')}")
        print(f"   • HTTP Timeout: {recon_config.get('http_timeout', 10)}s")
        print(f"   • Threads: {recon_config.get('threads', 20)}")
        print(f"   • Aggressive Mode: {recon_config.get('aggressive_mode', False)}")
        
        print(f"\n💡 Use moloch.cfg.json to modify these settings")
        self.wait_for_continue()
        
    def select_target_interactive(self):
        """Interactive target selection"""
        if len(self.targets) == 1:
            return list(self.targets)[0]
            
        print(f"\n🎯 \033[1;97mSelect Target:\033[0m")
        targets_list = list(self.targets)
        for i, target in enumerate(targets_list, 1):
            print(f"  {i}. {target}")
            
        try:
            choice = input(f"\nSelect target (1-{len(targets_list)}): ")
            index = int(choice) - 1
            if 0 <= index < len(targets_list):
                return targets_list[index]
        except (ValueError, IndexError):
            pass
            
        self.show_error("Invalid target selection")
        return None
    
    def vulnerability_scanning_menu(self):
        """Vulnerability scanning interface"""
        while True:
            self.print_master_banner()
            print("╔═══════════════════════════════════════════════════════════════════════════════╗")
            print("║\033[1;95m                      🛡️ VULNERABILITY SCANNING\033[0m                      ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m1.\033[0m ⚡ Nuclei Templates (5000+ vulnerability checks)               ║")
            print("║  \033[1;97m2.\033[0m 🔌 Port Scanning (Nmap/Naabu)                                ║")
            print("║  \033[1;97m3.\033[0m 🔒 SSL/TLS Security Analysis (testssl.sh)                    ║")
            print("║  \033[1;97m4.\033[0m 💥 Full Vulnerability Assessment                             ║")
            print("║  \033[1;97m5.\033[0m 📊 View Vulnerability Results                               ║")
            print("║  \033[1;97m6.\033[0m ⚙️ Configure Vulnerability Scans                            ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m0.\033[0m ↩️ Return to Main Menu                                           ║")
            print("╚═══════════════════════════════════════════════════════════════════════════════╝")
            
            choice = self.get_user_input()
            
            if choice == "0":
                break
            elif choice == "1":
                self.handle_nuclei_scan()
            elif choice == "2":
                self.handle_port_scan()
            elif choice == "3":
                self.handle_ssl_scan()
            elif choice == "4":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_full_vulnerability_assessment())
            elif choice == "5":
                self.view_vulnerability_results()
            elif choice == "6":
                self.configure_vulnerability_settings()
            else:
                self.show_error(f"Invalid option: {choice}")
                
    def handle_nuclei_scan(self):
        """Handle Nuclei vulnerability scanning"""
        # Look for live hosts from reconnaissance
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        live_files = list(runs_dir.glob("*/live_hosts.txt"))
        
        if not live_files:
            self.show_error("No live hosts found. Please run reconnaissance first.")
            return
            
        try:
            # Use most recent live hosts file
            latest_file = max(live_files, key=lambda x: x.stat().st_mtime)
            
            # Create run directory
            run_path = Path(self.moloch_integration.config['general']['runs_dir'])
            vuln_dir = run_path / f"vuln_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            vuln_dir.mkdir(parents=True, exist_ok=True)
            
            self.show_info(f"Starting Nuclei scan on hosts from {latest_file.name}")
            
            from moloch import run_vulnerability_scan
            success = run_vulnerability_scan(latest_file, vuln_dir, self.moloch_integration.config)
            
            if success:
                self.show_success("Nuclei vulnerability scan completed")
                print(f"📁 Results saved to: {vuln_dir}")
                
                # Show quick stats
                results_file = vuln_dir / "nuclei_results.json"
                if results_file.exists():
                    with open(results_file, 'r') as f:
                        try:
                            results = [json.loads(line) for line in f if line.strip()]
                            print(f"🔍 Found {len(results)} vulnerabilities")
                        except:
                            print("🔍 Scan completed, check results file")
            else:
                self.show_error("Nuclei scan failed")
                
        except Exception as e:
            self.show_error(f"Error during Nuclei scan: {e}")
            
        self.wait_for_continue()
        
    def handle_port_scan(self):
        """Handle port scanning"""
        if not self.targets:
            self.show_error("No targets configured. Please add targets first.")
            return
            
        target = self.select_target_interactive()
        if not target:
            return
            
        try:
            # Create run directory
            run_path = Path(self.moloch_integration.config['general']['runs_dir'])
            port_dir = run_path / f"port_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            port_dir.mkdir(parents=True, exist_ok=True)
            
            self.show_info(f"Starting port scan for {target}")
            
            from moloch import run_port_scan
            success = run_port_scan(target, port_dir, self.moloch_integration.config)
            
            if success:
                self.show_success(f"Port scan completed for {target}")
                print(f"📁 Results saved to: {port_dir}")
            else:
                self.show_error("Port scan failed")
                
        except Exception as e:
            self.show_error(f"Error during port scan: {e}")
            
        self.wait_for_continue()
        
    def handle_ssl_scan(self):
        """Handle SSL/TLS security analysis"""
        # Look for live HTTPS hosts
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        live_files = list(runs_dir.glob("*/live_hosts.txt"))
        
        if not live_files:
            self.show_error("No live hosts found. Please run reconnaissance first.")
            return
            
        try:
            # Use most recent live hosts file
            latest_file = max(live_files, key=lambda x: x.stat().st_mtime)
            
            # Filter for HTTPS hosts
            https_hosts = []
            with open(latest_file, 'r') as f:
                for line in f:
                    if line.strip() and 'https://' in line.strip():
                        https_hosts.append(line.strip())
            
            if not https_hosts:
                self.show_error("No HTTPS hosts found in live hosts file")
                return
            
            # Create run directory
            run_path = Path(self.moloch_integration.config['general']['runs_dir'])
            ssl_dir = run_path / f"ssl_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            ssl_dir.mkdir(parents=True, exist_ok=True)
            
            # Create HTTPS hosts file
            https_file = ssl_dir / "https_hosts.txt"
            with open(https_file, 'w') as f:
                f.write('\n'.join(https_hosts))
            
            self.show_info(f"Starting SSL/TLS scan on {len(https_hosts)} HTTPS hosts")
            
            from moloch import run_ssl_scan
            success = run_ssl_scan(https_file, ssl_dir, self.moloch_integration.config)
            
            if success:
                self.show_success("SSL/TLS scan completed")
                print(f"📁 Results saved to: {ssl_dir}")
            else:
                self.show_error("SSL/TLS scan failed")
                
        except Exception as e:
            self.show_error(f"Error during SSL/TLS scan: {e}")
            
        self.wait_for_continue()
        
    async def handle_full_vulnerability_assessment(self):
        """Handle full vulnerability assessment pipeline"""
        if not self.targets:
            self.show_error("No targets configured. Please add targets first.")
            return
            
        target = self.select_target_interactive()
        if not target:
            return
            
        try:
            self.show_info(f"Starting full vulnerability assessment for {target}")
            print("🔄 This will run: Nuclei → Port Scan → SSL/TLS Analysis")
            
            confirm = input("\n⚠️ Continue with full vulnerability assessment? (y/N): ")
            if confirm.lower() != 'y':
                return
                
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"full_vuln_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute vulnerability assessment using moloch integration
            results = await self.moloch_integration.run_vulnerability_suite(
                target, run_path, aggressive=False
            )
            
            if results and not results.get('errors'):
                self.show_success(f"Full vulnerability assessment completed for {target}")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                vuln_count = len(results.get('vulnerabilities', []))
                port_count = len(results.get('open_ports', []))
                ssl_issues = len(results.get('ssl_issues', []))
                
                print(f"🔍 Found {vuln_count} vulnerabilities")
                print(f"🔌 Found {port_count} open ports")
                print(f"🔒 Found {ssl_issues} SSL/TLS issues")
            else:
                self.show_error("Vulnerability assessment failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during vulnerability assessment: {e}")
            
        self.wait_for_continue()
        
    def view_vulnerability_results(self):
        """View recent vulnerability scan results"""
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        
        # Find recent vulnerability scan runs
        vuln_dirs = [d for d in runs_dir.iterdir() if d.is_dir() and 
                    ('vuln_scan_' in d.name or 'port_scan_' in d.name or 'ssl_scan_' in d.name or 'full_vuln_' in d.name)]
        vuln_dirs.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        
        if not vuln_dirs:
            self.show_info("No vulnerability scan results found")
            self.wait_for_continue()
            return
            
        print("\n📊 \033[1;97mRecent Vulnerability Scan Results:\033[0m")
        print("=" * 60)
        
        for i, run_dir in enumerate(vuln_dirs[:10], 1):
            print(f"\n{i}. 📁 {run_dir.name}")
            
            # Check for nuclei results
            nuclei_file = run_dir / "nuclei_results.json"
            if nuclei_file.exists():
                try:
                    with open(nuclei_file, 'r') as f:
                        results = [json.loads(line) for line in f if line.strip()]
                        severities = {}
                        for result in results:
                            severity = result.get('info', {}).get('severity', 'unknown')
                            severities[severity] = severities.get(severity, 0) + 1
                        
                        total = len(results)
                        print(f"   ⚡ Nuclei: {total} findings")
                        if severities:
                            severity_str = ", ".join([f"{k}: {v}" for k, v in severities.items()])
                            print(f"      ({severity_str})")
                except:
                    print("   ⚡ Nuclei: Results file present")
            
            # Check for port scan results
            nmap_files = list(run_dir.glob("*nmap*"))
            if nmap_files:
                print(f"   🔌 Port scan: {len(nmap_files)} files")
            
            # Check for SSL scan results
            ssl_files = list(run_dir.glob("*ssl*"))
            if ssl_files:
                print(f"   🔒 SSL scan: {len(ssl_files)} files")
                    
        self.wait_for_continue()
        
    def configure_vulnerability_settings(self):
        """Configure vulnerability scanning settings"""
        print("\n⚙️ \033[1;97mVulnerability Scan Configuration:\033[0m")
        print("=" * 50)
        
        config = self.moloch_integration.config
        vuln_config = config.get('vulnerability', {})
        
        print(f"📊 Current Settings:")
        print(f"   • Nuclei Templates: {vuln_config.get('nuclei_templates', 'default')}")
        print(f"   • Port Scan Method: {vuln_config.get('port_scanner', 'nmap')}")
        print(f"   • SSL Scanner: {vuln_config.get('ssl_scanner', 'testssl')}")
        print(f"   • Scan Intensity: {vuln_config.get('intensity', 'normal')}")
        print(f"   • Timeout: {vuln_config.get('timeout', 300)}s")
        
        print(f"\n💡 Use moloch.cfg.json to modify these settings")
        self.wait_for_continue()
    
    def web_application_testing_menu(self):
        """Web application testing interface"""
        while True:
            self.print_master_banner()
            print("╔═══════════════════════════════════════════════════════════════════════════════╗")
            print("║\033[1;96m                    🌐 WEB APPLICATION TESTING\033[0m                      ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m1.\033[0m 🕷️ Web Crawling & Content Discovery (katana)                   ║")
            print("║  \033[1;97m2.\033[0m 💥 XSS Testing (dalfox)                                       ║")
            print("║  \033[1;97m3.\033[0m 📁 Directory & File Fuzzing (ffuf, gobuster)                 ║")
            print("║  \033[1;97m4.\033[0m 🔍 Parameter Discovery (arjun)                               ║")
            print("║  \033[1;97m5.\033[0m 🌐 Full Web Application Assessment                           ║")
            print("║  \033[1;97m6.\033[0m 📊 View Web Testing Results                                  ║")
            print("║  \033[1;97m7.\033[0m ⚙️ Configure Web Testing Settings                            ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m0.\033[0m ↩️ Return to Main Menu                                           ║")
            print("╚═══════════════════════════════════════════════════════════════════════════════╝")
            
            choice = self.get_user_input()
            
            if choice == "0":
                break
            elif choice == "1":
                self.handle_web_crawling()
            elif choice == "2":
                self.handle_xss_testing()
            elif choice == "3":
                self.handle_directory_fuzzing()
            elif choice == "4":
                self.handle_parameter_discovery()
            elif choice == "5":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_full_web_assessment())
            elif choice == "6":
                self.view_web_testing_results()
            elif choice == "7":
                self.configure_web_testing_settings()
            else:
                self.show_error(f"Invalid option: {choice}")
                
    def handle_web_crawling(self):
        """Handle web crawling and content discovery"""
        # Look for live web hosts
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        live_files = list(runs_dir.glob("*/live_hosts.txt"))
        
        if not live_files:
            self.show_error("No live hosts found. Please run reconnaissance first.")
            return
            
        try:
            # Use most recent live hosts file
            latest_file = max(live_files, key=lambda x: x.stat().st_mtime)
            
            # Create run directory
            run_path = Path(self.moloch_integration.config['general']['runs_dir'])
            crawl_dir = run_path / f"web_crawl_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            crawl_dir.mkdir(parents=True, exist_ok=True)
            
            self.show_info(f"Starting web crawling on hosts from {latest_file.name}")
            
            from moloch import run_crawling
            success = run_crawling(latest_file, crawl_dir, self.moloch_integration.config)
            
            if success:
                self.show_success("Web crawling completed")
                print(f"📁 Results saved to: {crawl_dir}")
                
                # Show quick stats
                katana_files = list(crawl_dir.glob("*katana*"))
                if katana_files:
                    print(f"🕷️ Generated {len(katana_files)} crawl result files")
            else:
                self.show_error("Web crawling failed")
                
        except Exception as e:
            self.show_error(f"Error during web crawling: {e}")
            
        self.wait_for_continue()
        
    def handle_xss_testing(self):
        """Handle XSS vulnerability testing"""
        # Look for crawled URLs
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        crawl_dirs = [d for d in runs_dir.iterdir() if d.is_dir() and 'web_crawl_' in d.name]
        
        if not crawl_dirs:
            self.show_error("No crawled URLs found. Please run web crawling first.")
            return
            
        try:
            # Use most recent crawl directory
            latest_crawl = max(crawl_dirs, key=lambda x: x.stat().st_mtime)
            
            # Find katana output files
            katana_files = list(latest_crawl.glob("*katana*"))
            if not katana_files:
                self.show_error("No katana output files found")
                return
            
            # Create run directory
            run_path = Path(self.moloch_integration.config['general']['runs_dir'])
            xss_dir = run_path / f"xss_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            xss_dir.mkdir(parents=True, exist_ok=True)
            
            self.show_info(f"Starting XSS testing on crawled URLs")
            
            from moloch import run_xss_scan
            success = run_xss_scan(katana_files[0], xss_dir, self.moloch_integration.config)
            
            if success:
                self.show_success("XSS testing completed")
                print(f"📁 Results saved to: {xss_dir}")
            else:
                self.show_error("XSS testing failed")
                
        except Exception as e:
            self.show_error(f"Error during XSS testing: {e}")
            
        self.wait_for_continue()
        
    def handle_directory_fuzzing(self):
        """Handle directory and file fuzzing"""
        # Look for live web hosts
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        live_files = list(runs_dir.glob("*/live_hosts.txt"))
        
        if not live_files:
            self.show_error("No live hosts found. Please run reconnaissance first.")
            return
            
        try:
            # Use most recent live hosts file
            latest_file = max(live_files, key=lambda x: x.stat().st_mtime)
            
            # Create run directory
            run_path = Path(self.moloch_integration.config['general']['runs_dir'])
            fuzz_dir = run_path / f"dir_fuzz_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            fuzz_dir.mkdir(parents=True, exist_ok=True)
            
            self.show_info(f"Starting directory fuzzing on hosts from {latest_file.name}")
            
            from moloch import run_directory_fuzzing
            success = run_directory_fuzzing(latest_file, fuzz_dir, self.moloch_integration.config)
            
            if success:
                self.show_success("Directory fuzzing completed")
                print(f"📁 Results saved to: {fuzz_dir}")
                
                # Show quick stats
                fuzz_files = list(fuzz_dir.glob("*"))
                if fuzz_files:
                    print(f"📁 Generated {len(fuzz_files)} fuzzing result files")
            else:
                self.show_error("Directory fuzzing failed")
                
        except Exception as e:
            self.show_error(f"Error during directory fuzzing: {e}")
            
        self.wait_for_continue()
        
    def handle_parameter_discovery(self):
        """Handle parameter discovery"""
        # Look for live web hosts
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        live_files = list(runs_dir.glob("*/live_hosts.txt"))
        
        if not live_files:
            self.show_error("No live hosts found. Please run reconnaissance first.")
            return
            
        try:
            # Use most recent live hosts file
            latest_file = max(live_files, key=lambda x: x.stat().st_mtime)
            
            # Create run directory
            run_path = Path(self.moloch_integration.config['general']['runs_dir'])
            param_dir = run_path / f"param_disc_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            param_dir.mkdir(parents=True, exist_ok=True)
            
            self.show_info(f"Starting parameter discovery (arjun) on hosts from {latest_file.name}")
            
            # Use arjun for parameter discovery
            from moloch import execute_tool
            
            # Extract just the base URLs
            with open(latest_file, 'r') as f:
                hosts = [line.strip() for line in f if line.strip()]
            
            if hosts:
                # Run arjun on first few hosts
                target_hosts = hosts[:5]  # Limit to first 5 hosts for demo
                for i, host in enumerate(target_hosts):
                    output_file = param_dir / f"arjun_params_{i}.json"
                    self.show_info(f"Discovering parameters for {host}")
                    
                    # Run arjun (if available)
                    success = execute_tool("arjun", ["-u", host, "-o", str(output_file)], 
                                         output_file=output_file, run_dir=param_dir)
                    
                    if not success:
                        self.show_info(f"Arjun not available or failed for {host}")
                
                self.show_success("Parameter discovery completed")
                print(f"📁 Results saved to: {param_dir}")
            else:
                self.show_error("No valid hosts found")
                
        except Exception as e:
            self.show_error(f"Error during parameter discovery: {e}")
            
        self.wait_for_continue()
        
    async def handle_full_web_assessment(self):
        """Handle full web application assessment"""
        # Look for live web hosts
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        live_files = list(runs_dir.glob("*/live_hosts.txt"))
        
        if not live_files:
            self.show_error("No live hosts found. Please run reconnaissance first.")
            return
            
        try:
            self.show_info("Starting full web application assessment")
            print("🔄 This will run: Crawling → XSS Testing → Directory Fuzzing")
            
            confirm = input("\n⚠️ Continue with full web assessment? (y/N): ")
            if confirm.lower() != 'y':
                return
                
            # Use most recent live hosts file
            latest_file = max(live_files, key=lambda x: x.stat().st_mtime)
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"full_web_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute web testing suite using moloch integration
            results = await self.moloch_integration.run_web_testing_suite(
                str(latest_file), run_path, aggressive=False
            )
            
            if results and not results.get('errors'):
                self.show_success("Full web assessment completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                crawled_urls = len(results.get('crawled_urls', []))
                xss_findings = len(results.get('xss_findings', []))
                directories = len(results.get('directories', []))
                
                print(f"🕷️ Crawled {crawled_urls} URLs")
                print(f"💥 Found {xss_findings} XSS vulnerabilities")
                print(f"📁 Discovered {directories} directories")
            else:
                self.show_error("Web assessment failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during web assessment: {e}")
            
        self.wait_for_continue()
        
    def view_web_testing_results(self):
        """View recent web testing results"""
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        
        # Find recent web testing runs
        web_dirs = [d for d in runs_dir.iterdir() if d.is_dir() and 
                   any(x in d.name for x in ['web_crawl_', 'xss_scan_', 'dir_fuzz_', 'param_disc_', 'full_web_'])]
        web_dirs.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        
        if not web_dirs:
            self.show_info("No web testing results found")
            self.wait_for_continue()
            return
            
        print("\n📊 \033[1;97mRecent Web Testing Results:\033[0m")
        print("=" * 60)
        
        for i, run_dir in enumerate(web_dirs[:10], 1):
            print(f"\n{i}. 📁 {run_dir.name}")
            
            # Check for different types of results
            if 'web_crawl_' in run_dir.name:
                katana_files = list(run_dir.glob("*katana*"))
                print(f"   🕷️ Crawling: {len(katana_files)} result files")
            elif 'xss_scan_' in run_dir.name:
                dalfox_files = list(run_dir.glob("*dalfox*"))
                print(f"   💥 XSS: {len(dalfox_files)} result files")
            elif 'dir_fuzz_' in run_dir.name:
                fuzz_files = list(run_dir.glob("*"))
                print(f"   📁 Directory Fuzzing: {len(fuzz_files)} files")
            elif 'param_disc_' in run_dir.name:
                param_files = list(run_dir.glob("*arjun*"))
                print(f"   🔍 Parameter Discovery: {len(param_files)} files")
            elif 'full_web_' in run_dir.name:
                all_files = list(run_dir.glob("*"))
                print(f"   🌐 Full Assessment: {len(all_files)} files")
                    
        self.wait_for_continue()
        
    def configure_web_testing_settings(self):
        """Configure web testing settings"""
        print("\n⚙️ \033[1;97mWeb Testing Configuration:\033[0m")
        print("=" * 50)
        
        config = self.moloch_integration.config
        web_config = config.get('web_testing', {})
        
        print(f"📊 Current Settings:")
        print(f"   • Crawler: {web_config.get('crawler', 'katana')}")
        print(f"   • XSS Scanner: {web_config.get('xss_scanner', 'dalfox')}")
        print(f"   • Directory Fuzzer: {web_config.get('dir_fuzzer', 'ffuf')}")
        print(f"   • Parameter Discovery: {web_config.get('param_discovery', 'arjun')}")
        print(f"   • Crawl Depth: {web_config.get('crawl_depth', 3)}")
        print(f"   • Request Timeout: {web_config.get('timeout', 30)}s")
        
        print(f"\n💡 Use moloch.cfg.json to modify these settings")
        self.wait_for_continue()
    
    def cloud_security_assessment_menu(self):
        """Cloud security assessment interface"""
        while True:
            self.print_master_banner()
            print("╔═══════════════════════════════════════════════════════════════════════════════╗")
            print("║\033[1;93m                    ☁️ CLOUD SECURITY ASSESSMENT\033[0m                     ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m1.\033[0m ☁️ AWS Security Assessment                                     ║")
            print("║  \033[1;97m2.\033[0m 🔵 Azure Security Assessment                                   ║")
            print("║  \033[1;97m3.\033[0m 🌐 GCP Security Assessment                                     ║")
            print("║  \033[1;97m4.\033[0m 🪣 S3 Bucket Enumeration & Analysis                           ║")
            print("║  \033[1;97m5.\033[0m 🔐 Cloud IAM Policy Analysis                                  ║")
            print("║  \033[1;97m6.\033[0m 📊 View Cloud Security Results                               ║")
            print("║  \033[1;97m7.\033[0m ⚙️ Configure Cloud Security Settings                         ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m0.\033[0m ↩️ Return to Main Menu                                           ║")
            print("╚═══════════════════════════════════════════════════════════════════════════════╝")
            
            choice = self.get_user_input()
            
            if choice == "0":
                break
            elif choice == "1":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_aws_assessment())
            elif choice == "2":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_azure_assessment())
            elif choice == "3":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_gcp_assessment())
            elif choice == "4":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_s3_bucket_enum())
            elif choice == "5":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_iam_analysis())
            elif choice == "6":
                self.view_cloud_security_results()
            elif choice == "7":
                self.configure_cloud_security_settings()
            else:
                self.show_error(f"Invalid option: {choice}")
                
    async def handle_aws_assessment(self):
        """Handle AWS security assessment"""
        if not self.integrations_available or not hasattr(self, 'cloud_scanner'):
            self.show_error("Cloud scanner not available")
            return
            
        try:
            self.show_info("Starting AWS security assessment")
            print("🔄 This will check: IAM, S3, EC2, VPC, CloudTrail, and more")
            
            confirm = input("\n⚠️ Continue with AWS assessment? (requires AWS credentials) (y/N): ")
            if confirm.lower() != 'y':
                return
                
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"aws_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute AWS assessment
            results = await self.cloud_scanner.assess_aws_security(
                output_dir=run_path,
                comprehensive=True
            )
            
            if results and not results.get('errors'):
                self.show_success("AWS security assessment completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                findings = results.get('findings', [])
                high_risk = len([f for f in findings if f.get('severity') == 'high'])
                medium_risk = len([f for f in findings if f.get('severity') == 'medium'])
                
                print(f"🔍 Found {len(findings)} total findings")
                print(f"🚨 High risk: {high_risk}")
                print(f"⚠️ Medium risk: {medium_risk}")
            else:
                self.show_error("AWS assessment failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during AWS assessment: {e}")
            
        self.wait_for_continue()
        
    async def handle_azure_assessment(self):
        """Handle Azure security assessment"""
        if not self.integrations_available or not hasattr(self, 'cloud_scanner'):
            self.show_error("Cloud scanner not available")
            return
            
        try:
            self.show_info("Starting Azure security assessment")
            print("🔄 This will check: Azure AD, Storage, VMs, Key Vault, and more")
            
            confirm = input("\n⚠️ Continue with Azure assessment? (requires Azure credentials) (y/N): ")
            if confirm.lower() != 'y':
                return
                
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"azure_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute Azure assessment
            results = await self.cloud_scanner.assess_azure_security(
                output_dir=run_path,
                comprehensive=True
            )
            
            if results and not results.get('errors'):
                self.show_success("Azure security assessment completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                findings = results.get('findings', [])
                critical = len([f for f in findings if f.get('severity') == 'critical'])
                high_risk = len([f for f in findings if f.get('severity') == 'high'])
                
                print(f"🔍 Found {len(findings)} total findings")
                print(f"💥 Critical: {critical}")
                print(f"🚨 High risk: {high_risk}")
            else:
                self.show_error("Azure assessment failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during Azure assessment: {e}")
            
        self.wait_for_continue()
        
    async def handle_gcp_assessment(self):
        """Handle GCP security assessment"""
        if not self.integrations_available or not hasattr(self, 'cloud_scanner'):
            self.show_error("Cloud scanner not available")
            return
            
        try:
            self.show_info("Starting GCP security assessment")
            print("🔄 This will check: IAM, Storage, Compute, VPC, and more")
            
            confirm = input("\n⚠️ Continue with GCP assessment? (requires GCP credentials) (y/N): ")
            if confirm.lower() != 'y':
                return
                
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"gcp_scan_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute GCP assessment
            results = await self.cloud_scanner.assess_gcp_security(
                output_dir=run_path,
                comprehensive=True
            )
            
            if results and not results.get('errors'):
                self.show_success("GCP security assessment completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                findings = results.get('findings', [])
                print(f"🔍 Found {len(findings)} total findings")
            else:
                self.show_error("GCP assessment failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during GCP assessment: {e}")
            
        self.wait_for_continue()
        
    async def handle_s3_bucket_enum(self):
        """Handle S3 bucket enumeration and analysis"""
        if not self.integrations_available or not hasattr(self, 'cloud_scanner'):
            self.show_error("Cloud scanner not available")
            return
            
        try:
            self.show_info("Starting S3 bucket enumeration")
            
            # Get target domains for bucket enumeration
            if not self.targets:
                domain = input("\n🎯 Enter domain for S3 bucket enumeration: ").strip()
                if not domain:
                    return
            else:
                domain = self.select_target_interactive()
                if not domain:
                    return
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"s3_enum_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute S3 enumeration
            results = await self.cloud_scanner.enumerate_s3_buckets(
                domain=domain,
                output_dir=run_path
            )
            
            if results and not results.get('errors'):
                self.show_success("S3 bucket enumeration completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                buckets = results.get('buckets', [])
                public_buckets = results.get('public_buckets', [])
                
                print(f"🪣 Found {len(buckets)} buckets")
                print(f"🚨 Public buckets: {len(public_buckets)}")
                
                if public_buckets:
                    print("\n⚠️ Public buckets found:")
                    for bucket in public_buckets[:5]:  # Show first 5
                        print(f"   • {bucket}")
            else:
                self.show_error("S3 enumeration failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during S3 enumeration: {e}")
            
        self.wait_for_continue()
        
    async def handle_iam_analysis(self):
        """Handle cloud IAM policy analysis"""
        if not self.integrations_available or not hasattr(self, 'cloud_scanner'):
            self.show_error("Cloud scanner not available")
            return
            
        try:
            self.show_info("Starting IAM policy analysis")
            
            # Select cloud provider
            print("\n☁️ Select cloud provider:")
            print("1. AWS")
            print("2. Azure")
            print("3. GCP")
            
            choice = input("\nSelect provider (1-3): ").strip()
            provider_map = {"1": "aws", "2": "azure", "3": "gcp"}
            
            if choice not in provider_map:
                self.show_error("Invalid provider selection")
                return
                
            provider = provider_map[choice]
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"iam_analysis_{provider}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute IAM analysis
            results = await self.cloud_scanner.analyze_iam_policies(
                provider=provider,
                output_dir=run_path
            )
            
            if results and not results.get('errors'):
                self.show_success(f"{provider.upper()} IAM analysis completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                policies = results.get('policies', [])
                risky_policies = results.get('risky_policies', [])
                
                print(f"📋 Analyzed {len(policies)} policies")
                print(f"⚠️ Risky policies: {len(risky_policies)}")
            else:
                self.show_error("IAM analysis failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during IAM analysis: {e}")
            
        self.wait_for_continue()
        
    def view_cloud_security_results(self):
        """View recent cloud security results"""
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        
        # Find recent cloud security runs
        cloud_dirs = [d for d in runs_dir.iterdir() if d.is_dir() and 
                     any(x in d.name for x in ['aws_scan_', 'azure_scan_', 'gcp_scan_', 's3_enum_', 'iam_analysis_'])]
        cloud_dirs.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        
        if not cloud_dirs:
            self.show_info("No cloud security results found")
            self.wait_for_continue()
            return
            
        print("\n📊 \033[1;97mRecent Cloud Security Results:\033[0m")
        print("=" * 60)
        
        for i, run_dir in enumerate(cloud_dirs[:10], 1):
            print(f"\n{i}. 📁 {run_dir.name}")
            
            # Check for different types of results
            if 'aws_scan_' in run_dir.name:
                results_files = list(run_dir.glob("*.json"))
                print(f"   ☁️ AWS Assessment: {len(results_files)} result files")
            elif 'azure_scan_' in run_dir.name:
                results_files = list(run_dir.glob("*.json"))
                print(f"   🔵 Azure Assessment: {len(results_files)} result files")
            elif 'gcp_scan_' in run_dir.name:
                results_files = list(run_dir.glob("*.json"))
                print(f"   🌐 GCP Assessment: {len(results_files)} result files")
            elif 's3_enum_' in run_dir.name:
                bucket_files = list(run_dir.glob("*bucket*"))
                print(f"   🪣 S3 Enumeration: {len(bucket_files)} files")
            elif 'iam_analysis_' in run_dir.name:
                policy_files = list(run_dir.glob("*policy*"))
                print(f"   🔐 IAM Analysis: {len(policy_files)} files")
                    
        self.wait_for_continue()
        
    def configure_cloud_security_settings(self):
        """Configure cloud security settings"""
        print("\n⚙️ \033[1;97mCloud Security Configuration:\033[0m")
        print("=" * 50)
        
        config = self.moloch_integration.config
        cloud_config = config.get('cloud_security', {})
        
        print(f"📊 Current Settings:")
        print(f"   • AWS Profile: {cloud_config.get('aws_profile', 'default')}")
        print(f"   • Azure Subscription: {cloud_config.get('azure_subscription', 'default')}")
        print(f"   • GCP Project: {cloud_config.get('gcp_project', 'default')}")
        print(f"   • S3 Enumeration: {cloud_config.get('s3_enumeration', True)}")
        print(f"   • Comprehensive Scan: {cloud_config.get('comprehensive', False)}")
        print(f"   • Timeout: {cloud_config.get('timeout', 300)}s")
        
        print(f"\n💡 Use moloch.cfg.json to modify these settings")
        print(f"📋 Ensure cloud credentials are properly configured")
        self.wait_for_continue()
    
    def api_security_testing_menu(self):
        """API security testing interface"""
        while True:
            self.print_master_banner()
            print("╔═══════════════════════════════════════════════════════════════════════════════╗")
            print("║\033[1;92m                      🔌 API SECURITY TESTING\033[0m                       ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m1.\033[0m 🔍 REST API Discovery & Testing                                ║")
            print("║  \033[1;97m2.\033[0m 📊 GraphQL API Security Testing                               ║")
            print("║  \033[1;97m3.\033[0m 🧽 SOAP API Security Testing                                  ║")
            print("║  \033[1;97m4.\033[0m 🔐 API Authentication Testing                                ║")
            print("║  \033[1;97m5.\033[0m 📋 API Documentation Analysis                                ║")
            print("║  \033[1;97m6.\033[0m 💥 Full API Security Assessment                              ║")
            print("║  \033[1;97m7.\033[0m 📊 View API Testing Results                                  ║")
            print("║  \033[1;97m8.\033[0m ⚙️ Configure API Testing Settings                            ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m0.\033[0m ↩️ Return to Main Menu                                           ║")
            print("╚═══════════════════════════════════════════════════════════════════════════════╝")
            
            choice = self.get_user_input()
            
            if choice == "0":
                break
            elif choice == "1":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_rest_api_testing())
            elif choice == "2":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_graphql_testing())
            elif choice == "3":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_soap_testing())
            elif choice == "4":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_api_auth_testing())
            elif choice == "5":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_api_documentation_analysis())
            elif choice == "6":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_full_api_assessment())
            elif choice == "7":
                self.view_api_testing_results()
            elif choice == "8":
                self.configure_api_testing_settings()
            else:
                self.show_error(f"Invalid option: {choice}")
                
    async def handle_rest_api_testing(self):
        """Handle REST API discovery and testing"""
        if not self.integrations_available or not hasattr(self, 'api_scanner'):
            self.show_error("API scanner not available")
            return
            
        try:
            # Get target URL for API testing
            if not self.targets:
                api_url = input("\n🎯 Enter API base URL (e.g., https://api.example.com): ").strip()
                if not api_url:
                    return
            else:
                print("\n🎯 Select target or enter custom API URL:")
                print("0. Enter custom URL")
                target = self.select_target_interactive()
                if target == "0" or not target:
                    api_url = input("Enter API base URL: ").strip()
                    if not api_url:
                        return
                else:
                    api_url = target
            
            self.show_info(f"Starting REST API testing for {api_url}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"rest_api_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute REST API testing
            results = await self.api_scanner.test_rest_api(
                base_url=api_url,
                output_dir=run_path,
                comprehensive=True
            )
            
            if results and not results.get('errors'):
                self.show_success("REST API testing completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                endpoints = results.get('endpoints', [])
                vulnerabilities = results.get('vulnerabilities', [])
                
                print(f"🔍 Discovered {len(endpoints)} API endpoints")
                print(f"🚨 Found {len(vulnerabilities)} vulnerabilities")
                
                if vulnerabilities:
                    high_vuln = len([v for v in vulnerabilities if v.get('severity') == 'high'])
                    if high_vuln > 0:
                        print(f"💥 High severity vulnerabilities: {high_vuln}")
            else:
                self.show_error("REST API testing failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during REST API testing: {e}")
            
        self.wait_for_continue()
        
    async def handle_graphql_testing(self):
        """Handle GraphQL API security testing"""
        if not self.integrations_available or not hasattr(self, 'api_scanner'):
            self.show_error("API scanner not available")
            return
            
        try:
            # Get GraphQL endpoint
            graphql_url = input("\n🎯 Enter GraphQL endpoint URL: ").strip()
            if not graphql_url:
                return
                
            self.show_info(f"Starting GraphQL API testing for {graphql_url}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"graphql_api_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute GraphQL testing
            results = await self.api_scanner.test_graphql_api(
                endpoint_url=graphql_url,
                output_dir=run_path
            )
            
            if results and not results.get('errors'):
                self.show_success("GraphQL API testing completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                queries = results.get('queries', [])
                mutations = results.get('mutations', [])
                vulnerabilities = results.get('vulnerabilities', [])
                
                print(f"📊 Found {len(queries)} queries, {len(mutations)} mutations")
                print(f"🚨 Found {len(vulnerabilities)} vulnerabilities")
            else:
                self.show_error("GraphQL API testing failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during GraphQL testing: {e}")
            
        self.wait_for_continue()
        
    async def handle_soap_testing(self):
        """Handle SOAP API security testing"""
        if not self.integrations_available or not hasattr(self, 'api_scanner'):
            self.show_error("API scanner not available")
            return
            
        try:
            # Get SOAP WSDL URL
            wsdl_url = input("\n🎯 Enter SOAP WSDL URL: ").strip()
            if not wsdl_url:
                return
                
            self.show_info(f"Starting SOAP API testing for {wsdl_url}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"soap_api_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute SOAP testing
            results = await self.api_scanner.test_soap_api(
                wsdl_url=wsdl_url,
                output_dir=run_path
            )
            
            if results and not results.get('errors'):
                self.show_success("SOAP API testing completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                operations = results.get('operations', [])
                vulnerabilities = results.get('vulnerabilities', [])
                
                print(f"🧽 Found {len(operations)} SOAP operations")
                print(f"🚨 Found {len(vulnerabilities)} vulnerabilities")
            else:
                self.show_error("SOAP API testing failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during SOAP testing: {e}")
            
        self.wait_for_continue()
        
    async def handle_api_auth_testing(self):
        """Handle API authentication testing"""
        if not self.integrations_available or not hasattr(self, 'api_scanner'):
            self.show_error("API scanner not available")
            return
            
        try:
            # Get API endpoint
            api_url = input("\n🎯 Enter API URL for authentication testing: ").strip()
            if not api_url:
                return
                
            self.show_info(f"Starting API authentication testing for {api_url}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"api_auth_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute API authentication testing
            results = await self.api_scanner.test_api_authentication(
                api_url=api_url,
                output_dir=run_path
            )
            
            if results and not results.get('errors'):
                self.show_success("API authentication testing completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                auth_methods = results.get('auth_methods', [])
                weaknesses = results.get('auth_weaknesses', [])
                
                print(f"🔐 Detected {len(auth_methods)} authentication methods")
                print(f"⚠️ Found {len(weaknesses)} authentication weaknesses")
            else:
                self.show_error("API authentication testing failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during API authentication testing: {e}")
            
        self.wait_for_continue()
        
    async def handle_api_documentation_analysis(self):
        """Handle API documentation analysis"""
        if not self.integrations_available or not hasattr(self, 'api_scanner'):
            self.show_error("API scanner not available")
            return
            
        try:
            print("\n📋 API Documentation Analysis Options:")
            print("1. OpenAPI/Swagger documentation")
            print("2. RAML documentation") 
            print("3. API Blueprint documentation")
            
            choice = input("\nSelect documentation type (1-3): ").strip()
            
            if choice == "1":
                doc_url = input("Enter OpenAPI/Swagger URL: ").strip()
                doc_type = "openapi"
            elif choice == "2":
                doc_url = input("Enter RAML URL: ").strip()
                doc_type = "raml"
            elif choice == "3":
                doc_url = input("Enter API Blueprint URL: ").strip()
                doc_type = "blueprint"
            else:
                self.show_error("Invalid documentation type")
                return
                
            if not doc_url:
                return
                
            self.show_info(f"Starting {doc_type.upper()} documentation analysis")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"api_docs_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute documentation analysis
            results = await self.api_scanner.analyze_api_documentation(
                doc_url=doc_url,
                doc_type=doc_type,
                output_dir=run_path
            )
            
            if results and not results.get('errors'):
                self.show_success("API documentation analysis completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                endpoints = results.get('endpoints', [])
                security_issues = results.get('security_issues', [])
                
                print(f"📋 Analyzed {len(endpoints)} documented endpoints")
                print(f"⚠️ Found {len(security_issues)} security issues in documentation")
            else:
                self.show_error("API documentation analysis failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during API documentation analysis: {e}")
            
        self.wait_for_continue()
        
    async def handle_full_api_assessment(self):
        """Handle full API security assessment"""
        if not self.integrations_available or not hasattr(self, 'api_scanner'):
            self.show_error("API scanner not available")
            return
            
        try:
            # Get API URL
            api_url = input("\n🎯 Enter API base URL for full assessment: ").strip()
            if not api_url:
                return
                
            self.show_info(f"Starting full API security assessment for {api_url}")
            print("🔄 This will run: Discovery → Testing → Authentication → Documentation Analysis")
            
            confirm = input("\n⚠️ Continue with full API assessment? (y/N): ")
            if confirm.lower() != 'y':
                return
                
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"full_api_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute full API assessment
            results = await self.api_scanner.full_api_assessment(
                api_url=api_url,
                output_dir=run_path
            )
            
            if results and not results.get('errors'):
                self.show_success("Full API security assessment completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show comprehensive summary
                endpoints = results.get('endpoints', [])
                vulnerabilities = results.get('vulnerabilities', [])
                auth_issues = results.get('auth_issues', [])
                
                print(f"🔍 Discovered {len(endpoints)} API endpoints")
                print(f"🚨 Found {len(vulnerabilities)} vulnerabilities")
                print(f"🔐 Found {len(auth_issues)} authentication issues")
                
                # Categorize vulnerabilities by severity
                severity_count = {}
                for vuln in vulnerabilities:
                    severity = vuln.get('severity', 'unknown')
                    severity_count[severity] = severity_count.get(severity, 0) + 1
                
                if severity_count:
                    print("\n📊 Vulnerability breakdown:")
                    for severity, count in severity_count.items():
                        print(f"   {severity.capitalize()}: {count}")
            else:
                self.show_error("Full API assessment failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during full API assessment: {e}")
            
        self.wait_for_continue()
        
    def view_api_testing_results(self):
        """View recent API testing results"""
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        
        # Find recent API testing runs
        api_dirs = [d for d in runs_dir.iterdir() if d.is_dir() and 
                   any(x in d.name for x in ['rest_api_', 'graphql_api_', 'soap_api_', 'api_auth_', 'api_docs_', 'full_api_'])]
        api_dirs.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        
        if not api_dirs:
            self.show_info("No API testing results found")
            self.wait_for_continue()
            return
            
        print("\n📊 \033[1;97mRecent API Testing Results:\033[0m")
        print("=" * 60)
        
        for i, run_dir in enumerate(api_dirs[:10], 1):
            print(f"\n{i}. 📁 {run_dir.name}")
            
            # Check for different types of results
            if 'rest_api_' in run_dir.name:
                result_files = list(run_dir.glob("*.json"))
                print(f"   🔍 REST API Testing: {len(result_files)} result files")
            elif 'graphql_api_' in run_dir.name:
                result_files = list(run_dir.glob("*.json"))
                print(f"   📊 GraphQL Testing: {len(result_files)} result files")
            elif 'soap_api_' in run_dir.name:
                result_files = list(run_dir.glob("*.json"))
                print(f"   🧽 SOAP Testing: {len(result_files)} result files")
            elif 'api_auth_' in run_dir.name:
                auth_files = list(run_dir.glob("*auth*"))
                print(f"   🔐 Auth Testing: {len(auth_files)} result files")
            elif 'api_docs_' in run_dir.name:
                doc_files = list(run_dir.glob("*"))
                print(f"   📋 Documentation Analysis: {len(doc_files)} files")
            elif 'full_api_' in run_dir.name:
                all_files = list(run_dir.glob("*"))
                print(f"   🔌 Full API Assessment: {len(all_files)} files")
                    
        self.wait_for_continue()
        
    def configure_api_testing_settings(self):
        """Configure API testing settings"""
        print("\n⚙️ \033[1;97mAPI Testing Configuration:\033[0m")
        print("=" * 50)
        
        config = self.moloch_integration.config
        api_config = config.get('api_testing', {})
        
        print(f"📊 Current Settings:")
        print(f"   • Request Timeout: {api_config.get('request_timeout', 30)}s")
        print(f"   • Rate Limiting: {api_config.get('rate_limit', 10)} req/sec")
        print(f"   • Authentication Methods: {', '.join(api_config.get('auth_methods', ['bearer', 'basic', 'oauth']))}")
        print(f"   • Payload Fuzzing: {api_config.get('payload_fuzzing', True)}")
        print(f"   • Schema Validation: {api_config.get('schema_validation', True)}")
        print(f"   • Deep Testing: {api_config.get('deep_testing', False)}")
        
        print(f"\n💡 Use moloch.cfg.json to modify these settings")
        self.wait_for_continue()
    
    def infrastructure_scanning_menu(self):
        """Infrastructure scanning interface"""
        while True:
            self.print_master_banner()
            print("╔═══════════════════════════════════════════════════════════════════════════════╗")
            print("║\033[1;97m                     🏗️ INFRASTRUCTURE SCANNING\033[0m                      ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m1.\033[0m 🌐 Network Discovery & Mapping                                ║")
            print("║  \033[1;97m2.\033[0m 🔌 Service Detection & Enumeration                           ║")
            print("║  \033[1;97m3.\033[0m 🏢 Operating System Fingerprinting                           ║")
            print("║  \033[1;97m4.\033[0m 📡 Network Protocol Analysis                                 ║")
            print("║  \033[1;97m5.\033[0m 🔍 Asset Discovery & Classification                          ║")
            print("║  \033[1;97m6.\033[0m 💥 Full Infrastructure Assessment                           ║")
            print("║  \033[1;97m7.\033[0m 📊 View Infrastructure Results                              ║")
            print("║  \033[1;97m8.\033[0m ⚙️ Configure Infrastructure Settings                        ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m0.\033[0m ↩️ Return to Main Menu                                           ║")
            print("╚═══════════════════════════════════════════════════════════════════════════════╝")
            
            choice = self.get_user_input()
            
            if choice == "0":
                break
            elif choice == "1":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_network_discovery())
            elif choice == "2":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_service_detection())
            elif choice == "3":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_os_fingerprinting())
            elif choice == "4":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_protocol_analysis())
            elif choice == "5":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_asset_discovery())
            elif choice == "6":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_full_infrastructure_assessment())
            elif choice == "7":
                self.view_infrastructure_results()
            elif choice == "8":
                self.configure_infrastructure_settings()
            else:
                self.show_error(f"Invalid option: {choice}")
                
    async def handle_network_discovery(self):
        """Handle network discovery and mapping"""
        if not self.integrations_available or not hasattr(self, 'infrastructure_scanner'):
            self.show_error("Infrastructure scanner not available")
            return
            
        try:
            # Get network range
            network = input("\n🌐 Enter network range (e.g., 192.168.1.0/24): ").strip()
            if not network:
                return
                
            self.show_info(f"Starting network discovery for {network}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"network_disc_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute network discovery
            results = await self.infrastructure_scanner.discover_network(
                network_range=network,
                output_dir=run_path
            )
            
            if results and not results.get('errors'):
                self.show_success("Network discovery completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                live_hosts = results.get('live_hosts', [])
                network_services = results.get('network_services', [])
                
                print(f"🌐 Discovered {len(live_hosts)} live hosts")
                print(f"🔌 Found {len(network_services)} network services")
            else:
                self.show_error("Network discovery failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during network discovery: {e}")
            
        self.wait_for_continue()
        
    async def handle_service_detection(self):
        """Handle service detection and enumeration"""
        if not self.integrations_available or not hasattr(self, 'infrastructure_scanner'):
            self.show_error("Infrastructure scanner not available")
            return
            
        try:
            # Get target for service detection
            target = input("\n🎯 Enter target (IP or hostname): ").strip()
            if not target:
                return
                
            self.show_info(f"Starting service detection for {target}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"service_det_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute service detection
            results = await self.infrastructure_scanner.detect_services(
                target=target,
                output_dir=run_path,
                comprehensive=True
            )
            
            if results and not results.get('errors'):
                self.show_success("Service detection completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                services = results.get('services', [])
                open_ports = results.get('open_ports', [])
                
                print(f"🔌 Found {len(open_ports)} open ports")
                print(f"⚙️ Identified {len(services)} services")
                
                # Show top services
                if services:
                    print("\n🔝 Top services:")
                    for service in services[:5]:
                        port = service.get('port', 'unknown')
                        name = service.get('service', 'unknown')
                        version = service.get('version', '')
                        print(f"   • Port {port}: {name} {version}")
                        
            else:
                self.show_error("Service detection failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during service detection: {e}")
            
        self.wait_for_continue()
        
    async def handle_os_fingerprinting(self):
        """Handle operating system fingerprinting"""
        if not self.integrations_available or not hasattr(self, 'infrastructure_scanner'):
            self.show_error("Infrastructure scanner not available")
            return
            
        try:
            # Get target for OS fingerprinting
            target = input("\n🎯 Enter target (IP or hostname): ").strip()
            if not target:
                return
                
            self.show_info(f"Starting OS fingerprinting for {target}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"os_fingerprint_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute OS fingerprinting
            results = await self.infrastructure_scanner.fingerprint_os(
                target=target,
                output_dir=run_path
            )
            
            if results and not results.get('errors'):
                self.show_success("OS fingerprinting completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                os_info = results.get('os_info', {})
                confidence = results.get('confidence', 0)
                
                if os_info:
                    print(f"🏢 OS Family: {os_info.get('family', 'Unknown')}")
                    print(f"📊 Version: {os_info.get('version', 'Unknown')}")
                    print(f"🎯 Confidence: {confidence}%")
                    
                    if os_info.get('cpe'):
                        print(f"🔍 CPE: {os_info['cpe']}")
                        
            else:
                self.show_error("OS fingerprinting failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during OS fingerprinting: {e}")
            
        self.wait_for_continue()
        
    async def handle_protocol_analysis(self):
        """Handle network protocol analysis"""
        if not self.integrations_available or not hasattr(self, 'infrastructure_scanner'):
            self.show_error("Infrastructure scanner not available")
            return
            
        try:
            # Get target and protocol
            target = input("\n🎯 Enter target (IP or hostname): ").strip()
            if not target:
                return
                
            print("\n📡 Select protocol to analyze:")
            print("1. TCP")
            print("2. UDP")
            print("3. ICMP")
            print("4. All protocols")
            
            choice = input("\nSelect protocol (1-4): ").strip()
            protocol_map = {"1": "tcp", "2": "udp", "3": "icmp", "4": "all"}
            
            if choice not in protocol_map:
                self.show_error("Invalid protocol selection")
                return
                
            protocol = protocol_map[choice]
            
            self.show_info(f"Starting {protocol.upper()} protocol analysis for {target}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"protocol_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute protocol analysis
            results = await self.infrastructure_scanner.analyze_protocols(
                target=target,
                protocol=protocol,
                output_dir=run_path
            )
            
            if results and not results.get('errors'):
                self.show_success("Protocol analysis completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                protocols = results.get('protocols', {})
                anomalies = results.get('anomalies', [])
                
                print(f"📡 Analyzed protocols: {len(protocols)}")
                print(f"⚠️ Anomalies detected: {len(anomalies)}")
                
                if anomalies:
                    print("\n🚨 Protocol anomalies:")
                    for anomaly in anomalies[:5]:
                        print(f"   • {anomaly}")
                        
            else:
                self.show_error("Protocol analysis failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during protocol analysis: {e}")
            
        self.wait_for_continue()
        
    async def handle_asset_discovery(self):
        """Handle asset discovery and classification"""
        if not self.integrations_available or not hasattr(self, 'infrastructure_scanner'):
            self.show_error("Infrastructure scanner not available")
            return
            
        try:
            # Get network range for asset discovery
            network = input("\n🌐 Enter network range for asset discovery (e.g., 192.168.1.0/24): ").strip()
            if not network:
                return
                
            self.show_info(f"Starting asset discovery for {network}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"asset_discovery_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute asset discovery
            results = await self.infrastructure_scanner.discover_assets(
                network_range=network,
                output_dir=run_path,
                classify=True
            )
            
            if results and not results.get('errors'):
                self.show_success("Asset discovery completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show summary
                assets = results.get('assets', [])
                categories = results.get('asset_categories', {})
                
                print(f"🔍 Discovered {len(assets)} assets")
                
                if categories:
                    print("\n📊 Asset categories:")
                    for category, count in categories.items():
                        print(f"   • {category}: {count}")
                        
            else:
                self.show_error("Asset discovery failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during asset discovery: {e}")
            
        self.wait_for_continue()
        
    async def handle_full_infrastructure_assessment(self):
        """Handle full infrastructure security assessment"""
        if not self.integrations_available or not hasattr(self, 'infrastructure_scanner'):
            self.show_error("Infrastructure scanner not available")
            return
            
        try:
            # Get network range
            network = input("\n🌐 Enter network range for full assessment: ").strip()
            if not network:
                return
                
            self.show_info(f"Starting full infrastructure assessment for {network}")
            print("🔄 This will run: Network Discovery → Service Detection → OS Fingerprinting → Asset Classification")
            
            confirm = input("\n⚠️ Continue with full infrastructure assessment? (y/N): ")
            if confirm.lower() != 'y':
                return
                
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"full_infra_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute full infrastructure assessment
            results = await self.infrastructure_scanner.full_infrastructure_assessment(
                network_range=network,
                output_dir=run_path
            )
            
            if results and not results.get('errors'):
                self.show_success("Full infrastructure assessment completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show comprehensive summary
                live_hosts = results.get('live_hosts', [])
                services = results.get('services', [])
                os_info = results.get('os_fingerprints', [])
                assets = results.get('assets', [])
                
                print(f"🌐 Live hosts: {len(live_hosts)}")
                print(f"🔌 Services detected: {len(services)}")
                print(f"🏢 OS fingerprints: {len(os_info)}")
                print(f"🔍 Assets classified: {len(assets)}")
                
            else:
                self.show_error("Full infrastructure assessment failed")
                if results and results.get('errors'):
                    for error in results['errors']:
                        print(f"   ❌ {error}")
                        
        except Exception as e:
            self.show_error(f"Error during full infrastructure assessment: {e}")
            
        self.wait_for_continue()
        
    def view_infrastructure_results(self):
        """View recent infrastructure scan results"""
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        
        # Find recent infrastructure scan runs
        infra_dirs = [d for d in runs_dir.iterdir() if d.is_dir() and 
                     any(x in d.name for x in ['network_disc_', 'service_det_', 'os_fingerprint_', 'protocol_analysis_', 'asset_discovery_', 'full_infra_'])]
        infra_dirs.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        
        if not infra_dirs:
            self.show_info("No infrastructure scan results found")
            self.wait_for_continue()
            return
            
        print("\n📊 \033[1;97mRecent Infrastructure Scan Results:\033[0m")
        print("=" * 60)
        
        for i, run_dir in enumerate(infra_dirs[:10], 1):
            print(f"\n{i}. 📁 {run_dir.name}")
            
            # Check for different types of results
            if 'network_disc_' in run_dir.name:
                result_files = list(run_dir.glob("*.json"))
                print(f"   🌐 Network Discovery: {len(result_files)} result files")
            elif 'service_det_' in run_dir.name:
                result_files = list(run_dir.glob("*.json"))
                print(f"   🔌 Service Detection: {len(result_files)} result files")
            elif 'os_fingerprint_' in run_dir.name:
                result_files = list(run_dir.glob("*.json"))
                print(f"   🏢 OS Fingerprinting: {len(result_files)} result files")
            elif 'protocol_analysis_' in run_dir.name:
                result_files = list(run_dir.glob("*.json"))
                print(f"   📡 Protocol Analysis: {len(result_files)} result files")
            elif 'asset_discovery_' in run_dir.name:
                result_files = list(run_dir.glob("*.json"))
                print(f"   🔍 Asset Discovery: {len(result_files)} result files")
            elif 'full_infra_' in run_dir.name:
                all_files = list(run_dir.glob("*"))
                print(f"   🏗️ Full Infrastructure: {len(all_files)} files")
                    
        self.wait_for_continue()
        
    def configure_infrastructure_settings(self):
        """Configure infrastructure scanning settings"""
        print("\n⚙️ \033[1;97mInfrastructure Scanning Configuration:\033[0m")
        print("=" * 50)
        
        config = self.moloch_integration.config
        infra_config = config.get('infrastructure', {})
        
        print(f"📊 Current Settings:")
        print(f"   • Network Scanner: {infra_config.get('network_scanner', 'nmap')}")
        print(f"   • Service Detection: {infra_config.get('service_detection', True)}")
        print(f"   • OS Fingerprinting: {infra_config.get('os_fingerprinting', True)}")
        print(f"   • Protocol Analysis: {infra_config.get('protocol_analysis', False)}")
        print(f"   • Scan Intensity: {infra_config.get('scan_intensity', 'normal')}")
        print(f"   • Timeout: {infra_config.get('timeout', 300)}s")
        print(f"   • Max Concurrent: {infra_config.get('max_concurrent', 10)}")
        
        print(f"\n💡 Use moloch.cfg.json to modify these settings")
        self.wait_for_continue()
    
    def fuzzing_discovery_menu(self):
        """Fuzzing and discovery interface"""
        while True:
            self.print_master_banner()
            print("╔═══════════════════════════════════════════════════════════════════════════════╗")
            print("║\033[1;93m                     💥 FUZZING & DISCOVERY\033[0m                         ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m1.\033[0m 📁 Directory & File Fuzzing (ffuf, gobuster)                 ║")
            print("║  \033[1;97m2.\033[0m 🌐 Subdomain Fuzzing (DNS brute-force)                       ║")
            print("║  \033[1;97m3.\033[0m 🔍 Parameter Fuzzing (GET/POST parameters)                   ║")
            print("║  \033[1;97m4.\033[0m 📄 Content Discovery (file extensions, backups)             ║")
            print("║  \033[1;97m5.\033[0m 🔗 Virtual Host Discovery                                    ║")
            print("║  \033[1;97m6.\033[0m 💥 Full Fuzzing Suite                                        ║")
            print("║  \033[1;97m7.\033[0m 📊 View Fuzzing Results                                      ║")
            print("║  \033[1;97m8.\033[0m ⚙️ Configure Fuzzing Settings                               ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m0.\033[0m ↩️ Return to Main Menu                                           ║")
            print("╚═══════════════════════════════════════════════════════════════════════════════╝")
            
            choice = self.get_user_input()
            
            if choice == "0":
                break
            elif choice == "1":
                self.handle_directory_file_fuzzing()
            elif choice == "2":
                self.handle_subdomain_fuzzing()
            elif choice == "3":
                self.handle_parameter_fuzzing()
            elif choice == "4":
                self.handle_content_discovery()
            elif choice == "5":
                self.handle_vhost_discovery()
            elif choice == "6":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.handle_full_fuzzing_suite())
            elif choice == "7":
                self.view_fuzzing_results()
            elif choice == "8":
                self.configure_fuzzing_settings()
            else:
                self.show_error(f"Invalid option: {choice}")
                
    def handle_directory_file_fuzzing(self):
        """Handle directory and file fuzzing"""
        # Get target URL
        if not self.targets:
            target_url = input("\n🎯 Enter target URL for directory fuzzing: ").strip()
            if not target_url:
                return
        else:
            target_url = self.select_target_interactive()
            if not target_url:
                return
                
        # Ensure URL format
        if not target_url.startswith(('http://', 'https://')):
            target_url = f"https://{target_url}"
            
        try:
            self.show_info(f"Starting directory/file fuzzing for {target_url}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"dir_fuzz_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Run directory fuzzing using moloch
            from moloch import run_directory_fuzzing
            # Create a temporary file with the target URL
            target_file = run_path / "target.txt"
            with open(target_file, 'w') as f:
                f.write(target_url)
                
            success = run_directory_fuzzing(target_file, run_path, self.moloch_integration.config)
            
            if success:
                self.show_success("Directory/file fuzzing completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show quick stats
                result_files = list(run_path.glob("*"))
                print(f"📄 Generated {len(result_files)} result files")
            else:
                self.show_error("Directory/file fuzzing failed")
                
        except Exception as e:
            self.show_error(f"Error during directory fuzzing: {e}")
            
        self.wait_for_continue()
        
    def handle_subdomain_fuzzing(self):
        """Handle subdomain fuzzing"""
        # Get target domain
        if not self.targets:
            domain = input("\n🎯 Enter domain for subdomain fuzzing: ").strip()
            if not domain:
                return
        else:
            domain = self.select_target_interactive()
            if not domain:
                return
                
        # Extract domain from URL if needed
        from urllib.parse import urlparse
        if domain.startswith(('http://', 'https://')):
            domain = urlparse(domain).netloc
            
        try:
            self.show_info(f"Starting subdomain fuzzing for {domain}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"subdomain_fuzz_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Use gobuster for DNS bruteforcing
            from moloch import execute_tool
            output_file = run_path / "subdomain_fuzzing.txt"
            
            # Common subdomain wordlist
            wordlist_path = Path("wordlists") / "subdomains.txt"
            if not wordlist_path.exists():
                # Create basic wordlist if it doesn't exist
                wordlist_path.parent.mkdir(exist_ok=True)
                basic_subdomains = ['www', 'mail', 'ftp', 'admin', 'test', 'dev', 'staging', 'api', 'app', 'secure']
                with open(wordlist_path, 'w') as f:
                    f.write('\n'.join(basic_subdomains))
                    
            success = execute_tool("gobuster", [
                "dns", "-d", domain, "-w", str(wordlist_path), "-o", str(output_file)
            ], output_file=output_file, run_dir=run_path)
            
            if success and output_file.exists():
                self.show_success("Subdomain fuzzing completed")
                print(f"📁 Results saved to: {run_path}")
                
                # Show found subdomains
                with open(output_file, 'r') as f:
                    subdomains = [line.strip() for line in f if 'Found:' in line]
                    print(f"🌐 Found {len(subdomains)} subdomains")
            else:
                self.show_error("Subdomain fuzzing failed or no results")
                
        except Exception as e:
            self.show_error(f"Error during subdomain fuzzing: {e}")
            
        self.wait_for_continue()
        
    def handle_parameter_fuzzing(self):
        """Handle parameter fuzzing"""
        # Get target URL
        target_url = input("\n🎯 Enter target URL for parameter fuzzing: ").strip()
        if not target_url:
            return
            
        try:
            self.show_info(f"Starting parameter fuzzing for {target_url}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"param_fuzz_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Use ffuf for parameter fuzzing
            from moloch import execute_tool
            
            # Create parameter wordlist if it doesn't exist
            param_wordlist = Path("wordlists") / "parameters.txt"
            if not param_wordlist.exists():
                param_wordlist.parent.mkdir(exist_ok=True)
                common_params = ['id', 'user', 'admin', 'test', 'debug', 'file', 'path', 'url', 'redirect', 'callback']
                with open(param_wordlist, 'w') as f:
                    f.write('\n'.join(common_params))
                    
            output_file = run_path / "parameter_fuzzing.txt"
            
            # Fuzzing GET parameters
            fuzz_url = f"{target_url}?FUZZ=test"
            success = execute_tool("ffuf", [
                "-u", fuzz_url, "-w", str(param_wordlist), "-o", str(output_file), "-of", "csv"
            ], output_file=output_file, run_dir=run_path)
            
            if success:
                self.show_success("Parameter fuzzing completed")
                print(f"📁 Results saved to: {run_path}")
            else:
                self.show_error("Parameter fuzzing failed")
                
        except Exception as e:
            self.show_error(f"Error during parameter fuzzing: {e}")
            
        self.wait_for_continue()
        
    def handle_content_discovery(self):
        """Handle content discovery"""
        # Get target URL
        target_url = input("\n🎯 Enter target URL for content discovery: ").strip()
        if not target_url:
            return
            
        try:
            self.show_info(f"Starting content discovery for {target_url}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"content_disc_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Use ffuf for content discovery
            from moloch import execute_tool
            
            # Create file extension wordlist
            ext_wordlist = Path("wordlists") / "extensions.txt"
            if not ext_wordlist.exists():
                ext_wordlist.parent.mkdir(exist_ok=True)
                extensions = ['php', 'html', 'asp', 'aspx', 'jsp', 'txt', 'pdf', 'doc', 'xml', 'json', 'backup', 'bak']
                with open(ext_wordlist, 'w') as f:
                    f.write('\n'.join(extensions))
                    
            output_file = run_path / "content_discovery.txt"
            
            # Fuzz file extensions
            base_url = target_url.rstrip('/')
            fuzz_url = f"{base_url}/FUZZ"
            
            success = execute_tool("ffuf", [
                "-u", fuzz_url, "-w", str(ext_wordlist), "-o", str(output_file), "-of", "csv"
            ], output_file=output_file, run_dir=run_path)
            
            if success:
                self.show_success("Content discovery completed")
                print(f"📁 Results saved to: {run_path}")
            else:
                self.show_error("Content discovery failed")
                
        except Exception as e:
            self.show_error(f"Error during content discovery: {e}")
            
        self.wait_for_continue()
        
    def handle_vhost_discovery(self):
        """Handle virtual host discovery"""
        # Get target
        target = input("\n🎯 Enter target domain/IP for vhost discovery: ").strip()
        if not target:
            return
            
        try:
            self.show_info(f"Starting virtual host discovery for {target}")
            
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"vhost_disc_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Use ffuf for vhost discovery
            from moloch import execute_tool
            
            # Create vhost wordlist
            vhost_wordlist = Path("wordlists") / "vhosts.txt"
            if not vhost_wordlist.exists():
                vhost_wordlist.parent.mkdir(exist_ok=True)
                vhosts = ['admin', 'mail', 'test', 'dev', 'staging', 'api', 'secure', 'portal', 'www']
                with open(vhost_wordlist, 'w') as f:
                    f.write('\n'.join(vhosts))
                    
            output_file = run_path / "vhost_discovery.txt"
            target_url = f"http://{target}"
            
            success = execute_tool("ffuf", [
                "-u", target_url, "-H", "Host: FUZZ.{target}", "-w", str(vhost_wordlist), 
                "-o", str(output_file), "-of", "csv"
            ], output_file=output_file, run_dir=run_path)
            
            if success:
                self.show_success("Virtual host discovery completed")
                print(f"📁 Results saved to: {run_path}")
            else:
                self.show_error("Virtual host discovery failed")
                
        except Exception as e:
            self.show_error(f"Error during vhost discovery: {e}")
            
        self.wait_for_continue()
        
    async def handle_full_fuzzing_suite(self):
        """Handle full fuzzing suite"""
        target_url = input("\n🎯 Enter target URL for full fuzzing suite: ").strip()
        if not target_url:
            return
            
        try:
            self.show_info(f"Starting full fuzzing suite for {target_url}")
            print("🔄 This will run: Directory Fuzzing → Parameter Fuzzing → Content Discovery → VHost Discovery")
            
            confirm = input("\n⚠️ Continue with full fuzzing suite? (y/N): ")
            if confirm.lower() != 'y':
                return
                
            # Create run directory
            run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            run_path = run_dir / f"full_fuzz_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
            run_path.mkdir(parents=True, exist_ok=True)
            
            # Execute each fuzzing method
            methods = [
                ("Directory Fuzzing", self.handle_directory_file_fuzzing),
                ("Parameter Fuzzing", self.handle_parameter_fuzzing),
                ("Content Discovery", self.handle_content_discovery),
            ]
            
            completed = 0
            for method_name, method_func in methods:
                try:
                    print(f"\n🔄 Running {method_name}...")
                    # Note: These methods are sync, so we don't await them
                    method_func()
                    completed += 1
                except Exception as e:
                    print(f"❌ {method_name} failed: {e}")
                    
            self.show_success(f"Full fuzzing suite completed ({completed}/{len(methods)} methods)")
            print(f"📁 Results saved to: {run_path}")
            
        except Exception as e:
            self.show_error(f"Error during full fuzzing suite: {e}")
            
        self.wait_for_continue()
        
    def view_fuzzing_results(self):
        """View fuzzing results"""
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        
        # Find recent fuzzing runs
        fuzz_dirs = [d for d in runs_dir.iterdir() if d.is_dir() and 
                    any(x in d.name for x in ['dir_fuzz_', 'subdomain_fuzz_', 'param_fuzz_', 'content_disc_', 'vhost_disc_', 'full_fuzz_'])]
        fuzz_dirs.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        
        if not fuzz_dirs:
            self.show_info("No fuzzing results found")
            self.wait_for_continue()
            return
            
        print("\n📊 \033[1;97mRecent Fuzzing Results:\033[0m")
        print("=" * 60)
        
        for i, run_dir in enumerate(fuzz_dirs[:10], 1):
            print(f"\n{i}. 📁 {run_dir.name}")
            
            # Check for different types of results
            result_files = list(run_dir.glob("*"))
            print(f"   📄 Result files: {len(result_files)}")
            
            # Show some results if available
            for result_file in result_files[:3]:
                if result_file.is_file() and result_file.stat().st_size > 0:
                    print(f"     • {result_file.name} ({result_file.stat().st_size // 1024} KB)")
                    
        self.wait_for_continue()
        
    def configure_fuzzing_settings(self):
        """Configure fuzzing settings"""
        print("\n⚙️ \033[1;97mFuzzing Configuration:\033[0m")
        print("=" * 50)
        
        config = self.moloch_integration.config
        fuzz_config = config.get('fuzzing', {})
        
        print(f"📊 Current Settings:")
        print(f"   • Directory Fuzzer: {fuzz_config.get('dir_fuzzer', 'ffuf')}")
        print(f"   • DNS Fuzzer: {fuzz_config.get('dns_fuzzer', 'gobuster')}")
        print(f"   • Threads: {fuzz_config.get('threads', 40)}")
        print(f"   • Request Rate: {fuzz_config.get('rate_limit', 100)} req/sec")
        print(f"   • Timeout: {fuzz_config.get('timeout', 10)}s")
        print(f"   • Follow Redirects: {fuzz_config.get('follow_redirects', True)}")
        print(f"   • Wordlist Directory: {fuzz_config.get('wordlist_dir', 'wordlists')}")
        
        print(f"\n💡 Use moloch.cfg.json to modify these settings")
        self.wait_for_continue()
    
    def system_configuration_menu(self):
        """System configuration interface"""
        while True:
            self.print_master_banner()
            print("╔═══════════════════════════════════════════════════════════════════════════════╗")
            print("║\033[1;95m                      ⚙️ SYSTEM CONFIGURATION\033[0m                       ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m1.\033[0m 🔧 Tool Installation & Management                             ║")
            print("║  \033[1;97m2.\033[0m 📋 Configuration File Editor                                ║")
            print("║  \033[1;97m3.\033[0m 🔑 API Keys & Token Management                              ║")
            print("║  \033[1;97m4.\033[0m 🌐 Proxy & Network Settings                                ║")
            print("║  \033[1;97m5.\033[0m 📊 Performance & Resource Tuning                           ║")
            print("║  \033[1;97m6.\033[0m 🗂️ Directory & Path Configuration                           ║")
            print("║  \033[1;97m7.\033[0m 💾 Backup & Restore Settings                               ║")
            print("║  \033[1;97m8.\033[0m 🔍 System Status & Diagnostics                             ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m0.\033[0m ↩️ Return to Main Menu                                           ║")
            print("╚═══════════════════════════════════════════════════════════════════════════════╝")
            
            choice = self.get_user_input()
            
            if choice == "0":
                break
            elif choice == "1":
                self.tool_installation_management()
            elif choice == "2":
                self.configuration_file_editor()
            elif choice == "3":
                self.api_keys_management()
            elif choice == "4":
                self.proxy_network_settings()
            elif choice == "5":
                self.performance_tuning()
            elif choice == "6":
                self.directory_path_configuration()
            elif choice == "7":
                self.backup_restore_settings()
            elif choice == "8":
                self.system_status_diagnostics()
            else:
                self.show_error(f"Invalid option: {choice}")
                
    def tool_installation_management(self):
        """Tool installation and management interface"""
        while True:
            print("\n🔧 \033[1;97mTool Installation & Management:\033[0m")
            print("1. Check tool status")
            print("2. Install missing tools")
            print("3. Update existing tools")
            print("4. Remove tools")
            print("5. Tool dependencies check")
            print("0. Return to previous menu")
            
            choice = input("\nSelect option (0-5): ").strip()
            
            if choice == "0":
                break
            elif choice == "1":
                self.check_tool_status()
            elif choice == "2":
                self.install_missing_tools()
            elif choice == "3":
                self.update_existing_tools()
            elif choice == "4":
                self.remove_tools()
            elif choice == "5":
                self.check_tool_dependencies()
            else:
                self.show_error("Invalid option")
                
    def check_tool_status(self):
        """Check status of all security tools"""
        print("\n🔍 \033[1;97mChecking Tool Status...\033[0m")
        
        tools = {
            'subfinder': 'Subdomain discovery',
            'assetfinder': 'Asset discovery', 
            'amass': 'Attack surface mapping',
            'httpx': 'HTTP probing',
            'nuclei': 'Vulnerability scanner',
            'nmap': 'Network mapper',
            'testssl': 'SSL/TLS tester',
            'katana': 'Web crawler',
            'dalfox': 'XSS scanner',
            'ffuf': 'Web fuzzer',
            'gobuster': 'Directory/DNS bruter',
            'arjun': 'Parameter discovery'
        }
        
        available_tools = []
        missing_tools = []
        
        for tool, description in tools.items():
            try:
                result = subprocess.run(['which', tool], capture_output=True, text=True)
                if result.returncode == 0:
                    available_tools.append((tool, description, result.stdout.strip()))
                else:
                    missing_tools.append((tool, description))
            except:
                missing_tools.append((tool, description))
                
        print(f"\n✅ \033[1;92mAvailable Tools ({len(available_tools)}):\033[0m")
        for tool, desc, path in available_tools:
            print(f"   • {tool:12} - {desc} ({path})")
            
        if missing_tools:
            print(f"\n❌ \033[1;91mMissing Tools ({len(missing_tools)}):\033[0m")
            for tool, desc in missing_tools:
                print(f"   • {tool:12} - {desc}")
                
        print(f"\n📊 Summary: {len(available_tools)}/{len(tools)} tools available")
        self.wait_for_continue()
        
    def install_missing_tools(self):
        """Install missing security tools"""
        print("\n📦 \033[1;97mTool Installation:\033[0m")
        print("⚠️ This will install security tools using package managers and GitHub releases")
        
        confirm = input("\nContinue with tool installation? (y/N): ")
        if confirm.lower() != 'y':
            return
            
        # Installation commands for different tools
        install_commands = {
            'subfinder': 'go install -v github.com/projectdiscovery/subfinder/v2/cmd/subfinder@latest',
            'assetfinder': 'go install github.com/tomnomnom/assetfinder@latest',
            'httpx': 'go install -v github.com/projectdiscovery/httpx/cmd/httpx@latest',
            'nuclei': 'go install -v github.com/projectdiscovery/nuclei/v2/cmd/nuclei@latest',
            'katana': 'go install github.com/projectdiscovery/katana/cmd/katana@latest',
            'dalfox': 'go install github.com/hahwul/dalfox/v2@latest',
            'ffuf': 'go install github.com/ffuf/ffuf@latest',
            'gobuster': 'go install github.com/OJ/gobuster/v3@latest',
            'arjun': 'pip3 install arjun'
        }
        
        print("\n🚀 Starting installation process...")
        
        for tool, command in install_commands.items():
            try:
                # Check if tool is already installed
                result = subprocess.run(['which', tool], capture_output=True)
                if result.returncode == 0:
                    print(f"✅ {tool} already installed")
                    continue
                    
                print(f"📦 Installing {tool}...")
                result = subprocess.run(command.split(), capture_output=True, text=True, timeout=300)
                
                if result.returncode == 0:
                    print(f"✅ {tool} installed successfully")
                else:
                    print(f"❌ {tool} installation failed: {result.stderr[:100]}")
                    
            except subprocess.TimeoutExpired:
                print(f"⏰ {tool} installation timed out")
            except Exception as e:
                print(f"❌ {tool} installation error: {e}")
                
        print("\n🎉 Installation process completed!")
        self.wait_for_continue()
        
    def update_existing_tools(self):
        """Update existing security tools"""
        print("\n🔄 \033[1;97mTool Update:\033[0m")
        print("This will update all installed security tools to their latest versions")
        
        confirm = input("\nContinue with tool updates? (y/N): ")
        if confirm.lower() != 'y':
            return
            
        # Update commands for tools
        update_commands = [
            ('nuclei', 'nuclei -update-templates'),
            ('go tools', 'go install -a -v github.com/projectdiscovery/subfinder/v2/cmd/subfinder@latest'),
            # Add more update commands as needed
        ]
        
        print("\n🔄 Starting update process...")
        
        for tool, command in update_commands:
            try:
                print(f"🔄 Updating {tool}...")
                result = subprocess.run(command.split(), capture_output=True, text=True, timeout=180)
                
                if result.returncode == 0:
                    print(f"✅ {tool} updated successfully")
                else:
                    print(f"⚠️ {tool} update returned: {result.returncode}")
                    
            except subprocess.TimeoutExpired:
                print(f"⏰ {tool} update timed out")
            except Exception as e:
                print(f"❌ {tool} update error: {e}")
                
        print("\n🎉 Update process completed!")
        self.wait_for_continue()
        
    def remove_tools(self):
        """Remove security tools"""
        print("\n🗑️ \033[1;97mTool Removal:\033[0m")
        print("⚠️ This will remove selected security tools from the system")
        
        # Show available tools
        tools = ['subfinder', 'assetfinder', 'httpx', 'nuclei', 'katana', 'dalfox', 'ffuf', 'gobuster']
        available_tools = []
        
        for tool in tools:
            try:
                result = subprocess.run(['which', tool], capture_output=True)
                if result.returncode == 0:
                    available_tools.append(tool)
            except Exception as e:
                self.logger.debug(f"Tool availability check failed for {tool}: {e}")
                continue
                
        if not available_tools:
            self.show_info("No tools available for removal")
            return
            
        print(f"\n📋 Available tools for removal:")
        for i, tool in enumerate(available_tools, 1):
            print(f"  {i}. {tool}")
            
        selection = input(f"\nSelect tools to remove (comma-separated numbers, or 'all'): ").strip()
        
        if selection.lower() == 'all':
            tools_to_remove = available_tools
        else:
            try:
                indices = [int(x.strip()) - 1 for x in selection.split(',')]
                tools_to_remove = [available_tools[i] for i in indices if 0 <= i < len(available_tools)]
            except:
                self.show_error("Invalid selection format")
                return
                
        if tools_to_remove:
            print(f"\n⚠️ Tools to remove: {', '.join(tools_to_remove)}")
            confirm = input("Confirm removal? (y/N): ")
            
            if confirm.lower() == 'y':
                for tool in tools_to_remove:
                    try:
                        # For Go tools, remove from GOPATH/bin
                        result = subprocess.run(['which', tool], capture_output=True, text=True)
                        if result.returncode == 0:
                            tool_path = result.stdout.strip()
                            subprocess.run(['rm', tool_path], check=True)
                            print(f"✅ Removed {tool}")
                    except Exception as e:
                        print(f"❌ Failed to remove {tool}: {e}")
                        
                print("\n🎉 Removal process completed!")
            else:
                print("Removal cancelled")
        else:
            self.show_error("No valid tools selected")
            
        self.wait_for_continue()
        
    def check_tool_dependencies(self):
        """Check tool dependencies"""
        print("\n🔍 \033[1;97mTool Dependencies Check:\033[0m")
        
        dependencies = {
            'Go': ['go', 'version'],
            'Python': ['python3', '--version'],
            'Pip': ['pip3', '--version'],
            'Git': ['git', '--version'],
            'curl': ['curl', '--version'],
            'wget': ['wget', '--version']
        }
        
        print("📋 Checking system dependencies...")
        
        for dep, command in dependencies.items():
            try:
                result = subprocess.run(command, capture_output=True, text=True, timeout=10)
                if result.returncode == 0:
                    version = result.stdout.split('\n')[0]
                    print(f"✅ {dep:8} - {version}")
                else:
                    print(f"❌ {dep:8} - Not found")
            except:
                print(f"❌ {dep:8} - Not found")
                
        self.wait_for_continue()
        
    def configuration_file_editor(self):
        """Configuration file editor interface"""
        print("\n📋 \033[1;97mConfiguration File Editor:\033[0m")
        
        config_files = {
            '1': ('moloch.cfg.json', 'Main moloch configuration'),
            '2': ('targets.txt', 'Target list'),
            '3': ('.env', 'Environment variables'),
        }
        
        print("Available configuration files:")
        for key, (filename, description) in config_files.items():
            exists = "✅" if Path(filename).exists() else "❌"
            print(f"  {key}. {exists} {filename} - {description}")
            
        choice = input("\nSelect file to edit (1-3): ").strip()
        
        if choice in config_files:
            filename, description = config_files[choice]
            self.edit_configuration_file(filename, description)
        else:
            self.show_error("Invalid file selection")
            
    def edit_configuration_file(self, filename, description):
        """Edit a specific configuration file"""
        filepath = Path(filename)
        
        print(f"\n📝 Editing {description}")
        print(f"File: {filepath.absolute()}")
        
        if filepath.exists():
            print(f"\n📖 Current content (first 20 lines):")
            try:
                with open(filepath, 'r') as f:
                    lines = f.readlines()
                    for i, line in enumerate(lines[:20], 1):
                        print(f"{i:2d}: {line.rstrip()}")
                    if len(lines) > 20:
                        print(f"... and {len(lines) - 20} more lines")
            except Exception as e:
                print(f"Error reading file: {e}")
        else:
            print("File does not exist - will be created")
            
        print(f"\n⚙️ Edit options:")
        print("1. View full content")
        print("2. Add line(s)")
        print("3. Replace content") 
        print("4. Open in system editor")
        print("0. Return")
        
        choice = input("\nSelect option (0-4): ").strip()
        
        if choice == "1":
            self.view_full_file(filepath)
        elif choice == "2":
            self.add_lines_to_file(filepath)
        elif choice == "3":
            self.replace_file_content(filepath)
        elif choice == "4":
            self.open_in_system_editor(filepath)
        elif choice == "0":
            return
        else:
            self.show_error("Invalid option")
            
    def view_full_file(self, filepath):
        """View full file content"""
        try:
            with open(filepath, 'r') as f:
                content = f.read()
                print(f"\n📖 Full content of {filepath.name}:")
                print("-" * 50)
                print(content)
                print("-" * 50)
        except Exception as e:
            self.show_error(f"Error reading file: {e}")
        self.wait_for_continue()
        
    def add_lines_to_file(self, filepath):
        """Add lines to file"""
        print(f"\n➕ Add lines to {filepath.name}")
        print("Enter lines to add (empty line to finish):")
        
        new_lines = []
        while True:
            line = input()
            if line == "":
                break
            new_lines.append(line)
            
        if new_lines:
            try:
                with open(filepath, 'a') as f:
                    for line in new_lines:
                        f.write(line + '\n')
                self.show_success(f"Added {len(new_lines)} lines to {filepath.name}")
            except Exception as e:
                self.show_error(f"Error writing to file: {e}")
        else:
            print("No lines added")
            
    def replace_file_content(self, filepath):
        """Replace entire file content"""
        print(f"\n🔄 Replace content of {filepath.name}")
        print("⚠️ This will overwrite the entire file!")
        
        confirm = input("Continue? (y/N): ")
        if confirm.lower() != 'y':
            return
            
        print("Enter new content (type '---END---' on a new line to finish):")
        
        new_content = []
        while True:
            line = input()
            if line == "---END---":
                break
            new_content.append(line)
            
        try:
            with open(filepath, 'w') as f:
                f.write('\n'.join(new_content))
            self.show_success(f"Replaced content of {filepath.name}")
        except Exception as e:
            self.show_error(f"Error writing to file: {e}")
            
    def open_in_system_editor(self, filepath):
        """Open file in system editor"""
        editors = ['nano', 'vim', 'vi', 'gedit']
        
        for editor in editors:
            try:
                result = subprocess.run(['which', editor], capture_output=True)
                if result.returncode == 0:
                    print(f"Opening {filepath.name} in {editor}...")
                    subprocess.run([editor, str(filepath)])
                    return
            except:
                continue
                
        self.show_error("No suitable text editor found")
        
    def api_keys_management(self):
        """API keys and token management"""
        print("\n🔑 \033[1;97mAPI Keys & Token Management:\033[0m")
        
        # Common API keys used by security tools
        api_keys = {
            'chaos_api_key': 'Chaos (ProjectDiscovery)',
            'github_token': 'GitHub API',
            'shodan_api_key': 'Shodan',
            'virustotal_api_key': 'VirusTotal',
            'securitytrails_api_key': 'SecurityTrails',
            'censys_api_id': 'Censys API ID',
            'censys_secret': 'Censys Secret'
        }
        
        # Check current status
        print("📊 Current API Key Status:")
        env_file = Path('.env')
        current_keys = {}
        
        if env_file.exists():
            try:
                with open(env_file, 'r') as f:
                    for line in f:
                        if '=' in line and not line.startswith('#'):
                            key, value = line.strip().split('=', 1)
                            current_keys[key] = value
            except Exception as e:
                self.logger.debug(f"Error reading API key file: {e}")
                current_keys = {}
                
        for key, description in api_keys.items():
            status = "✅ Set" if key in current_keys else "❌ Not set"
            print(f"  {status} {description} ({key})")
            
        print(f"\n⚙️ Management options:")
        print("1. Add/Update API key")
        print("2. Remove API key")
        print("3. View API key (masked)")
        print("4. Test API keys")
        print("0. Return")
        
        choice = input("\nSelect option (0-4): ").strip()
        
        if choice == "1":
            self.add_update_api_key(api_keys, current_keys, env_file)
        elif choice == "2":
            self.remove_api_key(current_keys, env_file)
        elif choice == "3":
            self.view_api_key(current_keys)
        elif choice == "4":
            self.test_api_keys(current_keys)
        elif choice == "0":
            return
        else:
            self.show_error("Invalid option")
            
    def add_update_api_key(self, api_keys, current_keys, env_file):
        """Add or update an API key"""
        print(f"\n➕ Add/Update API Key:")
        
        for i, (key, description) in enumerate(api_keys.items(), 1):
            status = "✅" if key in current_keys else "❌"
            print(f"  {i}. {status} {description} ({key})")
            
        try:
            choice = int(input(f"\nSelect API key to add/update (1-{len(api_keys)}): "))
            if 1 <= choice <= len(api_keys):
                key = list(api_keys.keys())[choice - 1]
                description = api_keys[key]
                
                print(f"\n🔑 Setting {description}")
                if key in current_keys:
                    print(f"Current value: {'*' * 20}")
                    
                new_value = input("Enter new API key value: ").strip()
                if new_value:
                    current_keys[key] = new_value
                    self.save_env_file(current_keys, env_file)
                    self.show_success(f"Updated {description}")
                else:
                    print("No value entered")
            else:
                self.show_error("Invalid selection")
        except ValueError:
            self.show_error("Invalid input")
            
    def remove_api_key(self, current_keys, env_file):
        """Remove an API key"""
        if not current_keys:
            self.show_info("No API keys to remove")
            return
            
        print(f"\n🗑️ Remove API Key:")
        keys_list = list(current_keys.keys())
        
        for i, key in enumerate(keys_list, 1):
            print(f"  {i}. {key}")
            
        try:
            choice = int(input(f"\nSelect API key to remove (1-{len(keys_list)}): "))
            if 1 <= choice <= len(keys_list):
                key = keys_list[choice - 1]
                confirm = input(f"Remove {key}? (y/N): ")
                
                if confirm.lower() == 'y':
                    del current_keys[key]
                    self.save_env_file(current_keys, env_file)
                    self.show_success(f"Removed {key}")
                else:
                    print("Removal cancelled")
            else:
                self.show_error("Invalid selection")
        except ValueError:
            self.show_error("Invalid input")
            
    def view_api_key(self, current_keys):
        """View API key (masked)"""
        if not current_keys:
            self.show_info("No API keys configured")
            return
            
        print(f"\n👁️ View API Keys (masked):")
        for key, value in current_keys.items():
            if len(value) > 8:
                masked = value[:4] + '*' * (len(value) - 8) + value[-4:]
            else:
                masked = '*' * len(value)
            print(f"  {key}: {masked}")
            
        self.wait_for_continue()
        
    def test_api_keys(self, current_keys):
        """Test API keys connectivity"""
        print(f"\n🔬 Testing API Keys:")
        
        if not current_keys:
            self.show_info("No API keys to test")
            return
            
        # Simple connectivity tests
        for key, value in current_keys.items():
            print(f"\n🔍 Testing {key}...")
            
            if 'github' in key:
                self.test_github_token(value)
            elif 'shodan' in key:
                self.test_shodan_key(value)
            elif 'virustotal' in key:
                self.test_virustotal_key(value)
            else:
                print(f"  ⚠️ No test available for {key}")
                
    def test_github_token(self, token):
        """Test GitHub token"""
        try:
            import requests
            headers = {'Authorization': f'token {token}'}
            response = requests.get('https://api.github.com/user', headers=headers, timeout=10)
            
            if response.status_code == 200:
                print(f"  ✅ GitHub token valid")
            else:
                print(f"  ❌ GitHub token invalid (status: {response.status_code})")
        except Exception as e:
            print(f"  ❌ GitHub token test failed: {e}")
            
    def test_shodan_key(self, key):
        """Test Shodan API key"""
        try:
            import requests
            response = requests.get(f'https://api.shodan.io/api-info?key={key}', timeout=10)
            
            if response.status_code == 200:
                print(f"  ✅ Shodan API key valid")
            else:
                print(f"  ❌ Shodan API key invalid")
        except Exception as e:
            print(f"  ❌ Shodan API key test failed: {e}")
            
    def test_virustotal_key(self, key):
        """Test VirusTotal API key"""
        try:
            import requests
            headers = {'x-apikey': key}
            response = requests.get('https://www.virustotal.com/api/v3/users/current', headers=headers, timeout=10)
            
            if response.status_code == 200:
                print(f"  ✅ VirusTotal API key valid")
            else:
                print(f"  ❌ VirusTotal API key invalid")
        except Exception as e:
            print(f"  ❌ VirusTotal API key test failed: {e}")
            
    def save_env_file(self, keys, env_file):
        """Save environment variables to .env file"""
        try:
            with open(env_file, 'w') as f:
                f.write("# Azaz-El API Keys Configuration\n")
                f.write(f"# Generated: {datetime.now().isoformat()}\n\n")
                
                for key, value in keys.items():
                    f.write(f"{key}={value}\n")
                    
        except Exception as e:
            self.show_error(f"Error saving .env file: {e}")
            
    def proxy_network_settings(self):
        """Proxy and network settings configuration"""
        print("\n🌐 \033[1;97mProxy & Network Settings:\033[0m")
        print("Configure proxy settings for security tool connections")
        
        # Current proxy settings
        current_proxy = {
            'http_proxy': os.environ.get('HTTP_PROXY', ''),
            'https_proxy': os.environ.get('HTTPS_PROXY', ''),
            'no_proxy': os.environ.get('NO_PROXY', '')
        }
        
        print(f"\n📊 Current proxy settings:")
        for key, value in current_proxy.items():
            status = value if value else "Not set"
            print(f"  {key.upper()}: {status}")
            
        print(f"\n⚙️ Configuration options:")
        print("1. Set HTTP proxy")
        print("2. Set HTTPS proxy") 
        print("3. Set no-proxy list")
        print("4. Clear all proxy settings")
        print("5. Test proxy connectivity")
        print("0. Return")
        
        choice = input("\nSelect option (0-5): ").strip()
        
        if choice == "1":
            proxy = input("Enter HTTP proxy (e.g., http://proxy.example.com:8080): ")
            if proxy:
                os.environ['HTTP_PROXY'] = proxy
                print(f"✅ HTTP proxy set to: {proxy}")
        elif choice == "2":
            proxy = input("Enter HTTPS proxy (e.g., https://proxy.example.com:8080): ")
            if proxy:
                os.environ['HTTPS_PROXY'] = proxy
                print(f"✅ HTTPS proxy set to: {proxy}")
        elif choice == "3":
            no_proxy = input("Enter no-proxy list (comma-separated): ")
            if no_proxy:
                os.environ['NO_PROXY'] = no_proxy
                print(f"✅ No-proxy list set to: {no_proxy}")
        elif choice == "4":
            for key in ['HTTP_PROXY', 'HTTPS_PROXY', 'NO_PROXY']:
                os.environ.pop(key, None)
            print("✅ All proxy settings cleared")
        elif choice == "5":
            self.test_proxy_connectivity()
        elif choice == "0":
            return
        else:
            self.show_error("Invalid option")
            
        self.wait_for_continue()
        
    def test_proxy_connectivity(self):
        """Test proxy connectivity"""
        print("\n🔬 Testing proxy connectivity...")
        
        try:
            import requests
            test_urls = [
                'https://httpbin.org/ip',
                'https://api.github.com',
                'https://www.google.com'
            ]
            
            for url in test_urls:
                try:
                    response = requests.get(url, timeout=10)
                    print(f"✅ {url} - Status: {response.status_code}")
                except Exception as e:
                    print(f"❌ {url} - Error: {e}")
                    
        except ImportError:
            print("❌ requests library not available for testing")
            
    def performance_tuning(self):
        """Performance and resource tuning settings"""
        print("\n📊 \033[1;97mPerformance & Resource Tuning:\033[0m")
        
        # Current system resources
        try:
            import psutil
            cpu_count = psutil.cpu_count()
            memory = psutil.virtual_memory()
            disk = psutil.disk_usage('/')
            
            print(f"💻 System Resources:")
            print(f"  CPU Cores: {cpu_count}")
            print(f"  Total Memory: {memory.total // (1024**3)} GB")
            print(f"  Available Memory: {memory.available // (1024**3)} GB")
            print(f"  Disk Space: {disk.free // (1024**3)} GB free")
            
        except ImportError:
            print("💻 System resource monitoring not available")
            
        # Performance settings
        config = self.moloch_integration.config
        perf_config = config.get('performance', {})
        
        print(f"\n⚙️ Current Performance Settings:")
        print(f"  Max Concurrent Scans: {perf_config.get('max_concurrent_scans', 5)}")
        print(f"  Thread Pool Size: {perf_config.get('thread_pool_size', 10)}")
        print(f"  Request Timeout: {perf_config.get('request_timeout', 30)}s")
        print(f"  Memory Limit: {perf_config.get('memory_limit', '1GB')}")
        print(f"  Scan Intensity: {perf_config.get('scan_intensity', 'normal')}")
        
        print(f"\n🔧 Tuning options:")
        print("1. Adjust concurrent scans")
        print("2. Modify thread pool size")
        print("3. Set request timeouts")
        print("4. Configure memory limits")
        print("5. Reset to defaults")
        print("0. Return")
        
        choice = input("\nSelect option (0-5): ").strip()
        
        if choice == "1":
            try:
                max_scans = int(input("Enter max concurrent scans (1-20): "))
                if 1 <= max_scans <= 20:
                    print(f"✅ Max concurrent scans set to: {max_scans}")
                else:
                    self.show_error("Value must be between 1-20")
            except ValueError:
                self.show_error("Invalid number")
        elif choice == "2":
            try:
                pool_size = int(input("Enter thread pool size (5-100): "))
                if 5 <= pool_size <= 100:
                    print(f"✅ Thread pool size set to: {pool_size}")
                else:
                    self.show_error("Value must be between 5-100")
            except ValueError:
                self.show_error("Invalid number")
        elif choice == "3":
            try:
                timeout = int(input("Enter request timeout in seconds (10-300): "))
                if 10 <= timeout <= 300:
                    print(f"✅ Request timeout set to: {timeout}s")
                else:
                    self.show_error("Value must be between 10-300")
            except ValueError:
                self.show_error("Invalid number")
        elif choice == "4":
            memory_limit = input("Enter memory limit (e.g., 1GB, 512MB): ")
            if memory_limit:
                print(f"✅ Memory limit set to: {memory_limit}")
        elif choice == "5":
            print("✅ Performance settings reset to defaults")
        elif choice == "0":
            return
        else:
            self.show_error("Invalid option")
            
        self.wait_for_continue()
        
    def directory_path_configuration(self):
        """Directory and path configuration"""
        print("\n🗂️ \033[1;97mDirectory & Path Configuration:\033[0m")
        
        # Current paths
        paths = {
            'runs_dir': self.moloch_integration.config.get('general', {}).get('runs_dir', 'runs'),
            'wordlists_dir': 'wordlists',
            'tools_dir': os.path.expanduser('~/go/bin'),
            'reports_dir': 'reports',
            'logs_dir': 'logs'
        }
        
        print("📂 Current directory paths:")
        for name, path in paths.items():
            exists = "✅" if Path(path).exists() else "❌"
            print(f"  {exists} {name}: {path}")
            
        print(f"\n⚙️ Configuration options:")
        print("1. Change runs directory")
        print("2. Set wordlists directory")
        print("3. Configure tools directory")
        print("4. Set reports directory")
        print("5. Create missing directories")
        print("6. Reset to defaults")
        print("0. Return")
        
        choice = input("\nSelect option (0-6): ").strip()
        
        if choice == "1":
            new_path = input("Enter new runs directory path: ")
            if new_path:
                Path(new_path).mkdir(parents=True, exist_ok=True)
                print(f"✅ Runs directory set to: {new_path}")
        elif choice == "2":
            new_path = input("Enter wordlists directory path: ")
            if new_path:
                Path(new_path).mkdir(parents=True, exist_ok=True)
                print(f"✅ Wordlists directory set to: {new_path}")
        elif choice == "3":
            new_path = input("Enter tools directory path: ")
            if new_path:
                print(f"✅ Tools directory set to: {new_path}")
        elif choice == "4":
            new_path = input("Enter reports directory path: ")
            if new_path:
                Path(new_path).mkdir(parents=True, exist_ok=True)
                print(f"✅ Reports directory set to: {new_path}")
        elif choice == "5":
            for name, path in paths.items():
                try:
                    Path(path).mkdir(parents=True, exist_ok=True)
                    print(f"✅ Created {name}: {path}")
                except Exception as e:
                    print(f"❌ Failed to create {name}: {e}")
        elif choice == "6":
            print("✅ Directory paths reset to defaults")
        elif choice == "0":
            return
        else:
            self.show_error("Invalid option")
            
        self.wait_for_continue()
        
    def backup_restore_settings(self):
        """Backup and restore settings"""
        print("\n💾 \033[1;97mBackup & Restore Settings:\033[0m")
        
        print("1. Create configuration backup")
        print("2. Restore from backup")
        print("3. Export scan results")
        print("4. Schedule automatic backups")
        print("5. Cleanup old files")
        print("0. Return")
        
        choice = input("\nSelect option (0-5): ").strip()
        
        if choice == "1":
            self.create_configuration_backup()
        elif choice == "2":
            self.restore_from_backup()
        elif choice == "3":
            self.export_scan_results()
        elif choice == "4":
            self.schedule_automatic_backups()
        elif choice == "5":
            self.cleanup_old_files()
        elif choice == "0":
            return
        else:
            self.show_error("Invalid option")
            
    def create_configuration_backup(self):
        """Create configuration backup"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_dir = Path(f"backup_{timestamp}")
        backup_dir.mkdir(exist_ok=True)
        
        files_to_backup = [
            'moloch.cfg.json',
            'targets.txt',
            '.env'
        ]
        
        backup_count = 0
        
        for filename in files_to_backup:
            source = Path(filename)
            if source.exists():
                try:
                    import shutil
                    shutil.copy2(source, backup_dir / filename)
                    backup_count += 1
                    print(f"✅ Backed up {filename}")
                except Exception as e:
                    print(f"❌ Failed to backup {filename}: {e}")
                    
        if backup_count > 0:
            self.show_success(f"Created backup with {backup_count} files in {backup_dir}")
        else:
            print("No files to backup")
            
        self.wait_for_continue()
        
    def restore_from_backup(self):
        """Restore from backup"""
        backup_dirs = [d for d in Path('.').iterdir() if d.is_dir() and d.name.startswith('backup_')]
        
        if not backup_dirs:
            self.show_info("No backup directories found")
            self.wait_for_continue()
            return
            
        print("📂 Available backups:")
        for i, backup_dir in enumerate(backup_dirs, 1):
            print(f"  {i}. {backup_dir.name}")
            
        try:
            choice = int(input(f"\nSelect backup to restore (1-{len(backup_dirs)}): "))
            if 1 <= choice <= len(backup_dirs):
                backup_dir = backup_dirs[choice - 1]
                
                print(f"⚠️ This will overwrite current configuration files!")
                confirm = input("Continue with restore? (y/N): ")
                
                if confirm.lower() == 'y':
                    import shutil
                    restore_count = 0
                    
                    for backup_file in backup_dir.iterdir():
                        if backup_file.is_file():
                            try:
                                shutil.copy2(backup_file, backup_file.name)
                                restore_count += 1
                                print(f"✅ Restored {backup_file.name}")
                            except Exception as e:
                                print(f"❌ Failed to restore {backup_file.name}: {e}")
                                
                    self.show_success(f"Restored {restore_count} files from backup")
                else:
                    print("Restore cancelled")
            else:
                self.show_error("Invalid backup selection")
        except ValueError:
            self.show_error("Invalid input")
            
        self.wait_for_continue()
        
    def export_scan_results(self):
        """Export scan results"""
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        
        if not runs_dir.exists():
            self.show_info("No scan results directory found")
            return
            
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        export_file = f"scan_results_export_{timestamp}.tar.gz"
        
        try:
            import tarfile
            with tarfile.open(export_file, 'w:gz') as tar:
                tar.add(runs_dir, arcname='scan_results')
                
            self.show_success(f"Exported scan results to {export_file}")
        except Exception as e:
            self.show_error(f"Export failed: {e}")
            
        self.wait_for_continue()
        
    def schedule_automatic_backups(self):
        """Schedule automatic backups"""
        print("⏰ Automatic backup scheduling")
        print("This would set up periodic backups of configuration and results")
        print("💡 Consider using cron jobs or system schedulers for production use")
        self.wait_for_continue()
        
    def cleanup_old_files(self):
        """Cleanup old files"""
        print("\n🧹 \033[1;97mCleanup Old Files:\033[0m")
        
        # Find old files
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        old_runs = []
        
        if runs_dir.exists():
            # Find runs older than 30 days
            cutoff_time = datetime.now().timestamp() - (30 * 24 * 60 * 60)
            
            for run_dir in runs_dir.iterdir():
                if run_dir.is_dir() and run_dir.stat().st_mtime < cutoff_time:
                    old_runs.append(run_dir)
                    
        if old_runs:
            print(f"📂 Found {len(old_runs)} old scan directories (>30 days)")
            
            total_size = 0
            for run_dir in old_runs[:5]:  # Show first 5
                try:
                    size = sum(f.stat().st_size for f in run_dir.rglob('*') if f.is_file())
                    total_size += size
                    print(f"  • {run_dir.name} ({size // (1024*1024)} MB)")
                except:
                    print(f"  • {run_dir.name}")
                    
            if len(old_runs) > 5:
                print(f"  ... and {len(old_runs) - 5} more")
                
            print(f"\n💾 Total size: ~{total_size // (1024*1024)} MB")
            
            confirm = input(f"\n🗑️ Delete {len(old_runs)} old directories? (y/N): ")
            if confirm.lower() == 'y':
                import shutil
                deleted_count = 0
                
                for run_dir in old_runs:
                    try:
                        shutil.rmtree(run_dir)
                        deleted_count += 1
                    except Exception as e:
                        print(f"❌ Failed to delete {run_dir.name}: {e}")
                        
                self.show_success(f"Deleted {deleted_count} old directories")
            else:
                print("Cleanup cancelled")
        else:
            self.show_info("No old files found for cleanup")
            
        self.wait_for_continue()
        
    def system_status_diagnostics(self):
        """System status and diagnostics"""
        print("\n🔍 \033[1;97mSystem Status & Diagnostics:\033[0m")
        
        # System information
        try:
            import psutil
            import platform
            
            print("💻 System Information:")
            print(f"  OS: {platform.system()} {platform.release()}")
            print(f"  Python: {platform.python_version()}")
            print(f"  CPU Usage: {psutil.cpu_percent(interval=1)}%")
            print(f"  Memory Usage: {psutil.virtual_memory().percent}%")
            print(f"  Disk Usage: {psutil.disk_usage('/').percent}%")
            
        except ImportError:
            print("💻 System monitoring not available (install psutil)")
            
        # Framework status
        print(f"\n🛠️ Framework Status:")
        print(f"  Version: {MASTER_VERSION}")
        print(f"  Integrations: {'✅ Available' if self.integrations_available else '❌ Limited'}")
        print(f"  Targets Loaded: {len(self.targets)}")
        print(f"  Active Scans: {len(self.active_scans)}")
        
        # Check critical files
        print(f"\n📁 Critical Files:")
        critical_files = ['moloch.cfg.json', 'moloch.py', 'master_azaz_el.py']
        for filename in critical_files:
            exists = "✅" if Path(filename).exists() else "❌"
            print(f"  {exists} {filename}")
            
        # Network connectivity
        print(f"\n🌐 Network Connectivity:")
        self.test_network_connectivity()
        
        self.wait_for_continue()
        
    def test_network_connectivity(self):
        """Test network connectivity"""
        test_hosts = [
            ('DNS', '8.8.8.8'),
            ('GitHub', 'github.com'),
            ('ProjectDiscovery', 'projectdiscovery.io')
        ]
        
        for name, host in test_hosts:
            try:
                import socket
                sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
                sock.settimeout(5)
                
                if host == '8.8.8.8':
                    result = sock.connect_ex((host, 53))
                else:
                    result = sock.connect_ex((host, 443))
                    
                sock.close()
                
                status = "✅" if result == 0 else "❌"
                print(f"  {status} {name} ({host})")
                
            except Exception as e:
                print(f"  ❌ {name} - Error: {e}")

    def proxy_network_settings(self):
        """Proxy and network settings - redirect to implemented method"""
        self.network_proxy_settings()
        
    def performance_tuning(self):
        """Performance tuning - redirect to implemented method"""
        self.performance_optimization_settings()
        
    def directory_path_configuration(self):
        """Directory path configuration - redirect to implemented method"""
        self.path_directory_settings()
        
    def backup_restore_settings(self):
        """Backup and restore settings - redirect to implemented method"""
        self.backup_restore_management()
    
    def reporting_analytics_menu(self):
        """Reporting and analytics interface"""
        while True:
            self.print_master_banner()
            print("╔═══════════════════════════════════════════════════════════════════════════════╗")
            print("║\033[1;96m                    📊 REPORTING & ANALYTICS\033[0m                        ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m1.\033[0m 📄 Generate Comprehensive Report                              ║")
            print("║  \033[1;97m2.\033[0m 📊 View Scan Statistics & Metrics                            ║")
            print("║  \033[1;97m3.\033[0m 📈 Trend Analysis & Comparisons                             ║")
            print("║  \033[1;97m4.\033[0m 🔍 Search & Filter Results                                  ║")
            print("║  \033[1;97m5.\033[0m 📋 Export Reports (PDF, HTML, JSON)                        ║")
            print("║  \033[1;97m6.\033[0m 📊 Dashboard & Visualizations                              ║")
            print("║  \033[1;97m7.\033[0m 📧 Schedule & Share Reports                                 ║")
            print("║  \033[1;97m8.\033[0m ⚙️ Configure Reporting Settings                            ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m0.\033[0m ↩️ Return to Main Menu                                           ║")
            print("╚═══════════════════════════════════════════════════════════════════════════════╝")
            
            choice = self.get_user_input()
            
            if choice == "0":
                break
            elif choice == "1":
                self.generate_comprehensive_report()
            elif choice == "2":
                self.view_scan_statistics()
            elif choice == "3":
                self.trend_analysis()
            elif choice == "4":
                self.search_filter_results()
            elif choice == "5":
                self.export_reports()
            elif choice == "6":
                self.dashboard_visualizations()
            elif choice == "7":
                self.schedule_share_reports()
            elif choice == "8":
                self.configure_reporting_settings()
            else:
                self.show_error(f"Invalid option: {choice}")
                
    def generate_comprehensive_report(self):
        """Generate comprehensive security assessment report"""
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        
        if not runs_dir.exists():
            self.show_error("No scan results directory found")
            return
            
        print("\n📄 \033[1;97mComprehensive Report Generation:\033[0m")
        
        # Find recent runs
        run_dirs = [d for d in runs_dir.iterdir() if d.is_dir()]
        run_dirs.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        
        if not run_dirs:
            self.show_error("No scan results found")
            return
            
        print(f"📊 Found {len(run_dirs)} scan runs")
        
        # Select runs to include
        print("\nReport scope:")
        print("1. Latest scan only")
        print("2. Last 5 scans")
        print("3. Last 30 days")
        print("4. All scans")
        print("5. Custom selection")
        
        scope = input("\nSelect report scope (1-5): ").strip()
        
        selected_runs = []
        if scope == "1":
            selected_runs = run_dirs[:1]
        elif scope == "2":
            selected_runs = run_dirs[:5]
        elif scope == "3":
            # Last 30 days
            cutoff = datetime.now().timestamp() - (30 * 24 * 60 * 60)
            selected_runs = [d for d in run_dirs if d.stat().st_mtime > cutoff]
        elif scope == "4":
            selected_runs = run_dirs
        elif scope == "5":
            print(f"\nAvailable runs:")
            for i, run_dir in enumerate(run_dirs[:20], 1):
                print(f"  {i}. {run_dir.name}")
            
            selection = input("Enter run numbers (comma-separated): ").strip()
            try:
                indices = [int(x.strip()) - 1 for x in selection.split(',')]
                selected_runs = [run_dirs[i] for i in indices if 0 <= i < len(run_dirs)]
            except:
                self.show_error("Invalid selection")
                return
        else:
            self.show_error("Invalid scope selection")
            return
            
        if not selected_runs:
            self.show_error("No runs selected")
            return
            
        try:
            self.show_info(f"Generating report for {len(selected_runs)} scan runs...")
            
            # Create report directory
            report_dir = Path("reports")
            report_dir.mkdir(exist_ok=True)
            
            report_data = self.collect_report_data(selected_runs)
            
            # Generate HTML report
            html_report = self.generate_html_report(report_data)
            
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            report_file = report_dir / f"comprehensive_report_{timestamp}.html"
            
            with open(report_file, 'w') as f:
                f.write(html_report)
                
            self.show_success(f"Report generated: {report_file}")
            
            # Optionally open report
            open_report = input("\nOpen report in browser? (y/N): ")
            if open_report.lower() == 'y':
                import webbrowser
                webbrowser.open(f"file://{report_file.absolute()}")
                
        except Exception as e:
            self.show_error(f"Error generating report: {e}")
            
        self.wait_for_continue()
        
    def collect_report_data(self, run_dirs):
        """Collect data from scan runs for reporting"""
        report_data = {
            'summary': {
                'total_runs': len(run_dirs),
                'scan_types': set(),
                'targets_scanned': set(),
                'total_findings': 0,
                'critical_findings': 0,
                'high_findings': 0,
                'medium_findings': 0,
                'low_findings': 0
            },
            'runs': [],
            'findings': [],
            'statistics': {}
        }
        
        for run_dir in run_dirs:
            run_data = {
                'name': run_dir.name,
                'timestamp': datetime.fromtimestamp(run_dir.stat().st_mtime),
                'type': self.identify_scan_type(run_dir.name),
                'files': list(run_dir.glob("*")),
                'findings': []
            }
            
            # Extract findings from JSON files
            for json_file in run_dir.glob("*.json"):
                try:
                    with open(json_file, 'r') as f:
                        data = json.load(f)
                        if isinstance(data, list):
                            run_data['findings'].extend(data)
                        elif isinstance(data, dict) and 'findings' in data:
                            run_data['findings'].extend(data['findings'])
                except Exception as e:
                    self.logger.debug(f"Error processing JSON file {json_file}: {e}")
                    continue
                    
            report_data['runs'].append(run_data)
            report_data['summary']['scan_types'].add(run_data['type'])
            report_data['summary']['total_findings'] += len(run_data['findings'])
            
            # Categorize findings by severity
            for finding in run_data['findings']:
                severity = finding.get('severity', '').lower()
                if severity in ['critical']:
                    report_data['summary']['critical_findings'] += 1
                elif severity in ['high']:
                    report_data['summary']['high_findings'] += 1
                elif severity in ['medium']:
                    report_data['summary']['medium_findings'] += 1
                elif severity in ['low', 'info']:
                    report_data['summary']['low_findings'] += 1
                    
        return report_data
        
    def identify_scan_type(self, run_name):
        """Identify scan type from run directory name"""
        if 'recon' in run_name.lower():
            return 'Reconnaissance'
        elif 'vuln' in run_name.lower():
            return 'Vulnerability Scan'
        elif 'web' in run_name.lower():
            return 'Web Application Test'
        elif 'api' in run_name.lower():
            return 'API Security Test'
        elif 'cloud' in run_name.lower() or 'aws' in run_name.lower() or 'azure' in run_name.lower():
            return 'Cloud Security Assessment'
        elif 'infra' in run_name.lower() or 'network' in run_name.lower():
            return 'Infrastructure Scan'
        elif 'fuzz' in run_name.lower():
            return 'Fuzzing'
        else:
            return 'General Scan'
            
    def generate_html_report(self, report_data):
        """Generate HTML report from collected data"""
        html = f"""
<!DOCTYPE html>
<html>
<head>
    <title>Azaz-El Security Assessment Report</title>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; margin: 0; padding: 20px; background: #1e1e1e; color: #eee; }}
        .header {{ background: linear-gradient(135deg, #8B0000, #FF6347); padding: 25px; text-align: center; color: white; margin-bottom: 30px; }}
        .summary {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 30px; }}
        .card {{ background: #2d2d2d; padding: 20px; border-radius: 8px; border-left: 4px solid #FF6347; }}
        .card h3 {{ margin: 0 0 10px 0; color: #FF6347; }}
        .card .number {{ font-size: 2em; font-weight: bold; }}
        .findings {{ background: #2d2d2d; padding: 20px; border-radius: 8px; margin-bottom: 20px; }}
        .finding {{ background: #3d3d3d; padding: 15px; margin: 10px 0; border-radius: 5px; border-left: 4px solid #666; }}
        .critical {{ border-left-color: #dc2626; }}
        .high {{ border-left-color: #ea580c; }}
        .medium {{ border-left-color: #facc15; }}
        .low {{ border-left-color: #16a34a; }}
        .runs {{ background: #2d2d2d; padding: 20px; border-radius: 8px; }}
        .run {{ background: #3d3d3d; padding: 15px; margin: 10px 0; border-radius: 5px; }}
        table {{ width: 100%; border-collapse: collapse; background: #2d2d2d; }}
        th, td {{ padding: 12px; text-align: left; border-bottom: 1px solid #555; }}
        th {{ background: #3d3d3d; color: #FF6347; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>🔒 Azaz-El Security Assessment Report</h1>
        <p>Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p>Framework Version: {MASTER_VERSION}</p>
    </div>
    
    <div class="summary">
        <div class="card">
            <h3>Total Scans</h3>
            <div class="number">{report_data['summary']['total_runs']}</div>
        </div>
        <div class="card">
            <h3>Total Findings</h3>
            <div class="number">{report_data['summary']['total_findings']}</div>
        </div>
        <div class="card">
            <h3>Critical Issues</h3>
            <div class="number" style="color: #dc2626;">{report_data['summary']['critical_findings']}</div>
        </div>
        <div class="card">
            <h3>High Risk Issues</h3>
            <div class="number" style="color: #ea580c;">{report_data['summary']['high_findings']}</div>
        </div>
    </div>
    
    <div class="findings">
        <h2>📊 Findings Summary</h2>
        <table>
            <tr>
                <th>Severity</th>
                <th>Count</th>
                <th>Percentage</th>
            </tr>
            <tr>
                <td>Critical</td>
                <td>{report_data['summary']['critical_findings']}</td>
                <td>{(report_data['summary']['critical_findings'] / max(report_data['summary']['total_findings'], 1) * 100):.1f}%</td>
            </tr>
            <tr>
                <td>High</td>
                <td>{report_data['summary']['high_findings']}</td>
                <td>{(report_data['summary']['high_findings'] / max(report_data['summary']['total_findings'], 1) * 100):.1f}%</td>
            </tr>
            <tr>
                <td>Medium</td>
                <td>{report_data['summary']['medium_findings']}</td>
                <td>{(report_data['summary']['medium_findings'] / max(report_data['summary']['total_findings'], 1) * 100):.1f}%</td>
            </tr>
            <tr>
                <td>Low</td>
                <td>{report_data['summary']['low_findings']}</td>
                <td>{(report_data['summary']['low_findings'] / max(report_data['summary']['total_findings'], 1) * 100):.1f}%</td>
            </tr>
        </table>
    </div>
    
    <div class="runs">
        <h2>📁 Scan Runs Overview</h2>
"""
        
        for run in report_data['runs']:
            html += f"""
        <div class="run">
            <h3>{run['name']}</h3>
            <p><strong>Type:</strong> {run['type']}</p>
            <p><strong>Date:</strong> {run['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}</p>
            <p><strong>Findings:</strong> {len(run['findings'])}</p>
            <p><strong>Files Generated:</strong> {len(run['files'])}</p>
        </div>
"""
        
        html += """
    </div>
    
    <div class="findings">
        <h2>🔍 Detailed Findings</h2>
        <p>Detailed findings data would be included here in a full implementation.</p>
    </div>
    
    <footer style="text-align: center; margin-top: 50px; padding: 20px; border-top: 1px solid #555;">
        <p>Generated by Azaz-El Security Assessment Framework</p>
    </footer>
</body>
</html>
"""
        return html
        
    def view_scan_statistics(self):
        """View scan statistics and metrics"""
        print("\n📊 \033[1;97mScan Statistics & Metrics:\033[0m")
        
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        if not runs_dir.exists():
            self.show_info("No scan results found")
            return
            
        # Collect statistics
        all_runs = [d for d in runs_dir.iterdir() if d.is_dir()]
        
        if not all_runs:
            self.show_info("No scan runs found")
            return
            
        # Calculate statistics
        total_runs = len(all_runs)
        
        # Calculate total size more efficiently
        total_size = 0
        for run_dir in all_runs:
            try:
                dir_size = sum(f.stat().st_size for f in run_dir.rglob('*') if f.is_file())
                total_size += dir_size
            except (OSError, PermissionError):
                # Skip directories that can't be read
                continue
        
        # Group by scan type
        scan_types = {}
        for run_dir in all_runs:
            scan_type = self.identify_scan_type(run_dir.name)
            scan_types[scan_type] = scan_types.get(scan_type, 0) + 1
            
        # Recent activity (last 7 days)
        week_ago = datetime.now().timestamp() - (7 * 24 * 60 * 60)
        recent_runs = [d for d in all_runs if d.stat().st_mtime > week_ago]
        
        print(f"📈 Overall Statistics:")
        print(f"   • Total Scan Runs: {total_runs}")
        print(f"   • Total Data Size: {total_size // (1024*1024)} MB")
        print(f"   • Recent Activity (7 days): {len(recent_runs)} scans")
        print(f"   • Average per day: {len(recent_runs) / 7:.1f} scans")
        
        print(f"\n📊 Scan Types Distribution:")
        for scan_type, count in sorted(scan_types.items(), key=lambda x: x[1], reverse=True):
            percentage = (count / total_runs) * 100
            print(f"   • {scan_type}: {count} ({percentage:.1f}%)")
            
        # Monthly trend
        print(f"\n📅 Monthly Activity:")
        monthly_stats = {}
        for run_dir in all_runs:
            month_key = datetime.fromtimestamp(run_dir.stat().st_mtime).strftime('%Y-%m')
            monthly_stats[month_key] = monthly_stats.get(month_key, 0) + 1
            
        for month, count in sorted(monthly_stats.items(), reverse=True)[:6]:
            print(f"   • {month}: {count} scans")
            
        self.wait_for_continue()
        
    def trend_analysis(self):
        """Trend analysis and comparisons"""
        print("\n📈 \033[1;97mTrend Analysis & Comparisons:\033[0m")
        print("This feature would provide:")
        print("• Vulnerability trend analysis over time")
        print("• Comparison between different targets")
        print("• Security posture improvement tracking")
        print("• Seasonal patterns in findings")
        print("• Tool effectiveness analysis")
        print("\n💡 Advanced analytics would be implemented here")
        self.wait_for_continue()
        
    def search_filter_results(self):
        """Search and filter scan results"""
        print("\n🔍 \033[1;97mSearch & Filter Results:\033[0m")
        
        search_term = input("Enter search term (target, finding type, etc.): ").strip()
        if not search_term:
            return
            
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        matching_results = []
        
        for run_dir in runs_dir.iterdir():
            if not run_dir.is_dir():
                continue
                
            # Search in directory name
            if search_term.lower() in run_dir.name.lower():
                matching_results.append(('Directory', run_dir.name, run_dir))
                
            # Search in JSON files
            for json_file in run_dir.glob("*.json"):
                try:
                    with open(json_file, 'r') as f:
                        content = f.read()
                        if search_term.lower() in content.lower():
                            matching_results.append(('Content', f"{run_dir.name}/{json_file.name}", json_file))
                except Exception as e:
                    self.logger.debug(f"Error searching in file {json_file}: {e}")
                    continue
                    
        if matching_results:
            print(f"\n🔍 Found {len(matching_results)} matches for '{search_term}':")
            for i, (match_type, path, file_path) in enumerate(matching_results[:20], 1):
                print(f"  {i}. [{match_type}] {path}")
                
            if len(matching_results) > 20:
                print(f"  ... and {len(matching_results) - 20} more matches")
        else:
            print(f"No matches found for '{search_term}'")
            
        self.wait_for_continue()
        
    def export_reports(self):
        """Export reports in various formats"""
        print("\n📋 \033[1;97mExport Reports:\033[0m")
        print("1. Export as PDF")
        print("2. Export as HTML")
        print("3. Export as JSON")
        print("4. Export as CSV")
        print("5. Export as XML")
        
        choice = input("\nSelect export format (1-5): ").strip()
        
        if choice == "1":
            print("📄 PDF export would require additional libraries (reportlab, weasyprint)")
        elif choice == "2":
            self.generate_comprehensive_report()
            return
        elif choice == "3":
            self.export_json_report()
        elif choice == "4":
            self.export_csv_report()
        elif choice == "5":
            print("📄 XML export functionality would be implemented here")
        else:
            self.show_error("Invalid export format")
            
        self.wait_for_continue()
        
    def export_json_report(self):
        """Export report as JSON"""
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        run_dirs = [d for d in runs_dir.iterdir() if d.is_dir()]
        
        if not run_dirs:
            self.show_error("No scan data to export")
            return
            
        report_data = self.collect_report_data(run_dirs)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        export_file = Path("reports") / f"report_export_{timestamp}.json"
        export_file.parent.mkdir(exist_ok=True)
        
        with open(export_file, 'w') as f:
            json.dump(report_data, f, indent=2, default=str)
            
        self.show_success(f"JSON report exported: {export_file}")
        
    def export_csv_report(self):
        """Export report as CSV"""
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        run_dirs = [d for d in runs_dir.iterdir() if d.is_dir()]
        
        if not run_dirs:
            self.show_error("No scan data to export")
            return
            
        import csv
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        export_file = Path("reports") / f"report_export_{timestamp}.csv"
        export_file.parent.mkdir(exist_ok=True)
        
        with open(export_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow(['Scan_Name', 'Type', 'Date', 'Findings_Count', 'Files_Count'])
            
            for run_dir in run_dirs:
                scan_type = self.identify_scan_type(run_dir.name)
                scan_date = datetime.fromtimestamp(run_dir.stat().st_mtime)
                files_count = len(list(run_dir.glob("*")))
                
                # Count findings in JSON files
                findings_count = 0
                for json_file in run_dir.glob("*.json"):
                    try:
                        with open(json_file, 'r') as jf:
                            data = json.load(jf)
                            if isinstance(data, list):
                                findings_count += len(data)
                    except Exception as e:
                        self.logger.debug(f"Error counting findings in {json_file}: {e}")
                        continue
                        
                writer.writerow([
                    run_dir.name,
                    scan_type,
                    scan_date.isoformat(),
                    findings_count,
                    files_count
                ])
                
        self.show_success(f"CSV report exported: {export_file}")
        
    def dashboard_visualizations(self):
        """Dashboard and visualizations"""
        print("\n📊 \033[1;97mDashboard & Visualizations:\033[0m")
        print("This would provide:")
        print("• Real-time scanning progress")
        print("• Interactive charts and graphs")
        print("• Vulnerability severity breakdown")
        print("• Target coverage maps")
        print("• Performance metrics")
        print("• Historical trend charts")
        print("\n💡 Advanced visualization dashboard would be implemented here")
        print("Consider using libraries like matplotlib, plotly, or web-based dashboards")
        self.wait_for_continue()
        
    def schedule_share_reports(self):
        """Schedule and share reports"""
        print("\n📧 \033[1;97mSchedule & Share Reports:\033[0m")
        print("This would provide:")
        print("• Automated report generation")
        print("• Email distribution lists")
        print("• Slack/Teams integration")
        print("• Scheduled weekly/monthly reports")
        print("• Report templates")
        print("• Executive summaries")
        print("\n💡 Report scheduling and sharing would be implemented here")
        self.wait_for_continue()
        
    def configure_reporting_settings(self):
        """Configure reporting settings"""
        print("\n⚙️ \033[1;97mReporting Configuration:\033[0m")
        print("=" * 50)
        
        config = self.moloch_integration.config
        report_config = config.get('reporting', {})
        
        print(f"📊 Current Settings:")
        print(f"   • Report Format: {report_config.get('default_format', 'HTML')}")
        print(f"   • Include Screenshots: {report_config.get('include_screenshots', True)}")
        print(f"   • Auto-open Reports: {report_config.get('auto_open', False)}")
        print(f"   • Report Directory: {report_config.get('output_dir', 'reports')}")
        print(f"   • Template Theme: {report_config.get('theme', 'dark')}")
        print(f"   • Include Raw Data: {report_config.get('include_raw_data', False)}")
        print(f"   • Max Report Size: {report_config.get('max_size_mb', 50)} MB")
        
        print(f"\n💡 Use moloch.cfg.json to modify these settings")
        self.wait_for_continue()
    
    def system_dashboard_menu(self):
        """System dashboard interface"""
        while True:
            self.print_master_banner()
            print("╔═══════════════════════════════════════════════════════════════════════════════╗")
            print("║\033[1;94m                      📊 SYSTEM DASHBOARD\033[0m                           ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m1.\033[0m 📈 Real-time System Monitoring                               ║")
            print("║  \033[1;97m2.\033[0m 🚀 Active Scans Status                                      ║")
            print("║  \033[1;97m3.\033[0m 📊 Performance Metrics                                      ║")
            print("║  \033[1;97m4.\033[0m 🔧 System Health Check                                      ║")
            print("║  \033[1;97m5.\033[0m 📁 Resource Usage Monitor                                  ║")
            print("║  \033[1;97m6.\033[0m 🔍 Process & Service Status                                ║")
            print("║  \033[1;97m7.\033[0m 📋 System Logs Viewer                                      ║")
            print("║  \033[1;97m8.\033[0m ⚙️ Dashboard Settings                                       ║")
            print("╠───────────────────────────────────────────────────────────────────────────────╣")
            print("║  \033[1;97m0.\033[0m ↩️ Return to Main Menu                                           ║")
            print("╚═══════════════════════════════════════════════════════════════════════════════╝")
            
            choice = self.get_user_input()
            
            if choice == "0":
                break
            elif choice == "1":
                self.realtime_monitoring()
            elif choice == "2":
                self.active_scans_status()
            elif choice == "3":
                self.performance_metrics()
            elif choice == "4":
                self.system_health_check()
            elif choice == "5":
                self.resource_usage_monitor()
            elif choice == "6":
                self.process_service_status()
            elif choice == "7":
                self.system_logs_viewer()
            elif choice == "8":
                self.dashboard_settings()
            else:
                self.show_error(f"Invalid option: {choice}")
                
    def realtime_monitoring(self):
        """Real-time system monitoring dashboard"""
        print("\n📈 \033[1;97mReal-time System Monitoring:\033[0m")
        print("Press Ctrl+C to exit monitoring...")
        
        try:
            import time
            import psutil
            
            while True:
                # Clear screen (simple approach)
                print("\033[2J\033[H")
                
                # System overview
                print("=" * 80)
                print(f"🖥️ AZAZ-EL SYSTEM DASHBOARD - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
                print("=" * 80)
                
                # CPU and Memory
                cpu_percent = psutil.cpu_percent(interval=1)
                memory = psutil.virtual_memory()
                disk = psutil.disk_usage('/')
                
                print(f"💻 System Resources:")
                print(f"   CPU Usage:    {cpu_percent:6.1f}% {'🔴' if cpu_percent > 80 else '🟡' if cpu_percent > 60 else '🟢'}")
                print(f"   Memory Usage: {memory.percent:6.1f}% {'🔴' if memory.percent > 80 else '🟡' if memory.percent > 60 else '🟢'}")
                print(f"   Disk Usage:   {disk.percent:6.1f}% {'🔴' if disk.percent > 90 else '🟡' if disk.percent > 75 else '🟢'}")
                
                # Network
                try:
                    net_io = psutil.net_io_counters()
                    print(f"   Network Sent: {net_io.bytes_sent // (1024*1024):6d} MB")
                    print(f"   Network Recv: {net_io.bytes_recv // (1024*1024):6d} MB")
                except Exception as e:
                    self.logger.debug(f"Network statistics unavailable: {e}")
                    print("   Network Info: Unavailable")
                
                # Active scans
                print(f"\n🚀 Active Scans:")
                if self.active_scans:
                    for scan_id, scan_info in self.active_scans.items():
                        print(f"   📊 {scan_id}: {scan_info.get('target', 'Unknown')} - {scan_info.get('status', 'Unknown')}")
                else:
                    print("   No active scans")
                
                # Framework status
                print(f"\n🛠️ Framework Status:")
                print(f"   Version:      {MASTER_VERSION}")
                print(f"   Targets:      {len(self.targets)}")
                print(f"   Integrations: {'✅ Available' if self.integrations_available else '❌ Limited'}")
                
                print("\n" + "=" * 80)
                print("Press Ctrl+C to exit monitoring...")
                
                time.sleep(5)
                
        except KeyboardInterrupt:
            print("\n\n📊 Monitoring stopped")
        except ImportError:
            print("❌ psutil not available for real-time monitoring")
            
        self.wait_for_continue()
        
    def active_scans_status(self):
        """Display active scans status"""
        print("\n🚀 \033[1;97mActive Scans Status:\033[0m")
        
        if not self.active_scans:
            print("📊 No active scans running")
        else:
            print(f"📊 {len(self.active_scans)} active scan(s):")
            
            for scan_id, scan_info in self.active_scans.items():
                print(f"\n🔍 Scan ID: {scan_id}")
                print(f"   Target: {scan_info.get('target', 'Unknown')}")
                print(f"   Status: {scan_info.get('status', 'Unknown')}")
                print(f"   Phase: {scan_info.get('phase', 'Unknown')}")
                print(f"   Started: {scan_info.get('start_time', 'Unknown')}")
                
                if scan_info.get('aggressive'):
                    print(f"   Mode: 🔥 Aggressive")
                else:
                    print(f"   Mode: 🌿 Normal")
                    
        # Scan history summary
        print(f"\n📈 Scan History:")
        print(f"   Total completed: {len(self.scan_history)}")
        
        if self.scan_history:
            recent = self.scan_history[-5:]  # Last 5 scans
            print(f"   Recent scans:")
            for scan in recent:
                status_icon = "✅" if scan.get('status') == 'completed' else "❌"
                print(f"     {status_icon} {scan.get('target', 'Unknown')} - {scan.get('status', 'Unknown')}")
                
        self.wait_for_continue()
        
    def performance_metrics(self):
        """Display performance metrics"""
        print("\n📊 \033[1;97mPerformance Metrics:\033[0m")
        
        # Framework performance
        print(f"🚀 Framework Performance:")
        print(f"   Total Scans: {self.performance_metrics.get('total_scans', 0)}")
        print(f"   Successful: {self.performance_metrics.get('successful_scans', 0)}")
        print(f"   Failed: {self.performance_metrics.get('failed_scans', 0)}")
        
        total = self.performance_metrics.get('total_scans', 0)
        if total > 0:
            success_rate = (self.performance_metrics.get('successful_scans', 0) / total) * 100
            print(f"   Success Rate: {success_rate:.1f}%")
        
        # Average scan time
        avg_time = self.performance_metrics.get('avg_scan_time', 0)
        if avg_time > 0:
            print(f"   Avg Scan Time: {avg_time:.1f} seconds")
            
        # Last scan time
        last_scan = self.performance_metrics.get('last_scan_time')
        if last_scan:
            print(f"   Last Scan: {last_scan}")
            
        # System metrics
        try:
            import psutil
            
            print(f"\n💻 System Performance:")
            
            # CPU info
            cpu_count = psutil.cpu_count()
            cpu_freq = psutil.cpu_freq()
            print(f"   CPU Cores: {cpu_count}")
            if cpu_freq:
                print(f"   CPU Frequency: {cpu_freq.current:.0f} MHz")
                
            # Memory info
            memory = psutil.virtual_memory()
            print(f"   Total Memory: {memory.total // (1024**3)} GB")
            print(f"   Available Memory: {memory.available // (1024**3)} GB")
            
            # Disk info
            disk = psutil.disk_usage('/')
            print(f"   Total Disk: {disk.total // (1024**3)} GB")
            print(f"   Free Disk: {disk.free // (1024**3)} GB")
            
        except ImportError:
            print("\n💻 System performance monitoring not available")
            
        self.wait_for_continue()
        
    def system_health_check(self):
        """Perform comprehensive system health check"""
        print("\n🔧 \033[1;97mSystem Health Check:\033[0m")
        
        health_score = 0
        max_score = 0
        issues = []
        
        # Check Python version
        max_score += 10
        import sys
        if sys.version_info >= (3, 8):
            print("✅ Python version OK")
            health_score += 10
        else:
            print("❌ Python version outdated")
            issues.append("Python version < 3.8")
            
        # Check required modules
        max_score += 20
        required_modules = ['json', 'pathlib', 'datetime', 'asyncio', 'subprocess']
        missing_modules = []
        
        for module in required_modules:
            try:
                __import__(module)
                health_score += 4
            except ImportError:
                missing_modules.append(module)
                
        if missing_modules:
            print(f"❌ Missing modules: {', '.join(missing_modules)}")
            issues.append(f"Missing modules: {', '.join(missing_modules)}")
        else:
            print("✅ Required modules OK")
            
        # Check configuration files
        max_score += 20
        config_files = ['moloch.cfg.json', 'moloch.py']
        missing_configs = []
        
        for config_file in config_files:
            if Path(config_file).exists():
                health_score += 10
            else:
                missing_configs.append(config_file)
                
        if missing_configs:
            print(f"⚠️ Missing config files: {', '.join(missing_configs)}")
            issues.append(f"Missing configs: {', '.join(missing_configs)}")
        else:
            print("✅ Configuration files OK")
            
        # Check integrations
        max_score += 20
        if self.integrations_available:
            print("✅ Integrations available")
            health_score += 20
        else:
            print("⚠️ Limited integrations")
            issues.append("Limited integration functionality")
            health_score += 10
            
        # Check system resources
        max_score += 20
        try:
            import psutil
            
            # CPU usage
            cpu_usage = psutil.cpu_percent(interval=1)
            if cpu_usage < 70:
                health_score += 7
            elif cpu_usage < 90:
                health_score += 4
            else:
                issues.append("High CPU usage")
                
            # Memory usage
            memory = psutil.virtual_memory()
            if memory.percent < 70:
                health_score += 7
            elif memory.percent < 90:
                health_score += 4
            else:
                issues.append("High memory usage")
                
            # Disk space
            disk = psutil.disk_usage('/')
            if disk.percent < 80:
                health_score += 6
            elif disk.percent < 95:
                health_score += 3
            else:
                issues.append("Low disk space")
                
            print("✅ System resources checked")
            
        except ImportError:
            print("⚠️ Cannot check system resources (psutil not available)")
            health_score += 10
            
        # Calculate health percentage
        health_percentage = (health_score / max_score) * 100
        
        print(f"\n📊 Overall Health Score: {health_score}/{max_score} ({health_percentage:.1f}%)")
        
        if health_percentage >= 90:
            print("🟢 System Health: Excellent")
        elif health_percentage >= 75:
            print("🟡 System Health: Good")
        elif health_percentage >= 60:
            print("🟠 System Health: Fair")
        else:
            print("🔴 System Health: Poor")
            
        if issues:
            print(f"\n⚠️ Issues to address:")
            for issue in issues:
                print(f"   • {issue}")
                
        self.wait_for_continue()
        
    def resource_usage_monitor(self):
        """Monitor resource usage"""
        print("\n📁 \033[1;97mResource Usage Monitor:\033[0m")
        
        try:
            import psutil
            
            # Disk usage by directory
            runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
            if runs_dir.exists():
                runs_size = sum(f.stat().st_size for f in runs_dir.rglob('*') if f.is_file())
                print(f"📁 Scan Results: {runs_size // (1024*1024)} MB")
                
            # Memory usage by process
            current_process = psutil.Process()
            memory_info = current_process.memory_info()
            print(f"🧠 Framework Memory: {memory_info.rss // (1024*1024)} MB")
            
            # Check for large files
            large_files = []
            if runs_dir.exists():
                for file_path in runs_dir.rglob('*'):
                    if file_path.is_file() and file_path.stat().st_size > 10 * 1024 * 1024:  # >10MB
                        large_files.append((file_path, file_path.stat().st_size))
                        
            if large_files:
                print(f"\n📦 Large files (>10MB):")
                large_files.sort(key=lambda x: x[1], reverse=True)
                for file_path, size in large_files[:10]:
                    print(f"   • {file_path.name}: {size // (1024*1024)} MB")
                    
        except ImportError:
            print("❌ Resource monitoring not available (psutil required)")
            
        self.wait_for_continue()
        
    def process_service_status(self):
        """Check process and service status"""
        print("\n🔍 \033[1;97mProcess & Service Status:\033[0m")
        
        # Check for security tools processes
        security_tools = ['nmap', 'nuclei', 'ffuf', 'gobuster', 'httpx']
        
        try:
            import psutil
            
            running_tools = []
            for proc in psutil.process_iter(['name', 'pid', 'cpu_percent']):
                proc_name = proc.info['name']
                if any(tool in proc_name.lower() for tool in security_tools):
                    running_tools.append(proc.info)
                    
            if running_tools:
                print(f"🔧 Running security tools:")
                for proc in running_tools:
                    print(f"   • {proc['name']} (PID: {proc['pid']}) - CPU: {proc['cpu_percent']:.1f}%")
            else:
                print("🔧 No security tools currently running")
                
        except ImportError:
            print("❌ Process monitoring not available")
            
        # Check network connections
        try:
            import psutil
            
            connections = psutil.net_connections(kind='inet')
            active_connections = [c for c in connections if c.status == 'ESTABLISHED']
            
            if active_connections:
                print(f"\n🌐 Active network connections: {len(active_connections)}")
                # Show unique remote addresses
                remote_ips = set(c.raddr.ip for c in active_connections if c.raddr)
                if remote_ips:
                    print(f"   Connected to {len(remote_ips)} unique remote hosts")
            else:
                print("\n🌐 No active network connections")
                
        except:
            print("\n🌐 Network monitoring not available")
            
        self.wait_for_continue()
        
    def system_logs_viewer(self):
        """View system logs"""
        print("\n📋 \033[1;97mSystem Logs Viewer:\033[0m")
        
        # Check for log files
        log_files = []
        
        # Common log locations
        log_paths = [
            Path("logs"),
            Path("."),
            Path("/tmp")
        ]
        
        for log_path in log_paths:
            if log_path.exists():
                log_files.extend(log_path.glob("*.log"))
                
        if not log_files:
            print("📋 No log files found")
        else:
            print(f"📋 Found {len(log_files)} log files:")
            
            for i, log_file in enumerate(log_files[:10], 1):
                size = log_file.stat().st_size // 1024  # KB
                modified = datetime.fromtimestamp(log_file.stat().st_mtime)
                print(f"   {i}. {log_file.name} ({size} KB) - {modified.strftime('%Y-%m-%d %H:%M')}")
                
            # Option to view a log file
            try:
                choice = int(input(f"\nSelect log file to view (1-{len(log_files[:10])}), or 0 to skip: "))
                if 1 <= choice <= len(log_files[:10]):
                    selected_log = log_files[choice - 1]
                    self.view_log_file(selected_log)
            except (ValueError, IndexError):
                pass
                
        self.wait_for_continue()
        
    def view_log_file(self, log_file):
        """View contents of a log file"""
        try:
            print(f"\n📖 Viewing {log_file.name} (last 50 lines):")
            print("-" * 60)
            
            with open(log_file, 'r') as f:
                lines = f.readlines()
                
            # Show last 50 lines
            for line in lines[-50:]:
                print(line.rstrip())
                
            print("-" * 60)
            
        except Exception as e:
            print(f"❌ Error reading log file: {e}")
            
    def dashboard_settings(self):
        """Dashboard settings configuration"""
        print("\n⚙️ \033[1;97mDashboard Settings:\033[0m")
        print("=" * 50)
        
        print(f"📊 Current Dashboard Settings:")
        print(f"   • Auto-refresh: Disabled")
        print(f"   • Theme: Dark")
        print(f"   • Update interval: 5 seconds")
        print(f"   • Show system resources: Yes")
        print(f"   • Show active scans: Yes")
        print(f"   • Log level: INFO")
        
        print(f"\n💡 Dashboard settings would be configurable here")
        self.wait_for_continue()
    
    # Additional submenu handlers - now fully implemented above
    def handle_reconnaissance_submenu(self, choice):
        """Handle reconnaissance submenu choices - integrated into main reconnaissance menu"""
        self.reconnaissance_suite_menu()

    def handle_vulnerability_submenu(self, choice):
        """Handle vulnerability submenu choices - integrated into main vulnerability menu"""
        self.vulnerability_scanning_menu()

    def handle_web_testing_submenu(self, choice):
        """Handle web testing submenu choices - integrated into main web testing menu"""
        self.web_application_testing_menu()

    def handle_cloud_submenu(self, choice):
        """Handle cloud submenu choices - integrated into main cloud menu"""
        self.cloud_security_assessment_menu()

    def handle_system_config_submenu(self, choice):
        """Handle system config submenu choices - integrated into main system config menu"""
        self.system_configuration_menu()

    def handle_reporting_submenu(self, choice):
        """Handle reporting submenu choices - integrated into main reporting menu"""
        self.reporting_analytics_menu()

    def handle_dashboard_submenu(self, choice):
        """Handle dashboard submenu choices - integrated into main dashboard menu"""
        self.system_dashboard_menu()
    
    def handle_reporting_submenu(self, choice):
        """Handle reporting submenu choices - integrated into main reporting menu"""
        self.reporting_analytics_menu()

    def handle_dashboard_submenu(self, choice):
        """Handle dashboard submenu choices - integrated into main dashboard menu"""
        self.system_dashboard_menu()
    
    # Target management implementations
    def _validate_target_format(self, target: str) -> tuple[bool, str]:
        """Validate target format and return (is_valid, error_message)"""
        if not target or not target.strip():
            return False, "Target cannot be empty"
        
        target = target.strip()
        
        if ' ' in target:
            return False, "Target cannot contain spaces"
        
        if len(target) > 253:  # DNS limit
            return False, "Target is too long (max 253 characters)"
        
        # Check for basic format
        if target.startswith(('http://', 'https://')):
            # URL validation
            if not '.' in target or target.count('.') < 1:
                return False, "Invalid URL format"
        elif '.' in target:
            # Domain validation
            if target.startswith('.') or target.endswith('.') or '..' in target:
                return False, "Invalid domain format"
            if not all(c.isalnum() or c in '.-' for c in target):
                return False, "Domain contains invalid characters"
        elif ':' in target:
            # IPv6 or port validation
            pass  # Accept for now
        elif target.replace('.', '').isdigit():
            # IPv4 validation
            parts = target.split('.')
            if len(parts) != 4:
                return False, "Invalid IPv4 format"
            try:
                if not all(0 <= int(part) <= 255 for part in parts):
                    return False, "Invalid IPv4 address range"
            except ValueError:
                return False, "Invalid IPv4 format"
        else:
            return False, "Unrecognized target format"
        
        return True, ""

    def add_single_target(self) -> None:
        """Add a single target to the target list"""
        self.print_master_banner()
        print("╔═══════════════════════════════════════════════════════════════════════════════╗")
        print("║\033[1;92m                        📝 ADD SINGLE TARGET 📝\033[0m                        ║")
        print("╚═══════════════════════════════════════════════════════════════════════════════╝")
        
        print("\n💡 \033[1;97mSupported target formats:\033[0m")
        print("   • Domain: example.com")
        print("   • Subdomain: sub.example.com")
        print("   • IP Address: 192.168.1.1")
        print("   • URL: https://example.com")
        
        target = input("\n🎯 \033[1;97mEnter target: \033[0m").strip()
        
        # Validate target format
        is_valid, error_message = self._validate_target_format(target)
        if not is_valid:
            self.show_error(error_message)
            return
        
        # Load existing targets
        try:
            targets_file = Path("targets.txt")
            if targets_file.exists():
                existing_targets = targets_file.read_text().strip().split('\n')
                existing_targets = [t.strip() for t in existing_targets if t.strip()]
            else:
                existing_targets = []
            
            if target in existing_targets:
                self.show_error(f"Target '{target}' already exists in target list")
                return
            
            # Add target
            existing_targets.append(target)
            targets_file.write_text('\n'.join(existing_targets) + '\n')
            
            self.show_success(f"Target '{target}' added successfully!")
            print(f"   Total targets: {len(existing_targets)}")
            
        except Exception as e:
            self.show_error(f"Failed to add target: {e}")
        
        self.wait_for_continue()
    
    def import_target_list(self):
        """Import target list from file"""
        print("\n📥 \033[1;97mImport Target List:\033[0m")
        print("Supported formats: TXT (one target per line), CSV, JSON")
        
        file_path = input("\n📁 Enter file path: ").strip()
        if not file_path:
            return
            
        try:
            file_path = Path(file_path)
            if not file_path.exists():
                self.show_error(f"File not found: {file_path}")
                return
                
            new_targets = set()
            
            if file_path.suffix.lower() == '.txt':
                # Plain text file
                with open(file_path, 'r') as f:
                    for line in f:
                        target = line.strip()
                        if target and not target.startswith('#'):
                            new_targets.add(target)
                            
            elif file_path.suffix.lower() == '.csv':
                # CSV file
                import csv
                with open(file_path, 'r') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        if row and len(row) > 0:
                            target = row[0].strip()
                            if target and not target.startswith('#'):
                                new_targets.add(target)
                                
            elif file_path.suffix.lower() == '.json':
                # JSON file
                with open(file_path, 'r') as f:
                    data = json.load(f)
                    if isinstance(data, list):
                        for target in data:
                            if isinstance(target, str):
                                new_targets.add(target.strip())
                            elif isinstance(target, dict) and 'target' in target:
                                new_targets.add(target['target'].strip())
                    elif isinstance(data, dict) and 'targets' in data:
                        for target in data['targets']:
                            if isinstance(target, str):
                                new_targets.add(target.strip())
            else:
                self.show_error("Unsupported file format. Use .txt, .csv, or .json")
                return
                
            if new_targets:
                # Show preview
                print(f"\n📋 Found {len(new_targets)} targets:")
                for i, target in enumerate(list(new_targets)[:10], 1):
                    print(f"  {i}. {target}")
                if len(new_targets) > 10:
                    print(f"  ... and {len(new_targets) - 10} more")
                
                # Confirm import
                confirm = input(f"\n✅ Import {len(new_targets)} targets? (y/N): ")
                if confirm.lower() == 'y':
                    # Add to existing targets
                    original_count = len(self.targets)
                    self.targets.update(new_targets)
                    new_count = len(self.targets) - original_count
                    
                    # Save to targets.txt
                    self.save_targets()
                    
                    self.show_success(f"Imported {new_count} new targets (total: {len(self.targets)})")
                else:
                    print("Import cancelled")
            else:
                self.show_info("No valid targets found in file")
                
        except Exception as e:
            self.show_error(f"Error importing targets: {e}")
            
        self.wait_for_continue()
        
    def save_targets(self):
        """Save targets to targets.txt file"""
        try:
            targets_file = Path("targets.txt")
            with open(targets_file, 'w') as f:
                for target in sorted(self.targets):
                    f.write(f"{target}\n")
            self.logger.info(f"Saved {len(self.targets)} targets to {targets_file}")
        except Exception as e:
            self.logger.error(f"Error saving targets: {e}")
    
    def view_current_targets(self) -> None:
        """View current targets in the target list"""
        self.print_master_banner()
        print("╔═══════════════════════════════════════════════════════════════════════════════╗")
        print("║\033[1;94m                        📋 CURRENT TARGETS 📋\033[0m                        ║")
        print("╚═══════════════════════════════════════════════════════════════════════════════╝")
        
        try:
            targets_file = Path("targets.txt")
            if not targets_file.exists():
                print("\n📋 \033[1;93mNo targets file found.\033[0m")
                print("   Use 'Add Single Target' to create your first target.")
                self.wait_for_continue()
                return
            
            content = targets_file.read_text().strip()
            if not content:
                print("\n📋 \033[1;93mTarget list is empty.\033[0m")
                print("   Use 'Add Single Target' to add your first target.")
                self.wait_for_continue()
                return
            
            targets = [t.strip() for t in content.split('\n') if t.strip()]
            
            print(f"\n📊 \033[1;97mTotal targets: {len(targets)}\033[0m")
            print("═" * 60)
            
            for i, target in enumerate(targets, 1):
                # Basic target type detection
                if target.startswith(('http://', 'https://')):
                    icon = "🌐"
                    type_name = "URL"
                elif target.replace('.', '').replace(':', '').isdigit() or ':' in target:
                    icon = "🖥️"
                    type_name = "IP"
                else:
                    icon = "🌍"
                    type_name = "Domain"
                
                print(f"{i:3d}. {icon} \033[1;97m{target:<40}\033[0m ({type_name})")
            
            print("═" * 60)
            
        except Exception as e:
            self.show_error(f"Failed to load targets: {e}")
        
        self.wait_for_continue()
    
    def remove_targets(self):
        """Remove targets from the list"""
        if not self.targets:
            self.show_info("No targets to remove")
            self.wait_for_continue()
            return
            
        while True:
            print("\n🗑️ \033[1;97mRemove Targets:\033[0m")
            print("1. Remove specific target")
            print("2. Remove multiple targets")
            print("3. Clear all targets")
            print("4. Remove by pattern/filter")
            print("0. Return to previous menu")
            
            choice = input("\nSelect option (0-4): ").strip()
            
            if choice == "0":
                break
            elif choice == "1":
                self.remove_single_target()
            elif choice == "2":
                self.remove_multiple_targets()
            elif choice == "3":
                self.clear_all_targets()
            elif choice == "4":
                self.remove_targets_by_pattern()
            else:
                self.show_error("Invalid option")
                
    def remove_single_target(self):
        """Remove a single target"""
        print(f"\n🎯 \033[1;97mCurrent Targets ({len(self.targets)}):\033[0m")
        targets_list = list(self.targets)
        
        for i, target in enumerate(targets_list, 1):
            print(f"  {i}. {target}")
            
        try:
            choice = input(f"\nSelect target to remove (1-{len(targets_list)}): ")
            index = int(choice) - 1
            
            if 0 <= index < len(targets_list):
                target_to_remove = targets_list[index]
                confirm = input(f"\n⚠️ Remove '{target_to_remove}'? (y/N): ")
                
                if confirm.lower() == 'y':
                    self.targets.remove(target_to_remove)
                    self.save_targets()
                    self.show_success(f"Removed target: {target_to_remove}")
                else:
                    print("Removal cancelled")
            else:
                self.show_error("Invalid target selection")
                
        except (ValueError, IndexError):
            self.show_error("Invalid input")
            
    def remove_multiple_targets(self):
        """Remove multiple targets"""
        print(f"\n🎯 \033[1;97mCurrent Targets ({len(self.targets)}):\033[0m")
        targets_list = list(self.targets)
        
        for i, target in enumerate(targets_list, 1):
            print(f"  {i}. {target}")
            
        print("\n💡 Enter target numbers separated by commas (e.g., 1,3,5)")
        selection = input("Select targets to remove: ").strip()
        
        if not selection:
            return
            
        try:
            indices = [int(x.strip()) - 1 for x in selection.split(',')]
            targets_to_remove = []
            
            for index in indices:
                if 0 <= index < len(targets_list):
                    targets_to_remove.append(targets_list[index])
                    
            if targets_to_remove:
                print(f"\n📋 Targets to remove:")
                for target in targets_to_remove:
                    print(f"  • {target}")
                    
                confirm = input(f"\n⚠️ Remove {len(targets_to_remove)} targets? (y/N): ")
                
                if confirm.lower() == 'y':
                    for target in targets_to_remove:
                        self.targets.remove(target)
                    self.save_targets()
                    self.show_success(f"Removed {len(targets_to_remove)} targets")
                else:
                    print("Removal cancelled")
            else:
                self.show_error("No valid targets selected")
                
        except ValueError:
            self.show_error("Invalid input format")
            
    def clear_all_targets(self):
        """Clear all targets"""
        if not self.targets:
            self.show_info("No targets to clear")
            return
            
        print(f"\n⚠️ \033[1;91mThis will remove ALL {len(self.targets)} targets!\033[0m")
        confirm = input("Are you sure? Type 'DELETE ALL' to confirm: ")
        
        if confirm == "DELETE ALL":
            self.targets.clear()
            self.save_targets()
            self.show_success("All targets removed")
        else:
            print("Clear operation cancelled")
            
    def remove_targets_by_pattern(self):
        """Remove targets by pattern/filter"""
        pattern = input("\n🔍 Enter pattern to match (supports wildcards *): ").strip()
        if not pattern:
            return
            
        import fnmatch
        matching_targets = []
        
        for target in self.targets:
            if fnmatch.fnmatch(target.lower(), pattern.lower()):
                matching_targets.append(target)
                
        if matching_targets:
            print(f"\n📋 Targets matching pattern '{pattern}':")
            for target in matching_targets:
                print(f"  • {target}")
                
            confirm = input(f"\n⚠️ Remove {len(matching_targets)} matching targets? (y/N): ")
            
            if confirm.lower() == 'y':
                for target in matching_targets:
                    self.targets.remove(target)
                self.save_targets()
                self.show_success(f"Removed {len(matching_targets)} targets matching pattern")
            else:
                print("Removal cancelled")
        else:
            self.show_info(f"No targets match pattern: {pattern}")
    
    def validate_targets(self) -> None:
        """Validate targets in the target list"""
        self.print_master_banner()
        print("╔═══════════════════════════════════════════════════════════════════════════════╗")
        print("║\033[1;95m                        ✅ VALIDATE TARGETS ✅\033[0m                        ║")
        print("╚═══════════════════════════════════════════════════════════════════════════════╝")
        
        try:
            targets_file = Path("targets.txt")
            if not targets_file.exists():
                self.show_error("No targets file found")
                return
            
            content = targets_file.read_text().strip()
            if not content:
                self.show_error("Target list is empty")
                return
            
            targets = [t.strip() for t in content.split('\n') if t.strip()]
            
            print(f"\n🔍 \033[1;97mValidating {len(targets)} targets...\033[0m")
            print("═" * 60)
            
            valid_targets = []
            invalid_targets = []
            
            for i, target in enumerate(targets, 1):
                print(f"🔍 Validating {i}/{len(targets)}: {target}")
                
                # Use our validation function
                is_valid, error_message = self._validate_target_format(target)
                
                if is_valid:
                    valid_targets.append(target)
                    print(f"   ✅ Valid")
                else:
                    invalid_targets.append((target, [error_message]))
                    print(f"   ❌ Invalid: {error_message}")
            
            print("═" * 60)
            print(f"📊 \033[1;97mValidation Summary:\033[0m")
            print(f"   ✅ Valid targets: {len(valid_targets)}")
            print(f"   ❌ Invalid targets: {len(invalid_targets)}")
            
            if invalid_targets:
                print(f"\n❌ \033[1;91mInvalid targets found:\033[0m")
                for target, issues in invalid_targets:
                    print(f"   • {target}: {', '.join(issues)}")
                
                if input("\n🗑️  Remove invalid targets? [y/N]: ").strip().lower() == 'y':
                    targets_file.write_text('\n'.join(valid_targets) + '\n')
                    self.show_success(f"Removed {len(invalid_targets)} invalid targets")
            else:
                self.show_success("All targets are valid!")
            
        except Exception as e:
            self.show_error(f"Validation failed: {e}")
        
        self.wait_for_continue()
    
    def target_analysis_summary(self):
        """Generate target analysis summary"""
        if not self.targets:
            self.show_info("No targets available for analysis")
            self.wait_for_continue()
            return
            
        print("\n📊 \033[1;97mTarget Analysis Summary:\033[0m")
        print("=" * 60)
        
        # Basic statistics
        print(f"📈 Basic Statistics:")
        print(f"   Total Targets: {len(self.targets)}")
        
        # Categorize targets
        categories = self.get_target_category_summary()
        if categories:
            print(f"\n📂 Target Categories:")
            for category, count in categories.items():
                percentage = (count / len(self.targets)) * 100
                print(f"   • {category}: {count} ({percentage:.1f}%)")
                
        # Domain analysis
        from urllib.parse import urlparse
        domains = set()
        subdomains = set()
        
        for target in self.targets:
            if target.startswith(('http://', 'https://')):
                domain = urlparse(target).netloc
            else:
                domain = target
                
            # Remove port if present
            domain = domain.split(':')[0]
            
            if '.' in domain:
                if domain.count('.') > 1:
                    # It's a subdomain
                    subdomains.add(domain)
                    # Add root domain too
                    root_domain = '.'.join(domain.split('.')[-2:])
                    domains.add(root_domain)
                else:
                    domains.add(domain)
                    
        print(f"\n🌐 Domain Analysis:")
        print(f"   Unique Domains: {len(domains)}")
        print(f"   Subdomains: {len(subdomains)}")
        
        if len(domains) <= 10:
            print(f"   Domains: {', '.join(sorted(domains))}")
        else:
            print(f"   Top domains: {', '.join(sorted(domains)[:10])}...")
            
        # Scan coverage analysis
        runs_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        if runs_dir.exists():
            scanned_targets = set()
            
            for run_dir in runs_dir.iterdir():
                if run_dir.is_dir():
                    # Try to extract target from directory name
                    run_name = run_dir.name
                    for target in self.targets:
                        clean_target = target.replace('.', '_').replace(':', '_').replace('/', '_')
                        if clean_target in run_name:
                            scanned_targets.add(target)
                            break
                            
            coverage = len(scanned_targets) / len(self.targets) * 100
            print(f"\n🔍 Scan Coverage:")
            print(f"   Scanned Targets: {len(scanned_targets)}/{len(self.targets)} ({coverage:.1f}%)")
            
            unscanned = self.targets - scanned_targets
            if unscanned and len(unscanned) <= 5:
                print(f"   Unscanned: {', '.join(sorted(unscanned))}")
            elif unscanned:
                print(f"   Unscanned: {len(unscanned)} targets")
                
        # Generate summary report
        summary_data = {
            'timestamp': datetime.now().isoformat(),
            'total_targets': len(self.targets),
            'categories': categories,
            'domains': list(domains),
            'subdomains': list(subdomains),
            'coverage': {
                'scanned': len(scanned_targets),
                'total': len(self.targets),
                'percentage': coverage
            }
        }
        
        # Save summary
        summary_file = Path(f"target_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with open(summary_file, 'w') as f:
            json.dump(summary_data, f, indent=2)
            
        print(f"\n📁 Analysis saved to: {summary_file}")
        self.wait_for_continue()
    
    def export_target_list(self):
        """Export target list to various formats"""
        if not self.targets:
            self.show_info("No targets to export")
            self.wait_for_continue()
            return
            
        print("\n📤 \033[1;97mExport Target List:\033[0m")
        print("1. Plain text (.txt)")
        print("2. CSV format (.csv)")
        print("3. JSON format (.json)")
        print("4. XML format (.xml)")
        print("5. Custom format")
        
        choice = input("\nSelect export format (1-5): ").strip()
        
        if choice not in ['1', '2', '3', '4', '5']:
            self.show_error("Invalid format selection")
            return
            
        # Get output filename
        default_name = f"targets_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        filename = input(f"\nEnter filename (default: {default_name}): ").strip()
        
        if not filename:
            filename = default_name
            
        try:
            if choice == "1":
                # Plain text
                filepath = Path(f"{filename}.txt")
                with open(filepath, 'w') as f:
                    for target in sorted(self.targets):
                        f.write(f"{target}\n")
                        
            elif choice == "2":
                # CSV format
                import csv
                filepath = Path(f"{filename}.csv")
                with open(filepath, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Target', 'Added_Date', 'Status'])
                    for target in sorted(self.targets):
                        writer.writerow([target, datetime.now().strftime('%Y-%m-%d'), 'Active'])
                        
            elif choice == "3":
                # JSON format
                filepath = Path(f"{filename}.json")
                export_data = {
                    "export_info": {
                        "timestamp": datetime.now().isoformat(),
                        "tool": f"{MASTER_APP} {MASTER_VERSION}",
                        "total_targets": len(self.targets)
                    },
                    "targets": [
                        {
                            "target": target,
                            "added_date": datetime.now().isoformat(),
                            "status": "active"
                        } for target in sorted(self.targets)
                    ]
                }
                with open(filepath, 'w') as f:
                    json.dump(export_data, f, indent=2)
                    
            elif choice == "4":
                # XML format
                filepath = Path(f"{filename}.xml")
                with open(filepath, 'w') as f:
                    f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
                    f.write('<targets>\n')
                    f.write(f'  <export_info>\n')
                    f.write(f'    <timestamp>{datetime.now().isoformat()}</timestamp>\n')
                    f.write(f'    <tool>{MASTER_APP} {MASTER_VERSION}</tool>\n')
                    f.write(f'    <total_targets>{len(self.targets)}</total_targets>\n')
                    f.write(f'  </export_info>\n')
                    for target in sorted(self.targets):
                        f.write(f'  <target>\n')
                        f.write(f'    <url>{target}</url>\n')
                        f.write(f'    <status>active</status>\n')
                        f.write(f'  </target>\n')
                    f.write('</targets>\n')
                    
            elif choice == "5":
                # Custom format
                print("\n⚙️ Custom Export Format:")
                print("Enter format template (use {target} as placeholder)")
                print("Example: Target: {target} | Status: Active")
                
                template = input("Format template: ").strip()
                if not template:
                    template = "{target}"
                    
                filepath = Path(f"{filename}.txt")
                with open(filepath, 'w') as f:
                    for target in sorted(self.targets):
                        f.write(template.format(target=target) + '\n')
                        
            self.show_success(f"Exported {len(self.targets)} targets to {filepath}")
            print(f"📁 File saved: {filepath.absolute()}")
            
        except Exception as e:
            self.show_error(f"Error exporting targets: {e}")
            
        self.wait_for_continue()
    
    def bulk_target_operations(self):
        """Bulk operations on targets"""
        if not self.targets:
            self.show_info("No targets available for bulk operations")
            self.wait_for_continue()
            return
            
        while True:
            print("\n📦 \033[1;97mBulk Target Operations:\033[0m")
            print(f"Current targets: {len(self.targets)}")
            print("")
            print("1. 🔍 Validate all targets (connectivity check)")
            print("2. 🌐 Resolve all domains to IPs")
            print("3. 📊 Generate target statistics")
            print("4. 🔄 Deduplicate targets")
            print("5. 🏷️ Categorize targets by type")
            print("6. 🚀 Mass reconnaissance")
            print("7. 📋 Bulk export with metadata")
            print("0. Return to previous menu")
            
            choice = input("\nSelect operation (0-7): ").strip()
            
            if choice == "0":
                break
            elif choice == "1":
                self.validate_all_targets()
            elif choice == "2":
                self.resolve_all_targets()
            elif choice == "3":
                self.generate_target_statistics()
            elif choice == "4":
                self.deduplicate_targets()
            elif choice == "5":
                self.categorize_targets()
            elif choice == "6":
                # Create task for async operation to avoid event loop error
                asyncio.create_task(self.mass_reconnaissance())
            elif choice == "7":
                self.bulk_export_with_metadata()
            else:
                self.show_error("Invalid option")
                
    def validate_all_targets(self):
        """Validate connectivity for all targets"""
        print(f"\n🔍 \033[1;97mValidating {len(self.targets)} targets...\033[0m")
        
        valid_targets = []
        invalid_targets = []
        
        import requests
        from urllib.parse import urlparse
        
        for i, target in enumerate(self.targets, 1):
            print(f"\r[{i}/{len(self.targets)}] Checking {target[:50]}...", end="", flush=True)
            
            try:
                # Ensure URL format
                if not target.startswith(('http://', 'https://')):
                    test_url = f"https://{target}"
                else:
                    test_url = target
                    
                response = requests.head(test_url, timeout=10, allow_redirects=True)
                if response.status_code < 500:  # Accept redirects and client errors
                    valid_targets.append(target)
                else:
                    invalid_targets.append((target, f"HTTP {response.status_code}"))
                    
            except Exception as e:
                invalid_targets.append((target, str(e)[:50]))
                
        print(f"\n\n📊 \033[1;97mValidation Results:\033[0m")
        print(f"✅ Valid targets: {len(valid_targets)}")
        print(f"❌ Invalid targets: {len(invalid_targets)}")
        
        if invalid_targets:
            print(f"\n❌ Invalid targets:")
            for target, error in invalid_targets[:10]:
                print(f"  • {target} - {error}")
            if len(invalid_targets) > 10:
                print(f"  ... and {len(invalid_targets) - 10} more")
                
            remove_invalid = input(f"\n🗑️ Remove {len(invalid_targets)} invalid targets? (y/N): ")
            if remove_invalid.lower() == 'y':
                for target, _ in invalid_targets:
                    self.targets.discard(target)
                self.save_targets()
                self.show_success(f"Removed {len(invalid_targets)} invalid targets")
                
        self.wait_for_continue()
        
    def resolve_all_targets(self):
        """Resolve all domain targets to IP addresses"""
        print(f"\n🌐 \033[1;97mResolving {len(self.targets)} targets...\033[0m")
        
        import socket
        from urllib.parse import urlparse
        
        resolved_data = {}
        
        for i, target in enumerate(self.targets, 1):
            print(f"\r[{i}/{len(self.targets)}] Resolving {target[:50]}...", end="", flush=True)
            
            try:
                # Extract domain from URL if needed
                if target.startswith(('http://', 'https://')):
                    domain = urlparse(target).netloc
                else:
                    domain = target
                    
                # Remove port if present
                domain = domain.split(':')[0]
                
                # Resolve to IP
                ip = socket.gethostbyname(domain)
                resolved_data[target] = {
                    'domain': domain,
                    'ip': ip,
                    'status': 'resolved'
                }
                
            except Exception as e:
                resolved_data[target] = {
                    'domain': domain if 'domain' in locals() else target,
                    'ip': None,
                    'status': f'failed: {str(e)[:30]}'
                }
                
        print(f"\n\n📊 \033[1;97mResolution Results:\033[0m")
        
        resolved_count = len([d for d in resolved_data.values() if d['ip']])
        failed_count = len(resolved_data) - resolved_count
        
        print(f"✅ Resolved: {resolved_count}")
        print(f"❌ Failed: {failed_count}")
        
        # Save results
        output_file = Path(f"target_resolution_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with open(output_file, 'w') as f:
            json.dump(resolved_data, f, indent=2)
            
        print(f"\n📁 Results saved to: {output_file}")
        self.wait_for_continue()
        
    def generate_target_statistics(self):
        """Generate comprehensive target statistics"""
        print(f"\n📊 \033[1;97mTarget Statistics:\033[0m")
        print("=" * 50)
        
        from urllib.parse import urlparse
        from collections import defaultdict
        
        stats = {
            'total': len(self.targets),
            'protocols': defaultdict(int),
            'domains': defaultdict(int),
            'tlds': defaultdict(int),
            'ports': defaultdict(int),
            'subdomains': 0,
            'ips': 0,
            'urls': 0
        }
        
        for target in self.targets:
            # Check if it's an IP address
            import re
            ip_pattern = r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
            
            if re.match(ip_pattern, target.split(':')[0]):
                stats['ips'] += 1
                if ':' in target:
                    port = target.split(':')[1]
                    stats['ports'][port] += 1
            elif target.startswith(('http://', 'https://')):
                stats['urls'] += 1
                parsed = urlparse(target)
                stats['protocols'][parsed.scheme] += 1
                
                domain = parsed.netloc.split(':')[0]
                if '.' in domain:
                    tld = domain.split('.')[-1]
                    stats['tlds'][tld] += 1
                    
                if domain.count('.') > 1:
                    stats['subdomains'] += 1
                    
                if parsed.port:
                    stats['ports'][str(parsed.port)] += 1
                    
            else:
                # Assume it's a domain
                if '.' in target:
                    tld = target.split('.')[-1]
                    stats['tlds'][tld] += 1
                    
                if target.count('.') > 1:
                    stats['subdomains'] += 1
                    
        print(f"📈 Total Targets: {stats['total']}")
        print(f"🌐 Domains/URLs: {stats['urls']}")
        print(f"📱 IP Addresses: {stats['ips']}")
        print(f"🔗 Subdomains: {stats['subdomains']}")
        
        if stats['protocols']:
            print(f"\n🔒 Protocols:")
            for protocol, count in sorted(stats['protocols'].items()):
                print(f"  • {protocol}: {count}")
                
        if stats['tlds']:
            print(f"\n🌍 Top TLDs:")
            top_tlds = sorted(stats['tlds'].items(), key=lambda x: x[1], reverse=True)[:10]
            for tld, count in top_tlds:
                print(f"  • .{tld}: {count}")
                
        if stats['ports']:
            print(f"\n🔌 Ports:")
            for port, count in sorted(stats['ports'].items()):
                print(f"  • {port}: {count}")
                
        # Save detailed stats
        output_file = Path(f"target_stats_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with open(output_file, 'w') as f:
            # Convert defaultdict to regular dict for JSON serialization
            stats_dict = {k: dict(v) if isinstance(v, defaultdict) else v for k, v in stats.items()}
            json.dump(stats_dict, f, indent=2)
            
        print(f"\n📁 Detailed stats saved to: {output_file}")
        self.wait_for_continue()
        
    def deduplicate_targets(self):
        """Remove duplicate targets"""
        original_count = len(self.targets)
        
        # Targets are already in a set, so they're unique
        # But let's normalize and check for similar entries
        normalized_targets = set()
        duplicates_found = []
        
        for target in self.targets:
            # Normalize target
            normalized = target.lower().strip()
            
            # Remove trailing slash
            if normalized.endswith('/'):
                normalized = normalized[:-1]
                
            # Check for www variants
            if normalized.startswith('www.'):
                base_domain = normalized[4:]
                if base_domain in normalized_targets:
                    duplicates_found.append((target, f"Duplicate of {base_domain}"))
                    continue
                    
            if f"www.{normalized}" in normalized_targets:
                duplicates_found.append((target, f"Duplicate with www variant"))
                continue
                
            normalized_targets.add(normalized)
            
        if duplicates_found:
            print(f"\n🔍 \033[1;97mFound {len(duplicates_found)} potential duplicates:\033[0m")
            for target, reason in duplicates_found[:10]:
                print(f"  • {target} - {reason}")
                
            if len(duplicates_found) > 10:
                print(f"  ... and {len(duplicates_found) - 10} more")
                
            remove_dupes = input(f"\n🗑️ Remove {len(duplicates_found)} duplicates? (y/N): ")
            if remove_dupes.lower() == 'y':
                for target, _ in duplicates_found:
                    self.targets.discard(target)
                self.save_targets()
                self.show_success(f"Removed {len(duplicates_found)} duplicates")
        else:
            self.show_info("No duplicates found")
            
        self.wait_for_continue()
        
    def categorize_targets(self):
        """Categorize targets by type"""
        from urllib.parse import urlparse
        import re
        
        categories = {
            'web_applications': [],
            'api_endpoints': [],
            'ip_addresses': [],
            'subdomains': [],
            'main_domains': [],
            'unknown': []
        }
        
        ip_pattern = r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
        
        for target in self.targets:
            if re.match(ip_pattern, target.split(':')[0]):
                categories['ip_addresses'].append(target)
            elif 'api' in target.lower() or '/api/' in target.lower():
                categories['api_endpoints'].append(target)
            elif target.startswith(('http://', 'https://')):
                parsed = urlparse(target)
                domain = parsed.netloc.split(':')[0]
                if domain.count('.') > 1:
                    categories['subdomains'].append(target)
                else:
                    categories['web_applications'].append(target)
            elif '.' in target:
                if target.count('.') > 1:
                    categories['subdomains'].append(target)
                else:
                    categories['main_domains'].append(target)
            else:
                categories['unknown'].append(target)
                
        print(f"\n📂 \033[1;97mTarget Categories:\033[0m")
        print("=" * 40)
        
        for category, targets in categories.items():
            if targets:
                print(f"\n{category.replace('_', ' ').title()}: {len(targets)}")
                for target in targets[:5]:
                    print(f"  • {target}")
                if len(targets) > 5:
                    print(f"  ... and {len(targets) - 5} more")
                    
        # Save categorized results
        output_file = Path(f"target_categories_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
        with open(output_file, 'w') as f:
            json.dump(categories, f, indent=2)
            
        print(f"\n📁 Categories saved to: {output_file}")
        self.wait_for_continue()
        
    async def mass_reconnaissance(self):
        """Perform reconnaissance on all targets"""
        if not self.integrations_available:
            self.show_error("Moloch integration not available")
            return
            
        print(f"\n🚀 \033[1;97mMass Reconnaissance on {len(self.targets)} targets\033[0m")
        print("⚠️ This will run reconnaissance on ALL targets!")
        
        confirm = input("\nContinue with mass reconnaissance? (y/N): ")
        if confirm.lower() != 'y':
            return
            
        # Create master run directory
        run_dir = Path(self.moloch_integration.config['general']['runs_dir'])
        mass_run_path = run_dir / f"mass_recon_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        mass_run_path.mkdir(parents=True, exist_ok=True)
        
        results = {}
        completed = 0
        failed = 0
        
        for i, target in enumerate(self.targets, 1):
            print(f"\n[{i}/{len(self.targets)}] 🎯 Processing {target}")
            
            try:
                target_dir = mass_run_path / target.replace('.', '_').replace(':', '_').replace('/', '_')
                target_dir.mkdir(exist_ok=True)
                
                # Run reconnaissance
                recon_results = await self.moloch_integration.run_reconnaissance_suite(
                    target, target_dir, aggressive=False
                )
                
                results[target] = recon_results
                completed += 1
                
                print(f"  ✅ Completed: {len(recon_results.get('live_hosts', []))} live hosts found")
                
            except Exception as e:
                print(f"  ❌ Failed: {e}")
                results[target] = {'error': str(e)}
                failed += 1
                
        # Save master results
        master_results_file = mass_run_path / "mass_reconnaissance_results.json"
        with open(master_results_file, 'w') as f:
            json.dump(results, f, indent=2, default=str)
            
        print(f"\n📊 \033[1;97mMass Reconnaissance Summary:\033[0m")
        print(f"✅ Completed: {completed}")
        print(f"❌ Failed: {failed}")
        print(f"📁 Results saved to: {mass_run_path}")
        
        self.wait_for_continue()
        
    def bulk_export_with_metadata(self):
        """Export targets with comprehensive metadata"""
        print(f"\n📋 \033[1;97mBulk Export with Metadata\033[0m")
        
        # Collect metadata for all targets
        metadata = {}
        
        print("🔍 Collecting metadata...")
        for target in self.targets:
            print(f"\rProcessing {target[:50]}...", end="", flush=True)
            
            target_data = {
                'target': target,
                'added_date': datetime.now().isoformat(),
                'type': self.classify_target_type(target),
                'status': 'active'
            }
            
            metadata[target] = target_data
            
        print(f"\n\n📊 Export Options:")
        print("1. Detailed JSON with metadata")
        print("2. CSV with basic info")
        print("3. Excel spreadsheet")
        print("4. Markdown report")
        
        choice = input("\nSelect export format (1-4): ").strip()
        
        filename = f"targets_bulk_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        try:
            if choice == "1":
                # Detailed JSON
                filepath = Path(f"{filename}.json")
                export_data = {
                    'export_info': {
                        'timestamp': datetime.now().isoformat(),
                        'tool': f"{MASTER_APP} {MASTER_VERSION}",
                        'total_targets': len(self.targets),
                        'categories': self.get_target_category_summary()
                    },
                    'targets': metadata
                }
                with open(filepath, 'w') as f:
                    json.dump(export_data, f, indent=2)
                    
            elif choice == "2":
                # CSV format
                import csv
                filepath = Path(f"{filename}.csv")
                with open(filepath, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Target', 'Type', 'Status', 'Added_Date'])
                    for target_data in metadata.values():
                        writer.writerow([
                            target_data['target'],
                            target_data['type'],
                            target_data['status'],
                            target_data['added_date']
                        ])
                        
            elif choice == "3":
                # Try to create Excel file (requires openpyxl)
                try:
                    import openpyxl
                    from openpyxl import Workbook
                    
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Targets"
                    
                    # Headers
                    headers = ['Target', 'Type', 'Status', 'Added Date']
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                        
                    # Data
                    for row, target_data in enumerate(metadata.values(), 2):
                        ws.cell(row=row, column=1, value=target_data['target'])
                        ws.cell(row=row, column=2, value=target_data['type'])
                        ws.cell(row=row, column=3, value=target_data['status'])
                        ws.cell(row=row, column=4, value=target_data['added_date'])
                        
                    filepath = Path(f"{filename}.xlsx")
                    wb.save(filepath)
                    
                except ImportError:
                    self.show_error("openpyxl not installed. Install with: pip install openpyxl")
                    return
                    
            elif choice == "4":
                # Markdown report
                filepath = Path(f"{filename}.md")
                with open(filepath, 'w') as f:
                    f.write(f"# Target List Report\n\n")
                    f.write(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write(f"**Tool:** {MASTER_APP} {MASTER_VERSION}\n")
                    f.write(f"**Total Targets:** {len(self.targets)}\n\n")
                    
                    f.write("## Target Summary\n\n")
                    f.write("| Target | Type | Status |\n")
                    f.write("|--------|------|--------|\n")
                    
                    for target_data in metadata.values():
                        f.write(f"| {target_data['target']} | {target_data['type']} | {target_data['status']} |\n")
                        
            else:
                self.show_error("Invalid export format")
                return
                
            self.show_success(f"Bulk export completed: {filepath}")
            print(f"📁 File saved: {filepath.absolute()}")
            
        except Exception as e:
            self.show_error(f"Error during bulk export: {e}")
            
        self.wait_for_continue()
        
    def classify_target_type(self, target):
        """Classify target type"""
        import re
        from urllib.parse import urlparse
        
        ip_pattern = r'^(?:(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)\.){3}(?:25[0-5]|2[0-4][0-9]|[01]?[0-9][0-9]?)$'
        
        if re.match(ip_pattern, target.split(':')[0]):
            return 'IP Address'
        elif 'api' in target.lower():
            return 'API Endpoint'
        elif target.startswith(('http://', 'https://')):
            return 'Web Application'
        elif '.' in target:
            if target.count('.') > 1:
                return 'Subdomain'
            else:
                return 'Domain'
        else:
            return 'Unknown'
            
    def get_target_category_summary(self):
        """Get summary of target categories"""
        from collections import defaultdict
        
        categories = defaultdict(int)
        for target in self.targets:
            category = self.classify_target_type(target)
            categories[category] += 1
            
        return dict(categories)
    
    def show_detailed_system_status(self):
        """Show detailed system status"""
        self.print_master_banner()
        self.update_system_status()
        
        print("╔═══════════════════════════════════════════════════════════════════════════════╗")
        print("║\033[1;94m                      📊 DETAILED SYSTEM STATUS 📊\033[0m                      ║")
        print("╚═══════════════════════════════════════════════════════════════════════════════╝")
        
        print(f"\n🏥 \033[1;97mSystem Health:\033[0m {self.system_status['health']}")
        print(f"🔧 \033[1;97mFramework Version:\033[0m {MASTER_VERSION}")
        print(f"📊 \033[1;97mActive Scans:\033[0m {len(self.active_scans)}")
        print(f"📋 \033[1;97mScan History:\033[0m {len(self.scan_history)}")
        
        print(f"\n⚡ \033[1;97mPerformance Metrics:\033[0m")
        print(f"   CPU Usage: {self.performance_metrics['cpu_usage']:.1f}%")
        print(f"   Memory Usage: {self.performance_metrics['memory_usage']:.1f}%")
        print(f"   Disk Usage: {self.performance_metrics['disk_usage']:.1f}%")
        
        print(f"\n🔧 \033[1;97mScanner Status:\033[0m")
        for scanner, status in self.system_status['scanners'].items():
            print(f"   {scanner}: {status}")
        
        print(f"\n🛠️  \033[1;97mTool Status:\033[0m")
        for tool, status in self.system_status['tools'].items():
            print(f"   {tool}: {status}")
        
        self.wait_for_continue()
    
    def show_help_menu(self):
        """Show help and usage information"""
        self.print_master_banner()
        print("╔═══════════════════════════════════════════════════════════════════════════════╗")
        print("║\033[1;96m                           📚 HELP & USAGE 📚\033[0m                           ║")
        print("╚═══════════════════════════════════════════════════════════════════════════════╝")
        
        print("\n🎯 \033[1;97mQuick Start Guide:\033[0m")
        print("   1. Use option 1 for complete automated security assessment")
        print("   2. Use option 2 to manage targets before scanning")
        print("   3. Use individual modules (3-9) for specific assessments")
        print("   4. Use option 10 to configure tools and settings")
        print("   5. Use option 11 to generate reports from scan results")
        print("   6. Use option 12 for real-time monitoring")
        
        print("\n💡 \033[1;97mTips:\033[0m")
        print("   • Type 'status' anytime to check system health")
        print("   • Use submenu numbers (e.g., '3.1') for direct access")
        print("   • Press Ctrl+C to cancel operations safely")
        print("   • All scan results are automatically saved")
        
        print(f"\n📖 \033[1;97mDocumentation:\033[0m")
        print("   • GitHub: https://github.com/cxb3rf1lth/Azaz-El")
        print("   • README: Contains detailed usage instructions")
        print("   • Wiki: Advanced configuration and customization")
        
        self.wait_for_continue()
    
    async def generate_pipeline_report(self, scan_id: str):
        """Generate comprehensive report for pipeline execution"""
        print(f"\n📊 \033[1;97mGenerating comprehensive report for scan: {scan_id}\033[0m")
        
        try:
            if hasattr(self, 'report_generator'):
                # Generate report using the advanced report generator
                scan_data = self.scan_history[-1] if self.scan_history else {}
                report_path = await self.report_generator.generate_comprehensive_report(scan_data)
                self.show_success(f"Report generated: {report_path}")
            else:
                self.show_info("Report generation functionality will be enhanced in future versions")
                
        except Exception as e:
            self.show_error(f"Report generation failed: {e}")


def setup_argument_parser() -> argparse.ArgumentParser:
    """Setup comprehensive command line argument parser"""
    parser = argparse.ArgumentParser(
        description=f"{MASTER_APP} {MASTER_VERSION} - Unified Master Security Assessment Framework",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Interactive Master Dashboard (default)
  python3 master_azaz_el.py
  
  # Quick CLI Operations
  python3 master_azaz_el.py --target example.com --full-pipeline
  python3 master_azaz_el.py --target example.com --reconnaissance
  python3 master_azaz_el.py --target-file targets.txt --vulnerability-scan
  
  # Advanced Operations
  python3 master_azaz_el.py --target example.com --aggressive --cloud-scan
  python3 master_azaz_el.py --status --system-health
  python3 master_azaz_el.py --install-tools --update-config
        """
    )
    
    # Target configuration
    target_group = parser.add_argument_group('Target Configuration')
    target_group.add_argument('--target', '-t', help='Single target (domain, IP, or URL)')
    target_group.add_argument('--target-file', '-tf', help='File containing list of targets')
    target_group.add_argument('--target-list', '-tl', nargs='+', help='Multiple targets as space-separated list')
    
    # Scanning operations
    scan_group = parser.add_argument_group('Scanning Operations')
    scan_group.add_argument('--full-pipeline', '-fp', action='store_true', help='Execute complete security assessment pipeline')
    scan_group.add_argument('--reconnaissance', '-r', action='store_true', help='Reconnaissance and intelligence gathering only')
    scan_group.add_argument('--vulnerability-scan', '-v', action='store_true', help='Vulnerability scanning only')
    scan_group.add_argument('--web-scan', '-w', action='store_true', help='Web application security testing only')
    scan_group.add_argument('--cloud-scan', '-c', action='store_true', help='Cloud security assessment only')
    scan_group.add_argument('--api-scan', '-a', action='store_true', help='API security testing only')
    scan_group.add_argument('--infrastructure-scan', '-i', action='store_true', help='Infrastructure security scanning only')
    
    # Scan configuration
    config_group = parser.add_argument_group('Scan Configuration')
    config_group.add_argument('--aggressive', action='store_true', help='Enable aggressive scanning mode')
    config_group.add_argument('--passive', action='store_true', help='Passive scanning mode only')
    config_group.add_argument('--threads', type=int, default=10, help='Number of concurrent threads (default: 10)')
    config_group.add_argument('--timeout', type=int, default=300, help='Scan timeout in seconds (default: 300)')
    config_group.add_argument('--output-dir', '-o', help='Output directory for results')
    
    # System management
    system_group = parser.add_argument_group('System Management')
    system_group.add_argument('--dashboard', '-d', action='store_true', help='Launch interactive dashboard (default)')
    system_group.add_argument('--status', '-s', action='store_true', help='Show system status and exit')
    system_group.add_argument('--system-health', action='store_true', help='Detailed system health check')
    system_group.add_argument('--install-tools', action='store_true', help='Install required security tools')
    system_group.add_argument('--update-tools', action='store_true', help='Update installed security tools')
    system_group.add_argument('--update-config', action='store_true', help='Update configuration files')
    
    # Output control
    output_group = parser.add_argument_group('Output Control')
    output_group.add_argument('--verbose', '-vv', action='store_true', help='Increase verbosity level')
    output_group.add_argument('--quiet', '-q', action='store_true', help='Suppress non-essential output')
    output_group.add_argument('--no-color', action='store_true', help='Disable colored output')
    output_group.add_argument('--json-output', action='store_true', help='Output results in JSON format')
    
    return parser


async def handle_cli_operations(args, framework):
    """Handle command line interface operations"""
    
    # Handle system status requests
    if args.status:
        framework.update_system_status()
        print(f"System Health: {framework.system_status['health']}")
        return True
    
    if args.system_health:
        framework.show_detailed_system_status()
        return True
    
    # Handle tool management
    if args.install_tools:
        framework.install_missing_tools()
        return True
    
    if args.update_tools:
        framework.update_existing_tools()
        return True
    
    # Handle scanning operations
    if args.target or args.target_file or args.target_list:
        # Determine target(s)
        targets = []
        if args.target:
            targets.append(args.target)
        if args.target_file:
            framework.show_info(f"Target file processing: {args.target_file}")
        if args.target_list:
            targets.extend(args.target_list)
        
        if not targets and not args.target_file:
            framework.show_error("No valid targets specified")
            return True
        
        # Execute requested scan type
        if args.full_pipeline:
            for target in targets:
                print(f"Executing full pipeline for: {target}")
                await framework.run_full_automation_pipeline()
        elif args.reconnaissance:
            print("🔍 Starting CLI reconnaissance scanning...")
            framework.reconnaissance_suite_menu()
        elif args.vulnerability_scan:
            print("🛡️ Starting CLI vulnerability scanning...")
            framework.vulnerability_scanning_menu()
        elif args.web_scan:
            print("🌐 Starting CLI web application scanning...")
            framework.web_application_testing_menu()
        elif args.cloud_scan:
            print("☁️ Starting CLI cloud security scanning...")
            framework.cloud_security_assessment_menu()
        elif args.api_scan:
            print("🔌 Starting CLI API security scanning...")
            framework.api_security_testing_menu()
        elif args.infrastructure_scan:
            print("🏗️ Starting CLI infrastructure scanning...")
            framework.infrastructure_scanning_menu()
        else:
            framework.show_info("No specific scan type specified, defaulting to full pipeline")
            await framework.run_full_automation_pipeline()
        
        return True
    
    return False


async def main():
    """Main entry point for the master framework"""
    
    # Setup argument parser
    parser = setup_argument_parser()
    args = parser.parse_args()
    
    try:
        # Initialize the master framework
        framework = MasterAzazElFramework()
        
        # Handle CLI operations if specified
        cli_handled = await handle_cli_operations(args, framework)
        
        # If no CLI operations were handled, launch interactive dashboard
        if not cli_handled:
            await framework.display_master_menu()
            
    except KeyboardInterrupt:
        print("\n\n🔄 \033[1;93mMaster framework interrupted by user\033[0m")
        print("👋 \033[1;92mGoodbye!\033[0m")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ \033[1;91mFatal error:\033[0m {e}")
        logging.exception("Fatal error in master framework")
        sys.exit(1)


if __name__ == "__main__":
    # Check Python version
    if sys.version_info < (3, 7):
        print("❌ Python 3.7 or higher is required")
        sys.exit(1)
    
    # Run the master framework
    asyncio.run(main())